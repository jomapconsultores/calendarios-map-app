# ------------------------------------------------------------
# Desarrollado por Marco Antonio Posligua San Martín
# ------------------------------------------------------------
"""Módulo DIRECTORIO: la base de datos de personas y empresas.

Tres exigencias del módulo que están resueltas aquí y no en el navegador,
porque el navegador se puede saltar:

  1) NO SE REPITEN REGISTROS. El número de documento es único. Se comprueba
     contra la base y, en una importación, también contra las demás filas del
     mismo archivo (dos filas repetidas dentro del mismo Excel son un duplicado
     igual de real que uno ya guardado). Cuando hay repetido, el sistema AVISA y
     NO INSERTA: no sobreescribe en silencio.

  2) TODA MODIFICACIÓN QUEDA REGISTRADA. Para guardar un cambio hay que declarar
     el motivo. Se anota quién, cuándo, por qué, y campo por campo el valor
     anterior y el nuevo. Sin motivo no hay cambio.

  3) SI ES RUC, SE VERIFICA CON EL SRI y se traen los datos que publica,
     incluidas todas las actividades económicas.
"""
import io
import json
from datetime import datetime, timezone

from flask import request, jsonify, render_template, redirect, flash, send_file
from flask_login import login_required, current_user

from . import sri as sri_mod
from . import documentos as docs_mod
from . import ia as ia_mod

# Columnas que el usuario puede escribir, con su nombre legible. El nombre
# legible es el que aparece en la bitácora: "Celular", no "mobile".
CAMPOS_EDITABLES = {
    'sector_id':     'Sector',
    'doc_type':      'Tipo de documento',
    'doc_number':    'Número de documento',
    'first_name':    'Nombres',
    'last_name':     'Apellidos',
    'business_name': 'Razón social',
    'trade_name':    'Nombre comercial',
    'mobile':        'Celular',
    'landline':      'Convencional',
    'email':         'Correo electrónico',
    'website':       'Dirección web',
    'socials':       'Redes sociales',
    'work_address':  'Dirección de trabajo',
    'home_address':  'Dirección domiciliaria',
    'city':          'Ciudad',
    'province':      'Provincia',
    'notes':         'Observaciones',
    'tags':          'Etiquetas',
    'active':        'Activo',
}

# Columnas que llena el SRI. El usuario no las escribe a mano: se refrescan
# consultando el servicio, y por eso quedan fuera de CAMPOS_EDITABLES.
CAMPOS_SRI = ('ruc_state', 'ruc_class', 'ruc_type', 'ruc_obligado_contabilidad',
              'ruc_start_date', 'ruc_end_date', 'ruc_activities',
              'ruc_establishments', 'ruc_raw', 'sri_verified')

MAX_ARCHIVO_MB = 15

# Tiempo máximo que puede pasar clasificando con IA dentro de UNA petición.
# Va por debajo del timeout de gunicorn (300 s en el Dockerfile) a propósito:
# si el modelo se pasa de aquí, el usuario recibe la vista previa de lo que sí
# se analizó y un aviso de cuánto quedó fuera, en vez de una petición cortada
# a media respuesta sin ninguna explicación.
LIMITE_IA_SEGUNDOS = 200


def _slug(texto):
    import re
    base = str(texto or '').strip().lower()
    base = base.translate(str.maketrans('áéíóúüñ', 'aeiouun'))
    base = re.sub(r'[^a-z0-9]+', '-', base).strip('-')
    return base or 'sector'


def _ip():
    return (request.headers.get('X-Forwarded-For', '').split(',')[0].strip()
            or request.remote_addr or '')


def _texto(valor):
    """Un valor cualquiera a texto plano para comparar y para la bitácora."""
    if valor is None:
        return ''
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False, sort_keys=True)
    return str(valor).strip()


def registrar_directorio(app, ctx):
    """Cuelga las rutas del directorio sobre la aplicación.

    `ctx` trae los ayudantes que viven en app/__init__.py (is_admin, user_can,
    _sanitize). Se pasan como parámetro en lugar de importarlos para no crear un
    ciclo de importación entre los dos módulos."""
    is_admin  = ctx['is_admin']
    user_can  = ctx['user_can']
    _sanitize = ctx['_sanitize']
    db        = lambda: app.supabase

    def _sin_acceso():
        return jsonify({'success': False, 'error': 'Sin acceso al módulo Directorio'}), 403

    def _puede_editar():
        """Escribir en el directorio exige el módulo; borrar exige ser admin."""
        return user_can('directorio')

    # ────────────────────────────────────────────────────────
    #  BITÁCORA
    # ────────────────────────────────────────────────────────
    def _anotar(contacto_id, doc_number, accion, motivo, cambios=None):
        """Escribe una fila por CAMPO modificado. Una sola fila por acción
        cuando no hay campos (alta, baja, importación)."""
        base = {
            'contact_id': contacto_id,
            'doc_number': doc_number,
            'action':     accion,
            'reason':     motivo,
            'user_id':    current_user.id,
            'user_name':  current_user.full_name or current_user.email,
            'ip':         _ip(),
        }
        filas = []
        for campo, (antes, despues) in (cambios or {}).items():
            filas.append(dict(base, field=campo,
                              field_label=CAMPOS_EDITABLES.get(campo, campo),
                              old_value=_texto(antes)[:2000],
                              new_value=_texto(despues)[:2000]))
        if not filas:
            filas = [base]
        try:
            db().insert('contact_audit', filas)
        except Exception as e:
            # La bitácora no debe tumbar la operación, pero sí dejar rastro en el log.
            print(f'[directorio] no se pudo escribir la bitácora: {e}')

    # ────────────────────────────────────────────────────────
    #  DUPLICADOS
    # ────────────────────────────────────────────────────────
    def _buscar_por_documento(numero):
        if not numero:
            return None
        filas = db().get('contacts', {'doc_number': numero},
                         select='id,doc_number,first_name,last_name,business_name,sector_id')
        return filas[0] if filas else None

    def _indice_documentos():
        """Todos los documentos ya guardados, en memoria. Una sola consulta
        para revisar un archivo entero en lugar de una por fila."""
        filas = db().get('contacts', select='id,doc_number,first_name,last_name,business_name') or []
        return {f['doc_number']: f for f in filas if f.get('doc_number')}

    # ────────────────────────────────────────────────────────
    #  NORMALIZACIÓN DE UN REGISTRO ENTRANTE
    # ────────────────────────────────────────────────────────
    def _normalizar(entrada, sectores_por_nombre=None):
        """Deja un registro (venga del formulario o de un archivo) listo para
        guardarse: documento en su forma canónica, tipo deducido, redes en
        lista, sector resuelto a su id."""
        registro = {}
        for campo in CAMPOS_EDITABLES:
            if campo in entrada:
                registro[campo] = entrada[campo]

        # Documento: se normaliza y se deduce el tipo cuando no viene o viene mal.
        crudo = entrada.get('doc_number') or ''
        tipo  = (entrada.get('doc_type') or '').strip().lower()
        if tipo not in ('cedula', 'ruc', 'pasaporte'):
            tipo = sri_mod.detectar_tipo(crudo)
        validacion = sri_mod.validar_documento(crudo, tipo)
        registro['doc_number'] = validacion['numero']
        registro['doc_type']   = validacion['tipo']
        registro['doc_valid']  = validacion['valido']

        for campo in ('first_name', 'last_name', 'business_name', 'trade_name',
                      'mobile', 'landline', 'email', 'website',
                      'work_address', 'home_address', 'city', 'province', 'tags'):
            if campo in registro:
                registro[campo] = _sanitize(registro[campo], 300) or None
        if 'notes' in registro:
            registro['notes'] = _sanitize(registro['notes'], 4000) or None

        # Redes sociales: siempre una lista de {red, url}, descartando lo vacío.
        redes = entrada.get('socials')
        if redes is not None:
            limpias = []
            for red in (redes if isinstance(redes, list) else []):
                if not isinstance(red, dict):
                    continue
                nombre = _sanitize(red.get('red') or red.get('nombre'), 60)
                enlace = _sanitize(red.get('url') or red.get('usuario'), 300)
                if nombre and enlace:
                    limpias.append({'red': nombre, 'url': enlace})
            registro['socials'] = limpias

        # Sector: puede llegar como id o como nombre (así lo devuelve la IA).
        if not registro.get('sector_id') and entrada.get('sector'):
            clave = _slug(entrada['sector'])
            registro['sector_id'] = (sectores_por_nombre or {}).get(clave)
        if not registro.get('sector_id'):
            registro.pop('sector_id', None)

        return registro, validacion

    def _completar_con_sri(registro, forzar=False):
        """Si el documento es un RUC válido, trae del SRI lo que publica.

        Los datos del SRI no pisan lo que el usuario escribió a mano: sólo
        rellenan lo que está vacío (salvo `forzar`, que es la re-consulta
        explícita desde la ficha)."""
        if registro.get('doc_type') != 'ruc' or not registro.get('doc_valid'):
            return {'consultado': False, 'ok': False, 'error': None}
        resultado = sri_mod.consultar_ruc(registro['doc_number'])
        if not resultado['ok']:
            registro['sri_verified'] = False
            return {'consultado': True, 'ok': False, 'error': resultado['error']}
        for campo, valor in sri_mod.contacto_desde_ruc(resultado['datos']).items():
            if valor in (None, '', [], {}):
                continue
            if forzar or campo in CAMPOS_SRI or not registro.get(campo):
                registro[campo] = valor
        return {'consultado': True, 'ok': True, 'error': None,
                'actividades': len(resultado['datos'].get('actividades') or [])}

    # ============================================================
    #  PÁGINA
    # ============================================================
    @app.route('/directorio')
    @login_required
    def directorio():
        if not user_can('directorio'):
            flash('No tienes acceso al módulo Directorio.', 'warning')
            return redirect('/dashboard')
        return render_template('directorio.html',
                               is_admin_user=is_admin(),
                               ia=ia_mod.estado(),
                               page_title='Directorio',
                               page_sub='Base de datos de clientes por sector de servicio')

    # ============================================================
    #  SECTORES
    # ============================================================
    @app.route('/directorio/api/sectores', methods=['GET'])
    @login_required
    def directorio_sectores():
        if not user_can('directorio'):
            return _sin_acceso()
        sectores = db().get('sectors', select='*') or []
        contactos = db().get('contacts', select='sector_id') or []
        conteo = {}
        for c in contactos:
            conteo[c.get('sector_id')] = conteo.get(c.get('sector_id'), 0) + 1
        for s in sectores:
            s['total'] = conteo.get(s['id'], 0)
        sectores.sort(key=lambda s: (not s.get('active'), (s.get('name') or '').lower()))
        return jsonify({'sectores': sectores, 'sin_sector': conteo.get(None, 0)})

    @app.route('/directorio/api/sectores', methods=['POST'])
    @login_required
    def directorio_crear_sector():
        if not _puede_editar():
            return _sin_acceso()
        cuerpo = request.get_json() or {}
        nombre = _sanitize(cuerpo.get('name'), 120)
        if not nombre:
            return jsonify({'success': False, 'error': 'El nombre del sector es obligatorio'})
        slug = _slug(nombre)
        if db().get('sectors', {'slug': slug}, select='id'):
            return jsonify({'success': False, 'error': f'Ya existe un sector llamado «{nombre}»'})
        fila = db().insert('sectors', {
            'name': nombre, 'slug': slug,
            'description': _sanitize(cuerpo.get('description'), 500) or None,
            'color': ctx['_sanitize_hex_color'](cuerpo.get('color')),
            'icon':  _sanitize(cuerpo.get('icon'), 8) or '📁',
            'created_by': current_user.id,
        })
        return jsonify({'success': bool(fila), 'sector': fila[0] if fila else None})

    @app.route('/directorio/api/sectores/<sid>', methods=['PATCH'])
    @login_required
    def directorio_editar_sector(sid):
        if not _puede_editar():
            return _sin_acceso()
        cuerpo = request.get_json() or {}
        cambios = {}
        if 'name' in cuerpo:
            nombre = _sanitize(cuerpo['name'], 120)
            if not nombre:
                return jsonify({'success': False, 'error': 'El nombre del sector es obligatorio'})
            otro = db().get('sectors', {'slug': _slug(nombre)}, select='id')
            if otro and otro[0]['id'] != sid:
                return jsonify({'success': False, 'error': f'Ya existe un sector llamado «{nombre}»'})
            cambios['name'] = nombre
            cambios['slug'] = _slug(nombre)
        if 'description' in cuerpo:
            cambios['description'] = _sanitize(cuerpo['description'], 500) or None
        if 'color' in cuerpo:
            cambios['color'] = ctx['_sanitize_hex_color'](cuerpo['color'])
        if 'icon' in cuerpo:
            cambios['icon'] = _sanitize(cuerpo['icon'], 8) or '📁'
        if 'active' in cuerpo:
            cambios['active'] = bool(cuerpo['active'])
        if not cambios:
            return jsonify({'success': False, 'error': 'Nada que actualizar'})
        return jsonify({'success': db().update('sectors', sid, cambios)})

    @app.route('/directorio/api/sectores/<sid>', methods=['DELETE'])
    @login_required
    def directorio_borrar_sector(sid):
        if not is_admin():
            return jsonify({'success': False, 'error': 'Sólo un administrador puede eliminar sectores'}), 403
        # Un sector con contactos no se borra: se perdería la clasificación de
        # esos registros. Se desactiva, que es lo que el usuario suele querer.
        usados = db().get('contacts', {'sector_id': sid}, select='id')
        if usados:
            return jsonify({'success': False,
                            'error': f'El sector tiene {len(usados)} registro(s). '
                                     'Muévelos a otro sector o desactívalo en lugar de eliminarlo.'})
        return jsonify({'success': db().delete('sectors', sid)})

    # ============================================================
    #  VALIDACIÓN DE DOCUMENTO + CONSULTA AL SRI
    # ============================================================
    @app.route('/directorio/api/validar-documento', methods=['POST'])
    @login_required
    def directorio_validar_documento():
        if not user_can('directorio'):
            return _sin_acceso()
        cuerpo = request.get_json() or {}
        validacion = sri_mod.validar_documento(cuerpo.get('doc_number'), cuerpo.get('doc_type'))
        respuesta = dict(validacion)

        # ¿Ya existe? Se avisa aquí, mientras el usuario escribe, no al guardar.
        existente = _buscar_por_documento(validacion['numero'])
        excluir = cuerpo.get('excluir_id')
        if existente and existente['id'] != excluir:
            nombre = (existente.get('business_name')
                      or f"{existente.get('first_name') or ''} {existente.get('last_name') or ''}".strip()
                      or existente['doc_number'])
            respuesta['duplicado'] = {'id': existente['id'], 'nombre': nombre}

        # Consulta al SRI sólo si el RUC pasó el dígito verificador.
        if validacion['tipo'] == 'ruc' and validacion['valido'] and cuerpo.get('consultar_sri', True):
            resultado = sri_mod.consultar_ruc(validacion['numero'])
            respuesta['sri'] = resultado
            if resultado['ok']:
                respuesta['campos'] = sri_mod.contacto_desde_ruc(resultado['datos'])
        return jsonify(respuesta)

    @app.route('/directorio/api/contactos/<cid>/refrescar-sri', methods=['POST'])
    @login_required
    def directorio_refrescar_sri(cid):
        if not _puede_editar():
            return _sin_acceso()
        filas = db().get('contacts', {'id': cid}, select='*')
        if not filas:
            return jsonify({'success': False, 'error': 'Registro no encontrado'}), 404
        actual = filas[0]
        if actual.get('doc_type') != 'ruc':
            return jsonify({'success': False, 'error': 'Este registro no tiene RUC'})
        resultado = sri_mod.consultar_ruc(actual['doc_number'])
        if not resultado['ok']:
            return jsonify({'success': False, 'error': resultado['error']})
        nuevos = sri_mod.contacto_desde_ruc(resultado['datos'])
        nuevos['updated_at'] = datetime.now(timezone.utc).isoformat()
        nuevos['updated_by'] = current_user.id
        ok = db().update('contacts', cid, nuevos)
        if ok:
            _anotar(cid, actual['doc_number'], 'update',
                    'Actualización de datos desde el SRI',
                    {'ruc_state': (actual.get('ruc_state'), nuevos.get('ruc_state')),
                     'ruc_activities': (actual.get('ruc_activities'), nuevos.get('ruc_activities'))})
        return jsonify({'success': ok, 'campos': nuevos,
                        'actividades': len(nuevos.get('ruc_activities') or [])})

    # ============================================================
    #  CONTACTOS — lectura
    # ============================================================
    @app.route('/directorio/api/contactos', methods=['GET'])
    @login_required
    def directorio_contactos():
        if not user_can('directorio'):
            return _sin_acceso()
        sector = request.args.get('sector_id')
        filas = (db().get('contacts', {'sector_id': sector}, select='*') if sector
                 else db().get('contacts', select='*')) or []
        busqueda = (request.args.get('q') or '').strip().lower()
        if busqueda:
            def coincide(c):
                campos = (c.get('doc_number'), c.get('first_name'), c.get('last_name'),
                          c.get('business_name'), c.get('trade_name'), c.get('email'),
                          c.get('mobile'), c.get('landline'), c.get('city'))
                return any(busqueda in str(v or '').lower() for v in campos)
            filas = [c for c in filas if coincide(c)]
        filas.sort(key=lambda c: ((c.get('last_name') or c.get('business_name') or '').lower(),
                                  (c.get('first_name') or '').lower()))
        return jsonify(filas)

    @app.route('/directorio/api/contactos/<cid>', methods=['GET'])
    @login_required
    def directorio_contacto(cid):
        if not user_can('directorio'):
            return _sin_acceso()
        filas = db().get('contacts', {'id': cid}, select='*')
        if not filas:
            return jsonify({'success': False, 'error': 'Registro no encontrado'}), 404
        return jsonify(filas[0])

    @app.route('/directorio/api/contactos/<cid>/bitacora', methods=['GET'])
    @login_required
    def directorio_bitacora(cid):
        if not user_can('directorio'):
            return _sin_acceso()
        filas = db().get('contact_audit', {'contact_id': cid}, select='*') or []
        filas.sort(key=lambda f: f.get('created_at') or '', reverse=True)
        return jsonify(filas)

    # ============================================================
    #  CONTACTOS — alta
    # ============================================================
    @app.route('/directorio/api/contactos', methods=['POST'])
    @login_required
    def directorio_crear_contacto():
        if not _puede_editar():
            return _sin_acceso()
        cuerpo = request.get_json() or {}
        registro, validacion = _normalizar(cuerpo)

        if not registro['doc_number']:
            return jsonify({'success': False, 'error': 'El número de documento es obligatorio'})
        if not validacion['valido']:
            # El documento mal formado se puede guardar a propósito (hay archivos
            # históricos con números incompletos), pero nunca por accidente.
            if not cuerpo.get('forzar_documento_invalido'):
                return jsonify({'success': False, 'error': validacion['mensaje'],
                                'documento_invalido': True})
        if not (registro.get('first_name') or registro.get('last_name')
                or registro.get('business_name')):
            return jsonify({'success': False,
                            'error': 'Escribe al menos los nombres, los apellidos o la razón social'})

        existente = _buscar_por_documento(registro['doc_number'])
        if existente:
            nombre = (existente.get('business_name')
                      or f"{existente.get('first_name') or ''} {existente.get('last_name') or ''}".strip())
            return jsonify({'success': False, 'duplicado': True, 'id_existente': existente['id'],
                            'error': f'Ya existe un registro con el documento {registro["doc_number"]} '
                                     f'({nombre or "sin nombre"}). No se puede repetir.'})

        info_sri = _completar_con_sri(registro)
        registro['created_by'] = current_user.id
        registro['updated_by'] = current_user.id
        registro['source'] = cuerpo.get('source') or 'manual'

        filas = db().insert('contacts', registro)
        if not filas:
            # El índice único de la base es la última red: si dos altas
            # simultáneas pasan la comprobación anterior, aquí falla una.
            if _buscar_por_documento(registro['doc_number']):
                return jsonify({'success': False, 'duplicado': True,
                                'error': 'Ese documento acaba de ser registrado por otro usuario.'})
            return jsonify({'success': False, 'error': 'No se pudo guardar el registro'})

        nuevo = filas[0]
        _anotar(nuevo['id'], nuevo['doc_number'], 'create',
                _sanitize(cuerpo.get('reason'), 500) or 'Alta de registro')
        return jsonify({'success': True, 'contacto': nuevo, 'sri': info_sri})

    # ============================================================
    #  CONTACTOS — modificación (exige motivo)
    # ============================================================
    @app.route('/directorio/api/contactos/<cid>', methods=['PATCH'])
    @login_required
    def directorio_editar_contacto(cid):
        if not _puede_editar():
            return _sin_acceso()
        cuerpo = request.get_json() or {}
        motivo = _sanitize(cuerpo.get('reason'), 500)
        if not motivo or len(motivo) < 5:
            return jsonify({'success': False, 'falta_motivo': True,
                            'error': 'Indica el motivo del cambio (mínimo 5 caracteres). '
                                     'Queda registrado en la bitácora.'})

        filas = db().get('contacts', {'id': cid}, select='*')
        if not filas:
            return jsonify({'success': False, 'error': 'Registro no encontrado'}), 404
        actual = filas[0]

        registro, validacion = _normalizar(cuerpo)
        # Sólo se tocan los campos que vinieron en la petición.
        registro = {k: v for k, v in registro.items() if k in cuerpo or k in ('doc_number', 'doc_type', 'doc_valid')}
        if 'doc_number' not in cuerpo:
            for campo in ('doc_number', 'doc_type', 'doc_valid'):
                registro.pop(campo, None)
        elif not validacion['valido'] and not cuerpo.get('forzar_documento_invalido'):
            return jsonify({'success': False, 'error': validacion['mensaje'],
                            'documento_invalido': True})

        # Cambiar el documento a uno que ya existe es, otra vez, un duplicado.
        if registro.get('doc_number') and registro['doc_number'] != actual.get('doc_number'):
            otro = _buscar_por_documento(registro['doc_number'])
            if otro and otro['id'] != cid:
                return jsonify({'success': False, 'duplicado': True, 'id_existente': otro['id'],
                                'error': f'Ya existe otro registro con el documento '
                                         f'{registro["doc_number"]}.'})
            _completar_con_sri(registro)

        # Qué cambió realmente. Un PATCH que reenvía el formulario entero trae
        # veinte campos iguales; en la bitácora sólo deben quedar los distintos.
        cambios = {campo: (actual.get(campo), valor)
                   for campo, valor in registro.items()
                   if campo in CAMPOS_EDITABLES and _texto(actual.get(campo)) != _texto(valor)}
        if not cambios:
            return jsonify({'success': True, 'sin_cambios': True,
                            'mensaje': 'No había nada distinto que guardar'})

        registro['updated_at'] = datetime.now(timezone.utc).isoformat()
        registro['updated_by'] = current_user.id
        ok = db().update('contacts', cid, registro)
        if ok:
            _anotar(cid, registro.get('doc_number') or actual.get('doc_number'),
                    'update', motivo, cambios)
        return jsonify({'success': ok, 'cambios': len(cambios)})

    @app.route('/directorio/api/contactos/<cid>', methods=['DELETE'])
    @login_required
    def directorio_borrar_contacto(cid):
        if not is_admin():
            return jsonify({'success': False, 'error': 'Sólo un administrador puede eliminar registros'}), 403
        motivo = _sanitize((request.get_json() or {}).get('reason'), 500)
        if not motivo or len(motivo) < 5:
            return jsonify({'success': False, 'falta_motivo': True,
                            'error': 'Indica el motivo de la eliminación (mínimo 5 caracteres).'})
        filas = db().get('contacts', {'id': cid}, select='*')
        if not filas:
            return jsonify({'success': False, 'error': 'Registro no encontrado'}), 404
        actual = filas[0]
        # La bitácora se escribe ANTES del borrado: contact_audit.contact_id
        # tiene ON DELETE CASCADE, así que la fila muere con el contacto — por eso
        # `doc_number` se guarda suelto y el historial de quién borró qué queda
        # consultable por documento aunque el registro ya no exista.
        _anotar(cid, actual.get('doc_number'), 'delete', motivo,
                {'doc_number': (actual.get('doc_number'), ''),
                 'business_name': (actual.get('business_name'), ''),
                 'first_name': (actual.get('first_name'), ''),
                 'last_name': (actual.get('last_name'), '')})
        return jsonify({'success': db().delete('contacts', cid)})

    # ============================================================
    #  IMPORTACIÓN EN BLOQUE — dos pasos: analizar y confirmar
    # ============================================================
    @app.route('/directorio/api/ia-estado', methods=['GET'])
    @login_required
    def directorio_ia_estado():
        return jsonify(ia_mod.estado())

    @app.route('/directorio/api/importar/analizar', methods=['POST'])
    @login_required
    def directorio_importar_analizar():
        """Lee el archivo, clasifica los datos y devuelve una VISTA PREVIA.

        No inserta nada. Devuelve cada fila con su estado (`nuevo`, `duplicado`,
        `documento_invalido`, `incompleto`) para que el usuario vea qué va a
        entrar y qué se va a rechazar antes de confirmar."""
        if not _puede_editar():
            return _sin_acceso()
        archivo = request.files.get('file')
        if not archivo:
            return jsonify({'success': False, 'error': 'No se recibió ningún archivo'})
        contenido = archivo.read()
        if len(contenido) > MAX_ARCHIVO_MB * 1024 * 1024:
            return jsonify({'success': False,
                            'error': f'El archivo supera los {MAX_ARCHIVO_MB} MB'})

        try:
            leido = docs_mod.extraer(archivo.filename, contenido)
        except docs_mod.FormatoNoSoportado as e:
            return jsonify({'success': False, 'error': str(e)})
        except Exception as e:
            return jsonify({'success': False, 'error': f'No se pudo leer el archivo: {str(e)[:200]}'})

        if not leido['filas'] and not leido['bloques']:
            return jsonify({'success': False,
                            'error': 'El archivo no tiene datos: sólo se encontró la fila de '
                                     'cabeceras (o está vacío).'})

        sectores = db().get('sectors', select='id,name,slug') or []
        por_slug = {s['slug']: s['id'] for s in sectores}
        por_slug.update({_slug(s['name']): s['id'] for s in sectores})

        usar_ia = (request.form.get('usar_ia', '1') == '1') and ia_mod.disponible()
        avisos, metodo = [], ''
        crudos = []

        # Camino barato primero: si el Excel trae cabeceras reconocibles, se mapea
        # sin gastar una llamada al modelo. La IA entra cuando no alcanzan.
        if leido['filas'] and not usar_ia:
            crudos = docs_mod.mapear_por_cabeceras(leido['filas'], leido['cabeceras'])
            metodo = 'cabeceras'
            if not crudos:
                avisos.append('No se reconocieron las cabeceras del archivo. '
                              'Activa la clasificación con IA para interpretarlo.')
        elif leido['filas'] and usar_ia:
            directo = docs_mod.mapear_por_cabeceras(leido['filas'], leido['cabeceras'])
            reconocidas = docs_mod.cabeceras_reconocidas(leido['cabeceras'])
            if directo and len(reconocidas) >= 4:
                crudos, metodo = directo, 'cabeceras'
            else:
                try:
                    crudos, avisos = ia_mod.clasificar_registros(
                        leido['bloques'], sectores, origen=leido['tipo'],
                        limite_segundos=LIMITE_IA_SEGUNDOS)
                    metodo = 'ia'
                except ia_mod.IANoDisponible as e:
                    return jsonify({'success': False, 'error': str(e)})
                except Exception as e:
                    return jsonify({'success': False,
                                    'error': f'La IA no pudo clasificar el archivo: {str(e)[:200]}'})
        else:
            # PDF o Word sin tablas: no hay otro camino que la IA.
            if not ia_mod.disponible():
                return jsonify({'success': False,
                                'error': 'Un PDF o un Word sin tablas necesita la clasificación con IA. '
                                         + (ia_mod.estado()['motivo'] or '')})
            try:
                crudos, avisos = ia_mod.clasificar_registros(
                    leido['bloques'], sectores, origen=leido['tipo'],
                        limite_segundos=LIMITE_IA_SEGUNDOS)
                metodo = 'ia'
            except Exception as e:
                return jsonify({'success': False,
                                'error': f'La IA no pudo clasificar el archivo: {str(e)[:200]}'})

        if not crudos:
            return jsonify({'success': False,
                            'error': 'No se encontró ningún registro aprovechable en el archivo',
                            'avisos': avisos})

        # Sector por defecto elegido en el formulario de importación.
        sector_forzado = request.form.get('sector_id') or None

        ya_guardados = _indice_documentos()
        vistos_en_archivo = {}
        vista = []
        for indice, crudo in enumerate(crudos):
            registro, validacion = _normalizar(crudo, por_slug)
            if sector_forzado:
                registro['sector_id'] = sector_forzado
            numero = registro.get('doc_number') or ''
            nombre = (registro.get('business_name')
                      or f"{registro.get('first_name') or ''} {registro.get('last_name') or ''}".strip())

            estado, detalle = 'nuevo', ''
            if not numero:
                estado = 'incompleto'
                detalle = 'Sin número de documento'
            elif numero in ya_guardados:
                otro = ya_guardados[numero]
                estado = 'duplicado'
                detalle = ('Ya está en la base: '
                           + (otro.get('business_name')
                              or f"{otro.get('first_name') or ''} {otro.get('last_name') or ''}".strip()
                              or numero))
            elif numero in vistos_en_archivo:
                estado = 'duplicado'
                detalle = f'Repetido dentro del archivo (fila {vistos_en_archivo[numero] + 1})'
            elif not validacion['valido']:
                estado = 'documento_invalido'
                detalle = validacion['mensaje']
            elif not nombre:
                estado = 'incompleto'
                detalle = 'Sin nombres ni razón social'

            if estado == 'nuevo':
                vistos_en_archivo[numero] = indice
            vista.append({
                'indice': indice, 'estado': estado, 'detalle': detalle,
                'confianza': crudo.get('confianza') or 'alta',
                'registro': registro, 'nombre': nombre or '(sin nombre)',
            })

        resumen = {'total': len(vista)}
        for fila in vista:
            resumen[fila['estado']] = resumen.get(fila['estado'], 0) + 1
        return jsonify({'success': True, 'metodo': metodo, 'tipo': leido['tipo'],
                        'archivo': archivo.filename, 'avisos': avisos,
                        'resumen': resumen, 'filas': vista})

    @app.route('/directorio/api/importar/confirmar', methods=['POST'])
    @login_required
    def directorio_importar_confirmar():
        """Guarda las filas que el usuario aprobó en la vista previa.

        Vuelve a comprobar los duplicados: entre el análisis y la confirmación
        pudo entrar el mismo documento por otra vía."""
        if not _puede_editar():
            return _sin_acceso()
        cuerpo = request.get_json() or {}
        filas = cuerpo.get('filas') or []
        if not filas:
            return jsonify({'success': False, 'error': 'No se enviaron registros para guardar'})
        motivo = _sanitize(cuerpo.get('reason'), 500) or f'Importación desde {cuerpo.get("archivo") or "archivo"}'
        consultar_sri = bool(cuerpo.get('consultar_sri'))

        ya_guardados = _indice_documentos()
        insertados, rechazados = 0, []
        vistos = set()

        for fila in filas:
            registro, validacion = _normalizar(fila)
            numero = registro.get('doc_number') or ''
            if not numero:
                rechazados.append({'nombre': fila.get('business_name') or fila.get('first_name') or '?',
                                   'motivo': 'Sin número de documento'})
                continue
            if numero in ya_guardados or numero in vistos:
                rechazados.append({'nombre': numero, 'motivo': 'Duplicado: ya existe en la base'})
                continue
            if not (registro.get('first_name') or registro.get('last_name')
                    or registro.get('business_name')):
                rechazados.append({'nombre': numero, 'motivo': 'Sin nombres ni razón social'})
                continue

            # La consulta al SRI en bloque es lenta (una llamada por RUC), así que
            # es opcional y el usuario decide si la quiere en esta importación.
            if consultar_sri:
                _completar_con_sri(registro)

            registro['created_by'] = current_user.id
            registro['updated_by'] = current_user.id
            registro['source'] = cuerpo.get('tipo') or 'excel'
            registro['source_file'] = _sanitize(cuerpo.get('archivo'), 200) or None

            guardado = db().insert('contacts', registro)
            if guardado:
                vistos.add(numero)
                insertados += 1
                _anotar(guardado[0]['id'], numero, 'import', motivo)
            else:
                rechazados.append({'nombre': numero,
                                   'motivo': 'La base rechazó el registro (posible duplicado)'})

        return jsonify({'success': True, 'insertados': insertados,
                        'rechazados': rechazados, 'total_rechazados': len(rechazados)})

    # ============================================================
    #  EXPORTACIÓN
    # ============================================================
    @app.route('/directorio/api/exportar')
    @login_required
    def directorio_exportar():
        if not user_can('directorio'):
            return jsonify({'error': 'Sin acceso'}), 403
        if not docs_mod.EXCEL_DISPONIBLE:
            return jsonify({'error': 'openpyxl no está instalado en el servidor'}), 500
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        sector = request.args.get('sector_id')
        filas = (db().get('contacts', {'sector_id': sector}, select='*') if sector
                 else db().get('contacts', select='*')) or []
        sectores = {s['id']: s['name'] for s in (db().get('sectors', select='id,name') or [])}

        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = 'Directorio'
        cabeceras = ['Sector', 'Tipo', 'Documento', 'Válido', 'Verificado SRI', 'Apellidos',
                     'Nombres', 'Razón social', 'Nombre comercial', 'Celular', 'Convencional',
                     'Correo', 'Dirección web', 'Redes sociales', 'Dirección de trabajo',
                     'Dirección domicilio', 'Ciudad', 'Provincia', 'Estado RUC',
                     'Actividades económicas', 'Observaciones']
        relleno = PatternFill('solid', fgColor='4F46E5')
        fuente = Font(color='FFFFFF', bold=True)
        for i, titulo in enumerate(cabeceras, 1):
            celda = hoja.cell(row=1, column=i, value=titulo)
            celda.fill, celda.font = relleno, fuente
            celda.alignment = Alignment(horizontal='center')

        xs = ctx['_xlsx_safe']
        for fi, c in enumerate(filas, 2):
            redes = '; '.join(f"{r.get('red')}: {r.get('url')}"
                              for r in (c.get('socials') or []) if isinstance(r, dict))
            actividades = ' | '.join(a.get('descripcion', '') for a in (c.get('ruc_activities') or [])
                                     if isinstance(a, dict))
            valores = [
                sectores.get(c.get('sector_id'), ''), c.get('doc_type'), xs(c.get('doc_number')),
                'Sí' if c.get('doc_valid') else 'No', 'Sí' if c.get('sri_verified') else 'No',
                xs(c.get('last_name')), xs(c.get('first_name')), xs(c.get('business_name')),
                xs(c.get('trade_name')), xs(c.get('mobile')), xs(c.get('landline')),
                xs(c.get('email')), xs(c.get('website')), xs(redes), xs(c.get('work_address')),
                xs(c.get('home_address')), xs(c.get('city')), xs(c.get('province')),
                c.get('ruc_state'), xs(actividades), xs(c.get('notes')),
            ]
            for ci, valor in enumerate(valores, 1):
                hoja.cell(row=fi, column=ci, value=valor)
        for columna in hoja.columns:
            ancho = max((len(str(celda.value or '')) for celda in columna), default=0)
            hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 3, 55)

        buffer = io.BytesIO()
        libro.save(buffer)
        buffer.seek(0)
        nombre = f'directorio_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(buffer, as_attachment=True, download_name=nombre,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return app
