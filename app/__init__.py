# ------------------------------------------------------------
# Desarrollado por Marco Antonio Posligua San Martín
# ------------------------------------------------------------
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory, send_file, g, current_app
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import generate_csrf, validate_csrf
from wtforms.validators import ValidationError
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from config.config import Config
from dotenv import load_dotenv
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
import google.auth.exceptions
from googleapiclient.discovery import build
from datetime import datetime, timedelta, timezone, date
from collections import defaultdict, OrderedDict
import os, requests as req_lib, traceback, pytz, json, re, time, calendar as _cal, io, threading, tempfile, secrets
from urllib.parse import quote as _url_quote, urlparse as _urlparse

# Web push (opcional: si la librería no está disponible la app sigue funcionando)
try:
    from pywebpush import webpush, WebPushException
    from py_vapid import Vapid
    import base64
    WEB_PUSH_AVAILABLE = True
except Exception:
    WEB_PUSH_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

GOOGLE_SCOPES = ['https://www.googleapis.com/auth/calendar']

# Microsoft Graph — To-Do
MS_AUTH_URL   = 'https://login.microsoftonline.com/consumers/oauth2/v2.0/authorize'
MS_TOKEN_URL  = 'https://login.microsoftonline.com/consumers/oauth2/v2.0/token'
MS_GRAPH_URL  = 'https://graph.microsoft.com/v1.0'
MS_SCOPES     = 'Tasks.ReadWrite offline_access User.Read'
GOOGLE_ACCOUNT_EMAIL = 'mposligua0000@gmail.com'

# WebAuthn / passkeys (Face ID, huella). Import protegido: si la librería aún
# no está instalada, la app arranca igual y la función queda deshabilitada.
try:
    from webauthn import (
        generate_registration_options, verify_registration_response,
        generate_authentication_options, verify_authentication_response,
        options_to_json,
    )
    from webauthn.helpers import bytes_to_base64url, base64url_to_bytes
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, ResidentKeyRequirement,
        UserVerificationRequirement, PublicKeyCredentialDescriptor,
    )
    WEBAUTHN_AVAILABLE = True
except Exception as _wa_err:  # pragma: no cover
    WEBAUTHN_AVAILABLE = False
    print(f'[webauthn] no disponible: {_wa_err}')

from .feriados import feriados as _feriados_ec, feriados_rango as _feriados_rango
from . import browser_sync as _browser_sync
# Módulos propios. No importan nada de este archivo (reciben los ayudantes que
# necesitan como parámetro), así que no se forma ciclo de importación.
from .directorio import registrar_directorio
from .cronograma import registrar_cronograma
from . import atlas_sync as _atlas

load_dotenv()
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
TIMEZONE = pytz.timezone('America/Guayaquil')
login_manager = LoginManager()
limiter = Limiter(key_func=get_remote_address, default_limits=[])


# ============================================================
#  TTL CACHE — in-process, protegido con lock para uso multi-hilo
# ============================================================
class TTLCache:
    """Lightweight TTL cache with LRU eviction."""
    def __init__(self, ttl=60, maxsize=256):
        self._data = OrderedDict()
        self._ts = {}
        self.ttl = ttl
        self.maxsize = maxsize
        self._lock = threading.Lock()

    def get(self, key):
        with self._lock:
            if key in self._data:
                if time.monotonic() - self._ts[key] < self.ttl:
                    self._data.move_to_end(key)
                    return self._data[key], True
                self._evict(key)
            return None, False

    def set(self, key, value):
        with self._lock:
            if key in self._data:
                self._data.move_to_end(key)
            self._data[key] = value
            self._ts[key] = time.monotonic()
            if len(self._data) > self.maxsize:
                oldest = next(iter(self._data))
                self._evict(oldest)

    def _evict(self, key):
        self._data.pop(key, None)
        self._ts.pop(key, None)

    def invalidate(self, key):
        with self._lock:
            self._evict(key)

    def invalidate_prefix(self, prefix):
        with self._lock:
            for k in [k for k in self._data if k.startswith(prefix)]:
                self._evict(k)

# Module-level caches (shared across requests in same worker)
_cal_cache      = TTLCache(ttl=300)   # calendar_config — 5 min
_user_cal_cache = TTLCache(ttl=10)    # user calendars  — 10 s (corto: multi-worker safe)
_google_cache   = TTLCache(ttl=120)   # google status   — 2 min
# Grants de un rol (módulos/calendarios/proyectos/cuentas MS). 10 s, como sus dos
# vecinos y por la misma razón: en producción corren DOS workers de gunicorn, cada
# uno con su memoria. Al guardar un rol sólo se invalida el caché del worker que
# atendió la petición; el otro seguía sirviendo lo viejo. Con 300 s eso significaba
# que un cambio de permisos tardaba hasta CINCO MINUTOS en notarse, y de forma
# intermitente según a qué worker cayera cada petición: el administrador marcaba
# una casilla, guardaba, y para el usuario no cambiaba nada. Parecía que la
# pantalla de Roles no hiciera nada. Un caché que hace dudar de si el sistema
# obedece cuesta más de lo que ahorra.
_role_cache       = TTLCache(ttl=10)
_user_roles_cache = TTLCache(ttl=10)   # roles asignados a un usuario — 10 s
_ms_cache         = TTLCache(ttl=120)  # ¿hay alguna cuenta Microsoft conectada? — 2 min


# ============================================================
#  USER MODEL
# ============================================================
class User(UserMixin):
    def __init__(self, d):
        self.id = d.get('id'); self.email = d.get('email')
        self.full_name = d.get('full_name'); self.role = d.get('role', 'staff')
        self.is_admin = d.get('role') == 'admin'
        # Cuenta desactivada por el administrador. Se asume activa cuando falta
        # el dato: nadie debe quedarse fuera porque una consulta no trajera la
        # columna.
        self.active = d.get('is_active', True) is not False
        raw = d.get('modules', 'calendar,planning') or 'calendar,planning'
        self.modules = [m.strip() for m in raw.split(',') if m.strip()]

# Módulos del sistema, en el orden en que se presentan al usuario: primero lo
# que se usa a diario (actividades y agenda), después lo que se planifica, y al
# final los datos maestros. Este orden es el que siguen el menú lateral y la
# pantalla de roles, para que el usuario encuentre lo mismo en los dos sitios.
ALL_MODULES = [
    ('todo',        '✅ Actividades (To-Do)'),
    ('planning',    '📋 Proyectos y tareas'),
    ('calendar',    '📅 Calendario de citas'),
    ('cronograma',  '📊 Cronograma (Gantt)'),
    ('directorio',  '🗂️ Directorio de clientes'),
]

# Niveles de rol (clasificación de negocio). Es una etiqueta/agrupación: el
# acceso real lo definen los grants marcados por el admin. El administrador del
# sistema real sigue siendo users.role == 'admin'.
# Escalas, en el mismo juego que usa ATLAS (admin, socio, secretaria, profesor,
# psicólogo). Antes sólo había tres y a una secretaria había que meterla en el
# cajón de «funcionario», perdiendo la distinción que en ATLAS sí existe.
ROLE_LEVELS = [
    ('administrador', '👑 Administrador'),
    ('socio',         '🤝 Socio'),
    ('funcionario',   '🧑‍💼 Funcionario'),
    ('secretaria',    '🗒️ Secretaría'),
    ('profesor',      '👩‍🏫 Profesor'),
    ('psicologo',     '🧠 Psicólogo'),
]
ROLE_LEVEL_IDS = {lid for lid, _ in ROLE_LEVELS}
DEFAULT_ROLE_LEVEL = 'funcionario'


# ============================================================
#  PERMISOS POR PERSONA Y SUBMÓDULO
#
#  El rol da el MÓDULO entero o nada. Eso deja fuera un caso muy común: que
#  alguien pueda consultar y editar el directorio pero NO importar en bloque ni
#  eliminar. ATLAS ya lo resolvía con `usuario_permisos` (una fila por usuario y
#  `familia.submodulo`), y aquí faltaba.
#
#  La regla es:
#    * Acción NO sensible  -> basta con tener el módulo en el rol activo.
#    * Acción SENSIBLE     -> además hace falta el permiso concreto concedido a
#                             esa persona (o ser administrador).
#
#  Así el rol sigue siendo la base y esto es la capa fina encima, que se afina
#  persona a persona sin inventar un rol nuevo para cada excepción.
# ============================================================
SUBMODULOS = {
    'todo': [
        ('ver',        'Ver los pendientes',            False),
        ('crear',      'Crear pendientes',              False),
        ('editar',     'Editar pendientes',             False),
        ('sincronizar','Sincronizar con Microsoft',     True),
        ('eliminar',   'Eliminar pendientes',           True),
    ],
    'planning': [
        ('ver',        'Ver tareas y proyectos',        False),
        ('crear',      'Crear tareas',                  False),
        ('editar',     'Editar tareas',                 False),
        ('proyectos',  'Crear y editar proyectos',      True),
        ('eliminar',   'Eliminar tareas',               True),
    ],
    'calendar': [
        ('ver',        'Ver la agenda',                 False),
        ('agendar',    'Agendar citas',                 False),
        ('aprobar',    'Aprobar o rechazar citas',      True),
        ('eliminar',   'Eliminar citas',                True),
    ],
    'cronograma': [
        ('ver',           'Ver cronogramas',            False),
        ('crear',         'Crear cronogramas',          False),
        ('editar',        'Editar actividades',         False),
        ('planificar_ia', 'Planificar con IA',          True),
        ('eliminar',      'Eliminar cronogramas',       True),
    ],
    'directorio': [
        ('ver',       'Consultar el directorio',        False),
        ('crear',     'Crear registros',                False),
        ('editar',    'Editar registros',               False),
        ('importar',  'Importar archivos en bloque',    True),
        ('exportar',  'Exportar a Excel',               True),
        ('sectores',  'Administrar sectores',           True),
        ('eliminar',  'Eliminar registros',             True),
    ],
}

# Acciones que exigen permiso explícito, en forma de conjunto para consultarlo rápido.
ACCIONES_SENSIBLES = {f'{mod}.{acc}'
                     for mod, lista in SUBMODULOS.items()
                     for acc, _, sensible in lista if sensible}

_user_perms_cache = TTLCache(ttl=30, maxsize=256)


# ============================================================
#  MÓDULO DE CUENTA — clave propia y restablecimiento administrativo
# ============================================================
# Longitud mínima exigida al definir una contraseña nueva.
MIN_PASSWORD = 8

# Rutas que un usuario con clave temporal SÍ puede visitar (si no, quedaría
# encerrado sin poder cambiarla ni cerrar sesión).
FREE_PATHS_PASSWORD = ('/account/password', '/logout', '/login', '/static/', '/sw.js')


def _gen_temp_password(length=12):
    """Clave temporal legible: sin caracteres ambiguos (0/O, 1/l/I) para poder
    dictarla por teléfono sin errores. Aleatoriedad criptográfica."""
    alphabet = 'ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789'
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def _temp_password_expired(row):
    """True si el usuario arrastra una clave temporal ya caducada."""
    if not row or not row.get('must_change_password'):
        return False
    exp = row.get('temp_password_expires')
    if not exp:
        return False
    try:
        return datetime.fromisoformat(str(exp).replace('Z', '+00:00')) < datetime.now(timezone.utc)
    except ValueError:
        return False


def _cuenta_activa(fila_usuario):
    """¿Está habilitada esta cuenta?

    `users.is_active` existía en la base desde hace tiempo pero NO se consultaba
    en ninguna línea del sistema: era una bandera que no cerraba ninguna puerta.
    A partir de aquí sí. Sólo `false` bloquea; que la consulta no traiga la
    columna no deja a nadie fuera —una comprobación de acceso que se equivoca
    tiene que equivocarse dejando pasar a quien ya entraba, no cerrándole a
    todos."""
    return (fila_usuario or {}).get('is_active', True) is not False


def _log_password(app, user_id, action, executed_by=None):
    """Bitácora de cambios de clave. Nunca guarda la clave, solo el hecho."""
    try:
        app.supabase.insert('password_log', {
            'user_id': user_id, 'action': action, 'executed_by': executed_by,
            'ip': (request.headers.get('X-Forwarded-For', '').split(',')[0].strip()
                   or request.remote_addr or ''),
        })
    except Exception:
        # Auxiliar: si la tabla aún no está migrada no debe frenar el cambio.
        pass


def _auditar_acceso(app, accion, target_id=None, target_email=None, detalle=''):
    """Deja constancia de un cambio de accesos: quién, a quién, qué y cuándo.

    Un sistema donde el administrador reparte permisos y después nadie puede
    reconstruir quién dio qué no es un sistema de permisos, es una costumbre.
    Se guarda también el CORREO de las dos partes, no sólo su identificador: si
    mañana se borra la cuenta, el registro tiene que seguir diciendo a quién se
    le concedió aquello.

    Nunca interrumpe la operación que la llama: apuntar el hecho es importante,
    pero menos que el hecho mismo."""
    try:
        app.supabase.insert('permission_audit', {
            'actor_id':       str(current_user.id) if current_user.is_authenticated else None,
            'actor_email':    (current_user.email if current_user.is_authenticated else None),
            'target_user_id': str(target_id) if target_id else None,
            'target_email':   target_email,
            'accion':         accion,
            'detalle':        (detalle or '')[:1000],
            'ip': (request.headers.get('X-Forwarded-For', '').split(',')[0].strip()
                   or request.remote_addr or ''),
        })
    except Exception:
        pass


# ============================================================
#  SUPABASE CLIENT — persistent HTTP session (keep-alive)
# ============================================================
class SupabaseAPI:
    def __init__(self, url, key):
        self.url = url
        self._session = req_lib.Session()
        self._headers = {
            'apikey': key,
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
        }
        self._session.headers.update(self._headers)
        # Set sane timeouts for all calls
        self._timeout = (4, 10)   # (connect, read)

    def get(self, table, filters=None, select='*'):
        q = f'{self.url}/rest/v1/{table}?select={select}'
        if filters:
            for k, v in filters.items():
                q += f'&{k}=eq.{_url_quote(str(v), safe="")}'
        # Un reintento ante fallo pasajero. La primera consulta después de un
        # rato sin uso puede pasarse del límite de tiempo de PostgreSQL
        # («canceling statement due to statement timeout») o del nuestro; en
        # caliente la misma consulta tarda dos décimas. Sin reintento eso se
        # traducía en devolver [] y que el panel mostrara ceros como si de
        # verdad no hubiera nada: peor que un error, porque no se nota.
        # No se reintentan los errores de la consulta en sí (4xx): esos no
        # mejoran repitiéndolos.
        for intento in (1, 2):
            try:
                r = self._session.get(q, timeout=self._timeout)
                if r.status_code == 200:
                    return r.json()
                if intento == 1 and r.status_code >= 500:
                    continue
                print(f'[supabase.get] {table}: HTTP {r.status_code} {r.text[:120]}')
                return []
            except Exception as e:
                if intento == 1:
                    continue
                print(f'[supabase.get] {table}: {e}')
                return []
        return []

    def get_in(self, table, column, values, select='*'):
        """Single query WHERE column IN (values)."""
        if not values:
            return []
        ids = ','.join(_url_quote(str(v), safe='') for v in values)
        q = f'{self.url}/rest/v1/{table}?select={select}&{column}=in.({ids})'
        try:
            r = self._session.get(q, timeout=self._timeout)
            if r.status_code == 200:
                return r.json()
            print(f'[supabase.get_in] {table}: HTTP {r.status_code} {r.text[:120]}')
            return []
        except Exception as e:
            print(f'[supabase.get_in] {table}: {e}')
            return []

    def insert(self, table, data):
        h = {'Prefer': 'return=representation'}
        try:
            r = self._session.post(f'{self.url}/rest/v1/{table}', headers=h, json=data, timeout=self._timeout)
            if r.status_code in [200, 201]:
                body = r.json()
                return body if isinstance(body, list) else [body]
            print(f'[supabase.insert] {table}: HTTP {r.status_code} {r.text[:120]}')
            return None
        except Exception as e:
            print(f'[supabase.insert] {table}: {e}')
            return None

    def insert_ignore(self, table, data):
        """Insert and silently ignore unique-constraint conflicts."""
        h = {'Prefer': 'resolution=ignore-duplicates,return=minimal'}
        try:
            r = self._session.post(f'{self.url}/rest/v1/{table}', headers=h, json=data, timeout=self._timeout)
            if r.status_code in [200, 201, 204]:
                return True
            print(f'[supabase.insert_ignore] {table}: HTTP {r.status_code} {r.text[:120]}')
            return False
        except Exception as e:
            print(f'[supabase.insert_ignore] {table}: {e}')
            return False

    def insert_on_conflict(self, table, data, on_conflict, timeout=None):
        """Insert que IGNORA las filas en conflicto con el índice `on_conflict`.

        Evita duplicados a nivel de BD: si una fila choca (mismo source_id), se
        omite y el resto del lote sí se inserta. Devuelve la lista de filas
        realmente insertadas, o None si hubo error de red/servidor.
        """
        h = {'Prefer': 'resolution=ignore-duplicates,return=representation'}
        try:
            r = self._session.post(
                f'{self.url}/rest/v1/{table}?on_conflict={on_conflict}',
                headers=h, json=data, timeout=timeout or self._timeout)
            if r.status_code in (200, 201):
                body = r.json()
                return body if isinstance(body, list) else [body]
            if r.status_code == 204:
                return []
            print(f'[supabase.insert_on_conflict] {table}: HTTP {r.status_code} {r.text[:120]}')
            return None
        except Exception as e:
            print(f'[supabase.insert_on_conflict] {table}: {e}')
            return None

    def update(self, table, id_val, data, id_col='id'):
        h = {'Prefer': 'return=minimal'}
        try:
            r = self._session.patch(
                f'{self.url}/rest/v1/{table}?{id_col}=eq.{_url_quote(str(id_val), safe="")}',
                headers=h, json=data, timeout=self._timeout)
            if r.status_code in [200, 204]:
                return True
            print(f'[supabase.update] {table}: HTTP {r.status_code} {r.text[:120]}')
            return False
        except Exception as e:
            print(f'[supabase.update] {table}: {e}')
            return False

    def delete(self, table, id_val, id_col='id'):
        h = {'Prefer': 'return=minimal'}
        try:
            r = self._session.delete(
                f'{self.url}/rest/v1/{table}?{id_col}=eq.{_url_quote(str(id_val), safe="")}',
                headers=h, timeout=self._timeout)
            if r.status_code in [200, 204]:
                return True
            print(f'[supabase.delete] {table}: HTTP {r.status_code} {r.text[:120]}')
            return False
        except Exception as e:
            print(f'[supabase.delete] {table}: {e}')
            return False

    def update_where(self, table, filters, data):
        """PATCH condicional: solo aplica si TODAS las filas coinciden con `filters`
        (ej. {'id': aid, 'status': 'pending'}). Devuelve las filas actualizadas
        (lista vacía si ninguna coincidía, útil para evitar TOCTOU al 'reclamar' una fila)."""
        q = f'{self.url}/rest/v1/{table}?' + '&'.join(
            f'{k}=eq.{_url_quote(str(v), safe="")}' for k, v in filters.items())
        h = {'Prefer': 'return=representation'}
        try:
            r = self._session.patch(q, headers=h, json=data, timeout=self._timeout)
            if r.status_code in (200, 201):
                body = r.json()
                return body if isinstance(body, list) else [body]
            if r.status_code == 204:
                return []
            print(f'[supabase.update_where] {table}: HTTP {r.status_code} {r.text[:120]}')
            return []
        except Exception as e:
            print(f'[supabase.update_where] {table}: {e}')
            return []

    def get_q(self, table, query_params=None, select='*'):
        """Query with raw PostgREST filter params, e.g. {'status': 'eq.done'}."""
        q = f'{self.url}/rest/v1/{table}?select={select}'
        for k, v in (query_params or {}).items():
            q += f'&{k}={v}'
        try:
            r = self._session.get(q, timeout=self._timeout)
            if r.status_code == 200:
                return r.json()
            print(f'[supabase.get_q] {table}: HTTP {r.status_code} {r.text[:120]}')
            return []
        except Exception as e:
            print(f'[supabase.get_q] {table}: {e}')
            return []


# ============================================================
#  HELPERS
# ============================================================
def _is_invalid_grant(err):
    s = str(err).lower()
    return 'invalid_grant' in s or 'expired or revoked' in s or 'token has been expired' in s

def _sanitize(s, max_len=255):
    return str(s).strip()[:max_len] if s else ''

def _validate_email(email):
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email or ''))

def _xlsx_safe(v):
    """Neutraliza inyección de fórmulas: si un valor de texto controlado por el
    usuario empieza con =, +, -, @ (o tab/CR), Excel podría interpretarlo como
    fórmula al abrir el archivo exportado. Se le antepone una comilla simple
    para forzar que se trate como texto literal."""
    if isinstance(v, str) and v[:1] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + v
    return v

def _sanitize_hex_color(color, default='#4f46e5'):
    """Los colores de proyecto se insertan sin escapar en atributos style en planning.html;
    forzar que sean siempre un hex #rrggbb evita inyección de CSS/atributos."""
    if isinstance(color, str) and re.match(r'^#[0-9a-fA-F]{6}$', color):
        return color
    return default

def _safe_next_path(target, default='/dashboard'):
    """Solo permite redirigir a una ruta local propia (evita open redirect)."""
    if not target:
        return default
    parsed = _urlparse(target)
    if parsed.scheme or parsed.netloc or not target.startswith('/') or target.startswith('//'):
        return default
    return target


# ============================================================
#  GOOGLE CALENDAR EVENT BUILDER
# ============================================================
def _build_google_event(apt, attendees):
    """Build a Google Calendar event body, properly handling virtual vs presencial."""
    is_virtual = bool(apt.get('meeting_link'))
    desc = (f"Titulo: {apt.get('title', '')}\n"
            f"Encargado: {apt.get('encargado', '')}\n"
            f"Tema: {apt.get('tema', '')}")
    if apt.get('client_name'): desc += f"\nCliente: {apt['client_name']}"

    if is_virtual:
        desc += f"\n\n🔗 Enlace de reunion: {apt['meeting_link']}"
        location = apt['meeting_link']
    else:
        if apt.get('lugar'):     desc += f"\nLugar: {apt['lugar']}"
        if apt.get('direccion'): desc += f"\nDireccion: {apt['direccion']}"
        if apt.get('ciudad'):    desc += f"\nCiudad: {apt['ciudad']}, Ecuador"
        if apt.get('mapa'):      desc += f"\n📍 Mapa: {apt['mapa']}"
        location = ''
        if apt.get('direccion'):
            location = apt['direccion']
            if apt.get('ciudad'): location += f", {apt['ciudad']}, Ecuador"
            if apt.get('lugar'):  location = f"{apt['lugar']}, {location}"
        elif apt.get('lugar'):
            location = apt['lugar']

    if apt.get('notes'): desc += f"\nNotas: {apt['notes']}"

    event = {
        'summary': f"{apt.get('title', '')} - {apt.get('encargado', '')}",
        'description': desc,
        'start': {'dateTime': apt['start_time'], 'timeZone': 'America/Guayaquil'},
        'end':   {'dateTime': apt['end_time'],   'timeZone': 'America/Guayaquil'},
        'attendees': attendees,
        'reminders': {'useDefault': False, 'overrides': [
            {'method': 'email', 'minutes': 1440},
            {'method': 'popup', 'minutes': 30}]},
    }
    if location: event['location'] = location
    return event


# ============================================================
#  MICROSOFT TOKEN HELPER
# ============================================================
_ms_token_locks = {}
_ms_token_locks_guard = threading.Lock()

def _get_ms_token_lock(token_id):
    with _ms_token_locks_guard:
        lock = _ms_token_locks.get(token_id)
        if lock is None:
            lock = threading.Lock()
            _ms_token_locks[token_id] = lock
        return lock

def _refresh_ms_token(app, t):
    """Refresh a single MS token row. Returns new access_token or None.

    Serializado por fila (id) para que dos hilos no renueven el mismo
    refresh_token en paralelo y se pisen entre sí (MS puede rotarlo en cada uso).
    """
    lock = _get_ms_token_lock(t['id'])
    with lock:
        # Otro hilo pudo haber refrescado esta misma fila mientras esperábamos el lock.
        current = app.supabase.get('ms_tokens', {'id': t['id']}, select='*')
        if current and current[0].get('access_token') != t.get('access_token'):
            return current[0].get('access_token')
        refresh_token = (current[0].get('refresh_token') if current else None) or t.get('refresh_token', '')
        try:
            r = req_lib.post(MS_TOKEN_URL, data={
                'client_id':     app.config.get('MS_CLIENT_ID', ''),
                'client_secret': app.config.get('MS_CLIENT_SECRET', ''),
                'grant_type':    'refresh_token',
                'refresh_token': refresh_token,
                'scope': MS_SCOPES,
            }, timeout=(5, 15))
            if r.status_code != 200:
                return None
            d = r.json()
            new_exp = (datetime.now(timezone.utc)
                       + timedelta(seconds=d.get('expires_in', 3600))).isoformat()
            app.supabase.update('ms_tokens', t['id'], {
                'access_token':  d['access_token'],
                'refresh_token': d.get('refresh_token', refresh_token),
                'expires_at':    new_exp,
            })
            return d['access_token']
        except Exception:
            return None


def get_ms_token(app):
    """Return a valid MS Graph access_token for the first connected account."""
    tokens = app.supabase.get('ms_tokens', select='*')
    if not tokens: return None
    t = tokens[0]
    expiry_str = t.get('expires_at')
    if expiry_str:
        try:
            exp = datetime.fromisoformat(expiry_str.replace('Z', '+00:00'))
            if datetime.now(timezone.utc) >= exp - timedelta(minutes=5):
                return _refresh_ms_token(app, t)
        except Exception:
            pass
    return t.get('access_token')


def get_ms_token_for(app, ms_email):
    """Token válido (refresca si toca) para una cuenta MS específica."""
    rows = app.supabase.get('ms_tokens', {'email': ms_email}, select='*')
    if not rows: return None
    t = rows[0]
    expiry_str = t.get('expires_at')
    if expiry_str:
        try:
            exp = datetime.fromisoformat(expiry_str.replace('Z', '+00:00'))
            if datetime.now(timezone.utc) >= exp - timedelta(minutes=5):
                return _refresh_ms_token(app, t)
        except Exception:
            pass
    return t.get('access_token')


# Mapeos de estado/prioridad para empujar a MS
_TO_MS_STATUS = {'pending':'notStarted','in_progress':'inProgress',
                 'review':'waitingOnOthers','done':'completed','blocked':'deferred'}
_TO_MS_PRIO   = {'low':'low','medium':'normal','high':'high','urgent':'high'}

def _build_ms_task_body(task):
    body = {
        'title':      task.get('title','')[:255],
        'importance': _TO_MS_PRIO.get(task.get('priority'), 'normal'),
        'status':     _TO_MS_STATUS.get(task.get('status','pending'), 'notStarted'),
    }
    if task.get('description'):
        body['body'] = {'content': task['description'], 'contentType':'text'}
    if task.get('due_date'):
        body['dueDateTime'] = {'dateTime': f"{task['due_date']}T23:59:00",
                               'timeZone': 'America/Guayaquil'}
    return body

def push_task_to_ms(app, task):
    """Empuja un cambio del sistema a Microsoft To-Do.
    Devuelve (success: bool, new_source_id: str|None). Si se crea una tarea
    nueva en MS, new_source_id trae el ID asignado por Graph."""
    ms_email = task.get('ms_email'); list_id = task.get('ms_list_id')
    src_id   = task.get('source_id')
    if not (ms_email and list_id): return (False, None)
    token = get_ms_token_for(app, ms_email)
    if not token: return (False, None)
    headers = {'Authorization': f'Bearer {token}','Content-Type':'application/json'}
    try:
        if src_id:
            r = req_lib.patch(f'{MS_GRAPH_URL}/me/todo/lists/{list_id}/tasks/{src_id}',
                              headers=headers, json=_build_ms_task_body(task), timeout=(5,15))
            return (r.status_code in (200, 204), None)
        else:
            r = req_lib.post(f'{MS_GRAPH_URL}/me/todo/lists/{list_id}/tasks',
                             headers=headers, json=_build_ms_task_body(task), timeout=(5,15))
            if r.status_code in (200, 201):
                return (True, r.json().get('id'))
            return (False, None)
    except Exception:
        return (False, None)

def delete_task_in_ms(app, task):
    ms_email = task.get('ms_email'); list_id = task.get('ms_list_id')
    src_id   = task.get('source_id')
    if task.get('source') != 'ms_todo' or not (ms_email and list_id and src_id): return False
    token = get_ms_token_for(app, ms_email)
    if not token: return False
    try:
        r = req_lib.delete(f'{MS_GRAPH_URL}/me/todo/lists/{list_id}/tasks/{src_id}',
                           headers={'Authorization': f'Bearer {token}'}, timeout=(5,15))
        return r.status_code in (200, 204)
    except Exception:
        return False


def get_all_ms_tokens(app):
    """Return list of (email, access_token) for every connected MS account.
    Refreshes expired tokens automatically. Skips accounts whose refresh fails."""
    tokens = app.supabase.get('ms_tokens', select='*')
    out = []
    for t in tokens:
        access = t.get('access_token')
        expiry_str = t.get('expires_at')
        if expiry_str:
            try:
                exp = datetime.fromisoformat(expiry_str.replace('Z', '+00:00'))
                if datetime.now(timezone.utc) >= exp - timedelta(minutes=5):
                    access = _refresh_ms_token(app, t)
            except Exception:
                pass
        if access:
            out.append((t.get('email', 'microsoft'), access))
    return out


# ============================================================
#  DESTINO POR DEFECTO EN MICROSOFT TO-DO
#  Sin esto la sincronización iba en un solo sentido de hecho: lo que se creaba
#  en To-Do bajaba al sistema, pero lo que se creaba en el sistema sólo subía a
#  To-Do si alguien había elegido a mano una cuenta y una lista. Fijando un
#  destino por defecto, toda tarea nacida en el sistema aparece también en To-Do.
# ============================================================
_todo_target_cache = TTLCache(ttl=60)


def get_todo_default_target(app):
    """{'email', 'list_id', 'list_name'} o None si aún no se ha configurado."""
    val, hit = _todo_target_cache.get('target')
    if hit:
        return val
    destino = None
    try:
        filas = app.supabase.get('app_config', {'key': 'todo_default_target'}, select='value')
        if filas and filas[0].get('value'):
            destino = json.loads(filas[0]['value'])
    except Exception as e:
        print(f'[todo] no se pudo leer el destino por defecto: {e}')
    _todo_target_cache.set('target', destino)
    return destino


def set_todo_default_target(app, email, list_id, list_name=''):
    """Guarda (o borra, con email vacío) la lista de To-Do a la que van las
    tareas creadas en el sistema."""
    payload = (json.dumps({'email': email, 'list_id': list_id, 'list_name': list_name})
               if email and list_id else '')
    # app_config NO tiene columna `id`: su clave primaria es `key` (texto). Pedir
    # 'id' devolvía un 400 y la rama de actualización no se ejecutaba nunca, así
    # que guardar la lista por defecto fallaba en silencio y las tareas creadas
    # aquí no subían a Microsoft. Se busca y se actualiza por `key`.
    filas = app.supabase.get('app_config', {'key': 'todo_default_target'}, select='key')
    if filas:
        ok = app.supabase.update('app_config', 'todo_default_target',
                                 {'value': payload}, id_col='key')
    else:
        ok = bool(app.supabase.insert('app_config',
                                      {'key': 'todo_default_target', 'value': payload}))
    _todo_target_cache.invalidate('target')
    return ok


def _parse_iso_dt(s):
    """Parsea una fecha ISO-8601 a datetime UTC consciente; None si falla."""
    if not s: return None
    try:
        dt = datetime.fromisoformat(str(s).replace('Z', '+00:00'))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


# Serializa la sincronización dentro de un mismo worker para evitar que el
# scheduler en segundo plano y un disparo manual (botón admin) inserten la
# misma tarea a la vez y la dupliquen.
_SYNC_LOCK = threading.Lock()

# Filtro de ruido al importar desde Microsoft To-Do (para que el sistema quede
# enfocado en lo accionable y no se llene de miles de tareas históricas):
#   - IMPORT_FLAGGED_EMAILS=False -> no importar la lista "Correos marcados".
#   - IMPORT_COMPLETED=False      -> no importar tareas que YA vienen completadas.
# OJO: las tareas ya existentes en el sistema se SIGUEN actualizando (incluido
# marcarlas como completadas); el filtro solo evita crear ruido nuevo.
IMPORT_FLAGGED_EMAILS = False
IMPORT_COMPLETED      = False


def _prefetch_ms_todo_index(app):
    """Devuelve (dict {source_id: fila}, ok).

    Trae TODAS las tareas ya importadas de Microsoft To-Do para deduplicar. Pagina
    con cabecera Range y usa un timeout amplio. Si la lectura falla (timeout/error),
    ok=False: la sincronización DEBE cancelarse, porque un mapa vacío haría que todas
    las tareas se traten como nuevas y se re-inserten (causa de los duplicados).
    """
    sb = app.supabase
    url = (f"{sb.url}/rest/v1/tasks"
           f"?select=id,source_id,last_synced_at,progress_pct&source=eq.ms_todo")
    out = {}
    step = 1000
    start = 0
    while True:
        headers = {'Range-Unit': 'items', 'Range': f'{start}-{start + step - 1}'}
        try:
            r = sb._session.get(url, headers=headers, timeout=(5, 45))
        except Exception as e:
            print(f'[todo-sync] prefetch error: {e}')
            return {}, False
        if r.status_code not in (200, 206):
            print(f'[todo-sync] prefetch HTTP {r.status_code}: {r.text[:120]}')
            return {}, False
        try:
            rows = r.json()
        except Exception:
            return {}, False
        for row in rows:
            if row.get('source_id'):
                out[row['source_id']] = row
        if len(rows) < step:      # última página
            break
        start += step
    return out, True


def sync_ms_todo(app, accounts, created_by_id, deadline_seconds=90):
    """Sincroniza Microsoft To-Do → Sistema para las cuentas dadas.

    - Inserta las tareas nuevas que aún no existen en el sistema.
    - Actualiza las tareas ya importadas cuando el lado de Microsoft cambió
      después de nuestra última sincronización (política "gana el más reciente":
      compara lastModifiedDateTime de Graph contra last_synced_at local).

    La dirección Sistema → To-Do ya es automática (push_task_to_ms en cada edición),
    así que aquí solo traemos los cambios hechos directamente en Microsoft To-Do.
    Devuelve un dict con totales y detalle por cuenta.
    """
    with _SYNC_LOCK:
        return _sync_ms_todo_locked(app, accounts, created_by_id, deadline_seconds)


def _sync_ms_todo_locked(app, accounts, created_by_id, deadline_seconds=90):
    import time as _time
    DEADLINE = _time.monotonic() + deadline_seconds
    status_map = {
        'notStarted': 'pending', 'inProgress': 'in_progress',
        'completed': 'done', 'waitingOnOthers': 'review', 'deferred': 'blocked'
    }
    prio_map = {'low': 'low', 'normal': 'medium', 'high': 'high'}
    WELLKNOWN_NAME = {
        'flaggedEmails': '📧 Correos marcados',
        'defaultList':   '📌 Tareas (default)',
    }

    # Pre-fetch de las tareas ms_todo existentes: source_id -> fila (insertar/actualizar).
    # CRÍTICO: si no se puede leer el índice de existentes, se CANCELA la sync para no
    # duplicar (un mapa vacío re-insertaría todas las tareas como nuevas).
    existing_by_src, prefetch_ok = _prefetch_ms_todo_index(app)
    if not prefetch_ok:
        return {'success': False,
                'error': ('No se pudo leer las tareas existentes (timeout de BD). '
                          'Sincronización cancelada para evitar duplicados. '
                          'Verifica que existan los índices de la tabla tasks.'),
                'imported': 0, 'updated': 0, 'skipped': 0, 'errors': 0,
                'detail': [], 'partial': True}

    total_imported = 0; total_updated = 0; total_skipped = 0; total_errors = 0
    per_account = []
    sync_iso = datetime.now(timezone.utc).isoformat()
    partial = False

    for ms_email, token in accounts:
        if _time.monotonic() > DEADLINE:
            partial = True
            per_account.append(f'{ms_email}: pendiente (tiempo agotado, reintenta)')
            continue
        headers = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
        imported = 0; updated = 0; skipped = 0; errors = 0
        page_error = False
        r = req_lib.get(f'{MS_GRAPH_URL}/me/todo/lists', headers=headers, timeout=(8,20))
        if r.status_code == 401:
            per_account.append(f'{ms_email}: token expirado, reconecta'); continue
        if r.status_code != 200:
            per_account.append(f'{ms_email}: error {r.status_code}'); continue
        lists = r.json().get('value', [])
        for lst in lists:
            if _time.monotonic() > DEADLINE: partial = True; break
            list_id    = lst['id']
            wk = lst.get('wellknownListName', '')
            list_title = WELLKNOWN_NAME.get(wk, lst.get('displayName', 'To-Do'))
            is_flagged_list = (wk == 'flaggedEmails')
            # Filtro de ruido: saltar por completo la lista de "Correos marcados".
            if is_flagged_list and not IMPORT_FLAGGED_EMAILS:
                continue
            # Import liviano: las subtareas se cargan bajo demanda al abrir cada tarea.
            url = f'{MS_GRAPH_URL}/me/todo/lists/{list_id}/tasks?$top=100&$expand=linkedResources'
            while url:
                if _time.monotonic() > DEADLINE: partial = True; break
                tr = req_lib.get(url, headers=headers, timeout=(10,20))
                if tr.status_code != 200:
                    print(f'[todo-sync] {ms_email}/{list_title}: error HTTP {tr.status_code} paginando tareas')
                    page_error = True
                    partial = True
                    errors += 1
                    break
                tdata = tr.json()
                batch = []
                for task in tdata.get('value', []):
                    title = (task.get('title') or '').strip()
                    if not title: continue
                    tid = task.get('id', '')
                    # Datos comunes (sirven para insertar y para actualizar)
                    lr = (task.get('linkedResources') or [])
                    source_url = ''; source_app = ''
                    if lr:
                        source_url = lr[0].get('webUrl', '') or ''
                        source_app = lr[0].get('applicationName', '') or ''
                    elif is_flagged_list:
                        source_app = 'Outlook'
                    due = None
                    if task.get('dueDateTime'):
                        try: due = task['dueDateTime']['dateTime'][:10]
                        except Exception: pass
                    comp = None
                    if task.get('completedDateTime'):
                        try: comp = task['completedDateTime']['dateTime'][:10]
                        except Exception: pass
                    status  = status_map.get(task.get('status','notStarted'), 'pending')
                    is_done = (task.get('status') == 'completed')

                    existing = existing_by_src.get(tid)
                    if existing:
                        # ¿Microsoft cambió después de nuestra última sync? -> actualizar.
                        ms_mod     = _parse_iso_dt(task.get('lastModifiedDateTime'))
                        local_sync = _parse_iso_dt(existing.get('last_synced_at'))
                        if ms_mod and local_sync and ms_mod <= local_sync:
                            skipped += 1
                            continue
                        # No tocamos campos locales (project_id, assigned_to, notes…):
                        # solo los que son propiedad de Microsoft To-Do.
                        patch = {
                            'title':          title[:300],
                            'description':    (task.get('body') or {}).get('content', '')[:5000],
                            'status':         status,
                            'priority':       prio_map.get(task.get('importance','normal'), 'medium'),
                            'due_date':       due,
                            'completed_date': comp,
                            'source_url':     source_url,
                            'source_app':     source_app,
                            'phase':          (list_title or 'General')[:100],
                            'tags':           f'{list_title} · {ms_email}',
                            'ms_list_id':     list_id,
                            'last_synced_at': sync_iso,
                        }
                        if is_done:
                            patch['progress_pct'] = 100
                        elif (existing.get('progress_pct') or 0) >= 100:
                            patch['progress_pct'] = 0
                        if app.supabase.update('tasks', existing['id'], patch):
                            updated += 1
                            existing['last_synced_at'] = sync_iso
                        else:
                            errors += 1
                        continue

                    # Filtro de ruido: no crear tareas que YA vienen completadas
                    # (las completadas históricas no aportan; las completadas de
                    # tareas ya existentes sí se reflejan en el bloque de arriba).
                    if is_done and not IMPORT_COMPLETED:
                        skipped += 1
                        continue

                    # Tarea nueva -> insertar
                    td = {
                        'title':          title[:300],
                        'description':    (task.get('body') or {}).get('content', '')[:5000],
                        'status':         status,
                        'priority':       prio_map.get(task.get('importance','normal'), 'medium'),
                        'due_date':       due,
                        'completed_date': comp,
                        'tags':           f'{list_title} · {ms_email}',
                        'phase':          (list_title or 'General')[:100],
                        'source':         'ms_todo',
                        'source_id':      tid,
                        'source_url':     source_url,
                        'source_app':     source_app,
                        'ms_email':       ms_email,
                        'ms_list_id':     list_id,
                        'last_synced_at': sync_iso,
                        'created_by':     created_by_id,
                        'progress_pct':   100 if (is_done and comp) else 0,
                        'subtasks':       [],
                    }
                    batch.append(td)
                    existing_by_src[tid] = {'id': None, 'source_id': tid,
                                            'last_synced_at': sync_iso, 'progress_pct': 0}
                # BULK INSERT — un solo POST por página. ON CONFLICT(source_id) ignora
                # las que ya existan (defensa contra carreras/reintentos): nunca duplica.
                if batch:
                    res = app.supabase.insert_on_conflict('tasks', batch, 'source_id')
                    if res is None:
                        errors += len(batch)
                    else:
                        imported += len(res)
                        # Backfill del id real: si el mismo source_id reaparece más
                        # adelante en esta misma corrida, la rama de "existing" de
                        # arriba necesita el id real (no None) para poder actualizarlo.
                        for row in res:
                            src = row.get('source_id')
                            if src and src in existing_by_src:
                                existing_by_src[src]['id'] = row.get('id')
                url = tdata.get('@odata.nextLink')
        summary = f'{ms_email}: +{imported} nuevas, {updated} actualizadas, {skipped} sin cambios'
        if page_error:
            summary += ' (incompleto: error de red/API a mitad de paginación, reintenta)'
        per_account.append(summary)
        total_imported += imported; total_updated += updated
        total_skipped  += skipped;  total_errors  += errors

    return {'success': True, 'imported': total_imported, 'updated': total_updated,
            'skipped': total_skipped, 'errors': total_errors,
            'detail': per_account, 'partial': partial}


def _run_todo_autosync_once(app):
    """Una pasada de sincronización automática (sin sesión de usuario)."""
    try:
        accounts = get_all_ms_tokens(app)
        if not accounts:
            return
        admins = app.supabase.get('users', {'role': 'admin'}, select='id') or []
        created_by = admins[0]['id'] if admins else None
        res = sync_ms_todo(app, accounts, created_by)
        if res.get('imported') or res.get('updated'):
            print(f"[todo-autosync] +{res.get('imported',0)} nuevas, "
                  f"{res.get('updated',0)} actualizadas")
    except Exception as e:
        print(f'[todo-autosync] error: {e}')


def start_todo_autosync(app, interval_min=5):
    """Arranca un hilo que sincroniza To-Do → Sistema cada `interval_min` minutos.

    Con gunicorn (varios workers) usamos un flock para que SOLO un worker corra el
    scheduler y no se dupliquen las llamadas a Graph ni las inserciones. En Windows
    (dev local) no hay fcntl: el scheduler no arranca y se usa el botón manual.
    La dirección Sistema → To-Do sigue siendo inmediata en cada edición.
    """
    try:
        import fcntl
    except Exception:
        print('[todo-autosync] fcntl no disponible (dev local): scheduler desactivado')
        return
    try:
        lock_path = os.path.join(tempfile.gettempdir(), 'todo_autosync.lock')
        lock_file = open(lock_path, 'w')
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
        # Mantener la referencia viva para conservar el lock durante toda la vida del worker
        app._todo_autosync_lock = lock_file
    except Exception:
        # Otro worker ya tiene el lock: este no agenda nada
        return

    def _loop():
        while True:
            time.sleep(interval_min * 60)
            _run_todo_autosync_once(app)

    t = threading.Thread(target=_loop, name='todo-autosync', daemon=True)
    t.start()
    print(f'[todo-autosync] activo (cada {interval_min} min)')


# ============================================================
#  GOOGLE CREDENTIALS
# ============================================================
def get_google_creds(app):
    try:
        tokens = app.supabase.get('google_tokens', {'email': GOOGLE_ACCOUNT_EMAIL})
        if not tokens:
            return None
        t = tokens[0]
        expiry = None
        if t.get('token_expiry'):
            try:
                expiry = datetime.fromisoformat(t['token_expiry'].replace('Z', '+00:00'))
                if expiry.tzinfo is not None:
                    expiry = expiry.astimezone(timezone.utc).replace(tzinfo=None)
            except Exception:
                expiry = None
        creds = Credentials(
            token=t.get('token'), refresh_token=t.get('refresh_token'),
            token_uri='https://oauth2.googleapis.com/token',
            client_id=app.config['GOOGLE_CLIENT_ID'],
            client_secret=app.config['GOOGLE_CLIENT_SECRET'],
            scopes=GOOGLE_SCOPES, expiry=expiry)
        if not creds.refresh_token:
            return None
        needs_refresh = (expiry is None) or creds.expired
        if not needs_refresh and expiry is not None:
            try:
                if expiry <= datetime.utcnow() + timedelta(minutes=5):
                    needs_refresh = True
            except Exception:
                needs_refresh = True
        if needs_refresh:
            try:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                _save_token_fields(app, t['id'], creds)
            except google.auth.exceptions.RefreshError:
                return None
            except Exception:
                return None
        return creds
    except Exception:
        return None

def _save_token_fields(app, token_id, creds):
    data = {'token': creds.token, 'refresh_token': creds.refresh_token}
    if creds.expiry:
        data['token_expiry'] = creds.expiry.isoformat()
    if not app.supabase.update('google_tokens', token_id, data) and creds.expiry:
        app.supabase.update('google_tokens', token_id, {'token': creds.token, 'refresh_token': creds.refresh_token})

def save_google_creds(app, creds):
    """Update-si-existe / insert-si-no, para no dejar la cuenta sin token si el
    guardado falla a mitad de camino (antes se borraba la fila antes de insertar)."""
    existing = app.supabase.get('google_tokens', {'email': GOOGLE_ACCOUNT_EMAIL})
    refresh_token = creds.refresh_token
    if not refresh_token and existing and existing[0].get('refresh_token'):
        refresh_token = existing[0]['refresh_token']
    data = {'email': GOOGLE_ACCOUNT_EMAIL, 'token': creds.token, 'refresh_token': refresh_token}
    if creds.expiry:
        data['token_expiry'] = creds.expiry.isoformat()
    if existing:
        ok = app.supabase.update('google_tokens', existing[0]['id'], data)
        if not ok and creds.expiry:
            data.pop('token_expiry', None)
            ok = app.supabase.update('google_tokens', existing[0]['id'], data)
    else:
        result = app.supabase.insert('google_tokens', data)
        if not result and creds.expiry:
            data.pop('token_expiry', None)
            result = app.supabase.insert('google_tokens', data)
        ok = bool(result)
    if not ok:
        print(f'[save_google_creds] fallo al guardar credenciales de Google para {GOOGLE_ACCOUNT_EMAIL}')
    _google_cache.invalidate_prefix('google_status_')  # bust cache on reconnect
    if ok:
        # Reconexión manual recién hecha: se limpia el estado de avería y se
        # suben las citas que quedaron sin evento mientras estuvo caído.
        _guardar_estado_google(app, {'estado': 'ok', 'error': None, 'notificado_en': None})
        try:
            resincronizar_citas_google(app, creds)
        except Exception as e:
            print(f'[google] resincronización tras reconectar: {e}')
    return ok


# ============================================================
#  RECONEXIÓN AUTOMÁTICA CON GOOGLE CALENDAR
#
#  El refresco del token ya existía, pero era PASIVO: sólo ocurría si alguien
#  entraba a una pantalla que consultara el calendario. Eso deja dos agujeros:
#
#    1. Un token que nadie usa durante meses lo revoca Google por inactividad, y
#       nos enterábamos el día que hacía falta agendar.
#    2. Un corte de red momentáneo se trataba igual que una revocación real: la
#       aplicación decía «Google desconectado» y pedía reconectar a mano cuando
#       en realidad no había pasado nada.
#
#  Este bloque separa la avería transitoria (se reintenta sola) de la permanente
#  (hace falta que una persona vuelva a autorizar), mantiene el token vivo por su
#  cuenta, avisa por notificación cuando de verdad hay que intervenir, y al
#  recuperar la conexión sube solo las citas que se quedaron sin sincronizar.
# ============================================================
GOOGLE_HEALTH_KEY = 'google_health'
_google_health_cache = TTLCache(ttl=20)


def _leer_estado_google(app):
    """{'estado': 'ok'|'reauth'|'transitorio', 'error', 'notificado_en', 'ultimo_ok'}"""
    val, hit = _google_health_cache.get(GOOGLE_HEALTH_KEY)
    if hit:
        return val
    estado = {'estado': 'ok', 'error': None, 'notificado_en': None, 'ultimo_ok': None}
    try:
        filas = app.supabase.get('app_config', {'key': GOOGLE_HEALTH_KEY}, select='value')
        if filas and filas[0].get('value'):
            estado.update(json.loads(filas[0]['value']))
    except Exception:
        pass
    _google_health_cache.set(GOOGLE_HEALTH_KEY, estado)
    return estado


def _guardar_estado_google(app, cambios):
    """Persiste sólo cuando algo cambia de verdad. Se escribe en app_config y no
    en memoria porque el aviso al administrador no debe repetirse una vez por
    cada worker de gunicorn."""
    actual = _leer_estado_google(app)
    nuevo = dict(actual)
    nuevo.update(cambios)
    if nuevo == actual:
        return
    payload = json.dumps(nuevo)
    try:
        # Por `key`, no por `id`: app_config no tiene columna id (ver
        # set_todo_default_target).
        filas = app.supabase.get('app_config', {'key': GOOGLE_HEALTH_KEY}, select='key')
        if filas:
            app.supabase.update('app_config', GOOGLE_HEALTH_KEY,
                                {'value': payload}, id_col='key')
        else:
            app.supabase.insert('app_config', {'key': GOOGLE_HEALTH_KEY, 'value': payload})
    except Exception as e:
        print(f'[google] no se pudo guardar el estado: {e}')
    _google_health_cache.set(GOOGLE_HEALTH_KEY, nuevo)


def _es_avería_permanente(err):
    """True si el fallo exige que una persona vuelva a autorizar la cuenta.

    Un `RefreshError` de Google, o cualquier error cuyo texto traiga
    `invalid_grant`, significa que el permiso ya no existe: reintentarlo mil
    veces no lo va a arreglar. Todo lo demás (DNS, timeout, 5xx de Google) es
    pasajero y sí merece reintento."""
    if isinstance(err, google.auth.exceptions.RefreshError):
        return True
    return _is_invalid_grant(err)


def refrescar_token_google(app, intentos=3, espera_inicial=2):
    """Renueva el token reintentando los fallos pasajeros.

    Devuelve (creds|None, estado, mensaje_error). `estado` es 'ok', 'reauth'
    (hace falta reconectar a mano) o 'transitorio' (falló, pero se reintentará)."""
    try:
        tokens = app.supabase.get('google_tokens', {'email': GOOGLE_ACCOUNT_EMAIL})
    except Exception as e:
        return None, 'transitorio', f'No se pudo leer el token guardado: {str(e)[:120]}'
    if not tokens:
        return None, 'reauth', 'No hay ninguna cuenta de Google conectada'
    fila = tokens[0]
    if not fila.get('refresh_token'):
        return None, 'reauth', 'La cuenta está conectada sin permiso de refresco'

    ultimo_error = ''
    for intento in range(1, intentos + 1):
        try:
            from google.auth.transport.requests import Request
            creds = Credentials(
                token=fila.get('token'), refresh_token=fila.get('refresh_token'),
                token_uri='https://oauth2.googleapis.com/token',
                client_id=app.config['GOOGLE_CLIENT_ID'],
                client_secret=app.config['GOOGLE_CLIENT_SECRET'],
                scopes=GOOGLE_SCOPES)
            creds.refresh(Request())
            _save_token_fields(app, fila['id'], creds)
            _google_cache.invalidate_prefix('google_status_')
            return creds, 'ok', None
        except Exception as e:
            ultimo_error = str(e)[:200]
            if _es_avería_permanente(e):
                return None, 'reauth', ultimo_error
            if intento < intentos:
                time.sleep(espera_inicial * intento)   # 2 s, 4 s, 6 s…
    return None, 'transitorio', ultimo_error


def resincronizar_citas_google(app, creds, limite_segundos=60):
    """Sube a Google las citas confirmadas que se quedaron sin evento.

    Mientras Google está caído, aprobar una cita la guarda en el sistema pero no
    crea el evento. Sin este paso esas citas quedaban invisibles en el calendario
    para siempre, salvo que un administrador se acordara de pulsar el botón de
    sincronizar. Ahora se hace solo en cuanto vuelve la conexión."""
    if not creds:
        return {'subidas': 0, 'errores': 0, 'pendientes': 0}
    try:
        citas = app.supabase.get('appointments',
            select='id,title,encargado,tema,client_name,start_time,end_time,calendar_id,'
                   'invitados,direccion,ciudad,lugar,mapa,notes,meeting_link,status,'
                   'google_event_id') or []
    except Exception as e:
        print(f'[google] no se pudieron leer las citas: {e}')
        return {'subidas': 0, 'errores': 0, 'pendientes': 0}

    pendientes = [c for c in citas
                  if c.get('status') == 'confirmed' and not c.get('google_event_id')]
    if not pendientes:
        return {'subidas': 0, 'errores': 0, 'pendientes': 0}

    email_map, gcal_id_map = _make_cal_maps(_get_calendar_config(app))
    service = build('calendar', 'v3', credentials=creds)
    limite = time.monotonic() + limite_segundos
    subidas, errores = 0, 0

    for cita in pendientes:
        if time.monotonic() > limite:
            break
        gcal_id = gcal_id_map.get(cita.get('calendar_id'), 'primary')
        try:
            # Puede que el evento sí se creara y sólo se perdiera el id: se busca
            # antes de insertar para no duplicar la cita en el calendario.
            existente = service.events().list(
                calendarId=gcal_id, timeMin=cita['start_time'], timeMax=cita['end_time'],
                q=cita.get('title') or '', maxResults=1).execute()
            if existente.get('items'):
                app.supabase.update('appointments', cita['id'], {
                    'google_event_id': existente['items'][0]['id'], 'google_cal_id': gcal_id})
                continue
            evento = _build_google_event(cita, _build_attendees(cita, email_map))
            creado = service.events().insert(calendarId=gcal_id, body=evento,
                                             sendUpdates='all').execute()
            app.supabase.update('appointments', cita['id'], {
                'google_event_id': creado.get('id'), 'google_cal_id': gcal_id})
            subidas += 1
        except Exception as e:
            if _es_avería_permanente(e):
                break        # se volvió a caer: lo retoma el siguiente ciclo
            errores += 1
    if subidas:
        print(f'[google] resincronizadas {subidas} cita(s) pendientes')
    return {'subidas': subidas, 'errores': errores, 'pendientes': len(pendientes)}


def _avisar_reconexion_google(app, mensaje):
    """Notifica a los administradores UNA vez por avería, no en cada ciclo."""
    estado = _leer_estado_google(app)
    if estado.get('notificado_en'):
        return 0
    enviados = 0
    try:
        for admin in (app.supabase.get('users', {'role': 'admin'}, select='id') or []):
            enviados += send_push_to_user(
                app, admin['id'],
                '⚠️ Google Calendar necesita reconexión',
                'El permiso caducó o fue revocado. Las citas se aprueban igual, '
                'pero no llegan al calendario hasta reconectar.',
                '/auth/google')
    except Exception as e:
        print(f'[google] no se pudo avisar a los administradores: {e}')
    _guardar_estado_google(app, {
        'estado': 'reauth', 'error': mensaje,
        'notificado_en': datetime.now(timezone.utc).isoformat()})
    return enviados


def google_autoreconectar(app):
    """Un ciclo completo: refrescar, y si vuelve la conexión, poner al día.

    Es lo que llaman el hilo de fondo y la ruta de cron."""
    creds, estado, error = refrescar_token_google(app)
    previo = _leer_estado_google(app).get('estado')

    if estado == 'ok':
        _guardar_estado_google(app, {
            'estado': 'ok', 'error': None, 'notificado_en': None,
            'ultimo_ok': datetime.now(timezone.utc).isoformat()})
        resultado = resincronizar_citas_google(app, creds)
        if previo != 'ok':
            print('[google] conexión recuperada automáticamente')
        return {'success': True, 'estado': 'ok', 'recuperado': previo != 'ok', **resultado}

    if estado == 'reauth':
        avisados = _avisar_reconexion_google(app, error)
        return {'success': False, 'estado': 'reauth', 'error': error, 'avisados': avisados}

    # Transitorio: no se toca `notificado_en` ni se molesta a nadie; el próximo
    # ciclo lo reintenta. Marcarlo como avería aquí sería mentir sobre el estado.
    _guardar_estado_google(app, {'estado': 'transitorio', 'error': error})
    return {'success': False, 'estado': 'transitorio', 'error': error}


def start_google_autoheal(app, interval_min=60):
    """Mantiene vivo el token de Google en segundo plano.

    Mismo patrón que el autosync de To-Do: con gunicorn sólo un worker toma el
    flock y agenda; en Windows (sin fcntl) no arranca y queda el cron externo o
    el botón manual."""
    try:
        import fcntl
    except Exception:
        print('[google-autoheal] fcntl no disponible (dev local): scheduler desactivado')
        return
    try:
        ruta = os.path.join(tempfile.gettempdir(), 'google_autoheal.lock')
        archivo = open(ruta, 'w')
        fcntl.flock(archivo, fcntl.LOCK_EX | fcntl.LOCK_NB)
        app._google_autoheal_lock = archivo
    except Exception:
        return          # otro worker ya lo tiene

    def _bucle():
        # Primer ciclo a los dos minutos: da tiempo a que el despliegue termine
        # de levantar antes de salir a la red.
        time.sleep(120)
        while True:
            try:
                google_autoreconectar(app)
            except Exception as e:
                print(f'[google-autoheal] {e}')
            time.sleep(interval_min * 60)

    threading.Thread(target=_bucle, name='google-autoheal', daemon=True).start()
    print(f'[google-autoheal] activo (cada {interval_min} min)')


# ============================================================
#  CALENDAR ACCESS (with caching)
# ============================================================
def _get_calendar_config(app):
    """Cached calendar_config (5 min)."""
    val, hit = _cal_cache.get('all')
    if hit:
        return val
    result = app.supabase.get('calendar_config', select='calendar_id,name,email,color,google_cal_id')
    _cal_cache.set('all', result)
    return result

def _make_cal_maps(all_cals):
    """Build two maps from calendar_config list.
    Returns (email_map, gcal_id_map):
      email_map   — calendar_id → contact email (attendee)
      gcal_id_map — calendar_id → Google Calendar ID to use for events
    """
    email_map   = {c['calendar_id']: c['email']
                   for c in all_cals if c.get('email')}
    gcal_id_map = {c['calendar_id']: (c.get('google_cal_id') or 'primary')
                   for c in all_cals}
    return email_map, gcal_id_map

def _build_attendees(apt, email_map):
    """Build a deduplicated attendee list for a Google Calendar event.
    Uses lowercase comparison to avoid case-sensitive duplicates.
    """
    seen = set()
    attendees = []
    def _add(email):
        e = (email or '').strip().lower()
        if e and e not in seen:
            seen.add(e)
            attendees.append({'email': email.strip()})
    cal_email = email_map.get(apt.get('calendar_id', ''))
    if cal_email:
        _add(cal_email)
    if apt.get('invitados'):
        for inv in apt['invitados'].split(','):
            _add(inv)
    if not attendees:
        _add(GOOGLE_ACCOUNT_EMAIL)
    return attendees

def is_admin():
    """Admin EFECTIVO (memoizado por request). Depende del ROL ACTIVO: si el
    nivel del rol activo es 'administrador', el usuario tiene poderes totales;
    con cualquier otro nivel queda limitado a los grants de ese rol. Esto es lo
    que permite a un usuario alternar entre, p.ej., un rol Socio (limitado) y un
    rol Administrador (total) desde el selector de rol.

    Salvaguarda (bootstrap): un usuario marcado users.role=='admin' que todavía
    NO tiene roles de negocio asignados conserva admin, para no dejar el sistema
    sin ningún administrador tras la migración."""
    if not current_user.is_authenticated:
        return False
    if hasattr(g, '_is_admin'):
        return g._is_admin
    try:
        roles = get_user_roles(current_app, current_user.id)
        if not roles:
            result = (current_user.role == 'admin')
        else:
            result = (get_active_role_grants(current_app, current_user.id).get('level') == 'administrador')
    except Exception:
        result = (current_user.role == 'admin')
    g._is_admin = result
    return result

def browser_sync_allowed():
    """Acceso a la sincronización de navegadores: sólo el administrador dueño,
    y sólo si está habilitado en esta máquina (BROWSER_SYNC_ENABLED en el .env local)."""
    if not current_app.config.get('BROWSER_SYNC_ENABLED'):
        return False
    if not (current_user.is_authenticated and current_user.role == 'admin'):
        return False
    owner = current_app.config.get('BROWSER_SYNC_OWNER_EMAIL', '')
    return bool(owner) and (current_user.email or '').lower() == owner

# ============================================================
#  ROLES MULTIPLES — un usuario puede tener varios roles de negocio
#  (ej. "Encargado de Cuenca", "Auditor"), cada uno con su propio paquete de
#  modulos + calendarios + proyectos + cuentas MS. Reemplaza el modelo directo
#  de users.modules/calendar_permissions/ms_account_permissions como fuente de
#  autorizacion (esas tablas se conservan como respaldo, ver migrations/014-015).
# ============================================================
def get_user_roles(app, uid):
    """Roles asignados a un usuario: [{id, name, description}]."""
    val, hit = _user_roles_cache.get(uid)
    if hit:
        return val
    rows = app.supabase.get('user_roles', {'user_id': uid}, select='role_id')
    role_ids = [r['role_id'] for r in (rows or [])]
    result = app.supabase.get_in('roles', 'id', role_ids, select='id,name,description') if role_ids else []
    _user_roles_cache.set(uid, result)
    return result

def role_grants(app, role_id):
    """{modules, calendar_ids, project_ids, ms_emails, task_ids, narrowed_projects}
    otorgados por un rol.

    task_ids / narrowed_projects modelan el acceso a ACTIVIDADES (tareas):
      - narrowed_projects: proyectos en los que el rol tiene al menos una tarea
        marcada -> dentro de esos proyectos solo se ven las tareas de task_ids.
      - Un proyecto que NO está en narrowed_projects se ve completo (todas sus
        tareas), que es el comportamiento previo a esta función."""
    val, hit = _role_cache.get(role_id)
    if hit:
        return val
    role = app.supabase.get('roles', {'id': role_id}, select='modules,level')
    modules = [m.strip() for m in (role[0].get('modules') or '').split(',') if m.strip()] if role else []
    level = ((role[0].get('level') if role else None) or DEFAULT_ROLE_LEVEL)
    cals  = app.supabase.get('role_calendars',   {'role_id': role_id}, select='calendar_id')
    projs = app.supabase.get('role_projects',    {'role_id': role_id}, select='project_id')
    msacc = app.supabase.get('role_ms_accounts', {'role_id': role_id}, select='ms_email')
    tasks = app.supabase.get('role_tasks',       {'role_id': role_id}, select='task_id,project_id')
    result = {
        'level':        level,
        'modules':      modules,
        'calendar_ids': {c['calendar_id'] for c in (cals or [])},
        'project_ids':  {p['project_id']  for p in (projs or [])},
        'ms_emails':    {m['ms_email']    for m in (msacc or [])},
        'task_ids':          {t['task_id']    for t in (tasks or [])},
        'narrowed_projects': {t['project_id'] for t in (tasks or []) if t.get('project_id')},
    }
    _role_cache.set(role_id, result)
    return result

def _save_role_activities(app, role_id, task_ids, replace=False):
    """Persiste las actividades (tareas) marcadas para un rol en role_tasks,
    guardando project_id denormalizado. Si replace=True borra las previas."""
    if replace:
        for row in app.supabase.get('role_tasks', {'role_id': role_id}, select='id'):
            app.supabase.delete('role_tasks', row['id'])
    task_ids = [t for t in (task_ids or []) if t]
    if not task_ids:
        return
    rows = app.supabase.get_in('tasks', 'id', task_ids, select='id,project_id') or []
    pid_by_task = {r['id']: r.get('project_id') for r in rows}
    for tid in task_ids:
        if tid not in pid_by_task:        # tarea inexistente/borrada -> se ignora
            continue
        app.supabase.insert_ignore('role_tasks', {
            'role_id': role_id, 'task_id': tid, 'project_id': pid_by_task[tid]})

def get_active_role_id(app, uid):
    """Rol activo del usuario, memoizado por request. Si la elección guardada
    ya no le pertenece (rol borrado/reasignado), cae al primero de sus roles."""
    if hasattr(g, '_active_role_id'):
        return g._active_role_id
    roles = get_user_roles(app, uid)
    if not roles:
        g._active_role_id = None
        return None
    role_ids = {r['id'] for r in roles}
    choice = session.get('active_role_id')
    if choice not in role_ids:
        choice = roles[0]['id']
        session['active_role_id'] = choice
    g._active_role_id = choice
    return choice

def get_active_role_grants(app, uid):
    rid = get_active_role_id(app, uid)
    if not rid:
        return {'level': None, 'modules': [], 'calendar_ids': set(), 'project_ids': set(),
                'ms_emails': set(), 'task_ids': set(), 'narrowed_projects': set()}
    return role_grants(app, rid)

# ============================================================
#  LO CONCEDIDO A UNA PERSONA, AL MARGEN DE SU ROL
#
#  El rol es la base y responde bien a «todas las secretarias ven esto». Lo que
#  no sabía decir el sistema es «además, ELLA ve el calendario de JOMAP», y la
#  única salida era inventarle un rol entero para una excepción — de ahí el rol
#  «Acceso - {nombre}» que el propio sistema genera al aprobar una solicitud: un
#  parche que confiesa que faltaba esta capa.
#
#  Regla única, y por eso todo pasa por `grants_efectivos`: lo de la persona se
#  SUMA a lo del rol, nunca lo resta.
# ============================================================
_user_grants_cache = TTLCache(ttl=30, maxsize=256)

# tabla -> (columna del recurso, clave en el diccionario de grants)
_TABLAS_CONCESION = (
    ('user_modules',     'modulo',      'modules'),
    ('user_calendars',   'calendar_id', 'calendar_ids'),
    ('user_projects',    'project_id',  'project_ids'),
    ('user_ms_accounts', 'ms_email',    'ms_emails'),
)


def get_user_grants(app, uid):
    """Módulos, calendarios, proyectos y cuentas MS concedidos a ESTA persona.

    Las tablas pueden no existir todavía (migraciones 027 y 028 sin aplicar).
    En ese caso se devuelven conjuntos vacíos y el sistema se comporta como
    antes, gobernado sólo por el rol: una capa que se añade no puede tumbar la
    que ya funcionaba."""
    val, hit = _user_grants_cache.get(str(uid))
    if hit:
        return val
    concedido = {}
    for tabla, columna, clave in _TABLAS_CONCESION:
        try:
            filas = app.supabase.get(tabla, {'user_id': uid}, select=columna) or []
            concedido[clave] = {f[columna] for f in filas if f.get(columna)}
        except Exception:
            concedido[clave] = set()
    _user_grants_cache.set(str(uid), concedido)
    return concedido


def grants_efectivos(app, uid):
    """Lo que esta persona puede alcanzar: su ROL ACTIVO más lo suyo propio.

    Es el único sitio donde se combinan las dos fuentes. Todo lo que autoriza
    —módulos, calendarios, proyectos, cuentas de Microsoft y la visibilidad de
    las tareas— pasa por aquí, para que no puedan discrepar entre sí: un menú
    que enseña un módulo que luego no deja entrar es peor que no enseñarlo."""
    rol  = get_active_role_grants(app, uid)
    mio  = get_user_grants(app, uid)
    proyectos_propios = mio['project_ids']
    return {
        'level':        rol['level'],
        'modules':      set(rol['modules']) | mio['modules'],
        'calendar_ids': rol['calendar_ids'] | mio['calendar_ids'],
        'project_ids':  rol['project_ids']  | proyectos_propios,
        'ms_emails':    rol['ms_emails']    | mio['ms_emails'],
        'task_ids':     rol['task_ids'],
        # Un proyecto concedido a la persona se ve ENTERO. El recorte por
        # actividades pertenece al rol; lo que el administrador da aparte, a
        # mano y a alguien concreto, no debe llegar recortado por una regla
        # escrita pensando en otros.
        'narrowed_projects': rol['narrowed_projects'] - proyectos_propios,
    }


def get_user_calendars(app, uid):
    """Calendarios visibles: los del rol activo más los concedidos a la persona
    (admins ven todos)."""
    if is_admin():
        return _get_calendar_config(app)
    permitidos = grants_efectivos(app, uid)['calendar_ids']
    return [c for c in _get_calendar_config(app) if c['calendar_id'] in permitidos]

def user_has_calendar_access(app, uid, calendar_id):
    if is_admin():
        return True
    return calendar_id in grants_efectivos(app, uid)['calendar_ids']

def get_user_projects(app, uid):
    """Proyectos visibles: los del rol activo más los concedidos a la persona
    (admins ven todos)."""
    if is_admin():
        return app.supabase.get('projects', select='*') or []
    ids = list(grants_efectivos(app, uid)['project_ids'])
    return app.supabase.get_in('projects', 'id', ids, select='*') if ids else []

def user_has_project_access(app, uid, project_id):
    if is_admin():
        return True
    return project_id in grants_efectivos(app, uid)['project_ids']

def _grant_calendar_via_role(app, uid, calendar_id):
    """Aprobar una solicitud de /register ya no alcanza con marcar calendar_permissions
    como 'approved' (esa tabla dejó de ser autoritativa) — crea o extiende un rol
    personal del usuario ("Acceso - {nombre}") que le da el módulo 'calendar' + ese
    calendario, y se lo asigna, para que 'Aprobar' siga otorgando acceso real."""
    user = app.supabase.get('users', {'id': uid}, select='full_name')
    full_name = user[0]['full_name'] if user else uid
    role_name = f'Acceso - {full_name}'
    role_id = None
    for ur in (app.supabase.get('user_roles', {'user_id': uid}, select='role_id') or []):
        r = app.supabase.get('roles', {'id': ur['role_id']}, select='id,name,modules,created_by')
        # Coincide por nombre Y por haber sido generado por este mismo flujo --
        # evita reusar/mutar un rol creado a mano por el admin que casualmente
        # tenga el mismo nombre "Acceso - {nombre}".
        if r and r[0]['name'] == role_name and r[0].get('created_by') == 'admin_approve':
            role_id = r[0]['id']
            mods = {m for m in (r[0].get('modules') or '').split(',') if m}
            if 'calendar' not in mods:
                mods.add('calendar')
                app.supabase.update('roles', role_id, {'modules': ','.join(mods)})
            break
    if not role_id:
        created = app.supabase.insert('roles', {
            'name': role_name,
            'description': 'Rol generado al aprobar una solicitud de acceso a calendario.',
            'modules': 'calendar', 'created_by': 'admin_approve'})
        role_id = created[0]['id'] if created else None
        if role_id:
            app.supabase.insert('user_roles', {'user_id': uid, 'role_id': role_id})
    if role_id:
        app.supabase.insert_ignore('role_calendars', {'role_id': role_id, 'calendar_id': calendar_id})
        _role_cache.invalidate(role_id)
        _user_roles_cache.invalidate(uid)

def csrf_protect(view):
    """Exige un csrf_token válido (form field o header X-CSRFToken) en rutas
    admin sensibles. No se aplica globalmente porque el resto de la app usa
    fetch()/JSON sin token."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        token = request.form.get('csrf_token') or request.headers.get('X-CSRFToken')
        try:
            validate_csrf(token)
        except ValidationError:
            if request.is_json or request.headers.get('X-CSRFToken') is not None:
                return jsonify({'success': False, 'error': 'Sesión de seguridad expirada, recarga la página.'}), 400
            flash('La sesión de seguridad expiró o la solicitud no es válida. Intenta de nuevo.', 'danger')
            return redirect(request.referrer or '/dashboard')
        return view(*args, **kwargs)
    return wrapped

def get_user_ms_emails(app, uid):
    """Cuentas MS autorizadas: las del rol activo más las concedidas a la
    persona (admins ven todas). Son las que deciden qué listas de To-Do ve."""
    if is_admin():
        return [t['email'] for t in (app.supabase.get('ms_tokens', select='email') or []) if t.get('email')]
    return list(grants_efectivos(app, uid)['ms_emails'])

def _delete_user_cascade(app, uid):
    """Borra las filas relacionadas (permisos, credenciales) antes del usuario,
    para no dejar huérfanas en calendar_permissions/webauthn_credentials/etc."""
    for tbl in ['calendar_permissions', 'webauthn_credentials', 'face_descriptors',
                'ms_account_permissions', 'user_roles', 'user_permissions',
                'user_modules', 'user_calendars', 'user_projects', 'user_ms_accounts']:
        for row in app.supabase.get(tbl, {'user_id': uid}, select='id'):
            app.supabase.delete(tbl, row['id'])
    ok = app.supabase.delete('users', uid)
    _user_cal_cache.invalidate(uid)
    _user_grants_cache.invalidate(str(uid))
    _user_perms_cache.invalidate(str(uid))
    return ok

def _project_allowed(app, task, uid):
    """Alcance de una tarea. Combina dos capas:

      1) Proyecto: si la tarea tiene project_id, ese proyecto debe estar entre
         los del rol activo o entre los concedidos a la persona (las tareas
         manuales sueltas sin project_id no se ocultan — es aditivo, no
         retroactivo).
      2) Actividad: si el ROL tiene actividades marcadas en ESE proyecto
         (narrowed_projects), sólo se ve la tarea si está en task_ids. Un
         proyecto sin actividades marcadas se ve completo, y también el que se
         concedió directamente a la persona: ese recorte es del rol.

    Admins bypasean vía is_admin() en el llamador."""
    pid = task.get('project_id')
    if not pid:
        return True
    grants = grants_efectivos(app, uid)
    if pid not in grants['project_ids']:
        return False
    if pid in grants['narrowed_projects']:
        return task.get('id') in grants['task_ids']
    return True

def _filter_visible_tasks(app, rows, uid):
    """Misma regla de visibilidad usada en GET /planning/api/tasks: admins ven todo;
    el resto solo tareas MS de cuentas autorizadas o tareas manuales propias/asignadas,
    y si la tarea pertenece a un proyecto, ese proyecto debe estar autorizado."""
    if is_admin():
        return rows
    allowed_ms = set(get_user_ms_emails(app, uid))
    grants = grants_efectivos(app, uid)
    has_todo = 'todo' in grants['modules']
    has_plan = 'planning' in grants['modules']
    suid = str(uid)
    def visible(t):
        if t.get('source') == 'ms_todo':
            if not has_todo: return False
            if (t.get('ms_email') or '') not in allowed_ms: return False
            return _project_allowed(app, t, uid)
        if not has_plan: return False
        owns = (t.get('created_by') == suid or t.get('assigned_to') == suid or
                (t.get('assigned_email') or '').lower() == (current_user.email or '').lower())
        if not owns: return False
        return _project_allowed(app, t, uid)
    return [t for t in rows if visible(t)]

def _user_owns_task(app, task, uid):
    """Misma regla de propiedad usada en planning_bulk_update: admins la evitan (is_admin() aparte)."""
    if task.get('source') == 'ms_todo':
        if (task.get('ms_email') or '') not in set(get_user_ms_emails(app, uid)):
            return False
        return _project_allowed(app, task, uid)
    owns = (task.get('created_by') == str(uid) or task.get('assigned_to') == str(uid) or
            (task.get('assigned_email') or '').lower() == (current_user.email or '').lower())
    if not owns:
        return False
    return _project_allowed(app, task, uid)

# ============================================================
#  WEB PUSH (VAPID + notificaciones)
# ============================================================
def _b64url(b):
    return base64.urlsafe_b64encode(b).rstrip(b'=').decode('ascii')

def get_vapid_keys(app):
    """Devuelve (public_pem, private_pem, public_b64url_uncompressed_point).
    Si no existen en app_config las genera y guarda."""
    if not WEB_PUSH_AVAILABLE: return None
    rows = app.supabase.get('app_config', {'key': 'vapid'}, select='value')
    if rows and rows[0].get('value'):
        try:
            d = json.loads(rows[0]['value'])
            return d.get('public_pem'), d.get('private_pem'), d.get('public_b64')
        except Exception:
            pass
    # Generar
    v = Vapid()
    v.generate_keys()
    public_pem  = v.public_pem().decode('utf-8')
    private_pem = v.private_pem().decode('utf-8')
    raw_pub = v.public_key.public_numbers().x.to_bytes(32,'big') + v.public_key.public_numbers().y.to_bytes(32,'big')
    public_b64 = _b64url(b'\x04' + raw_pub)
    payload = json.dumps({'public_pem': public_pem, 'private_pem': private_pem,
                          'public_b64': public_b64})
    app.supabase.insert('app_config', {'key': 'vapid', 'value': payload})
    return public_pem, private_pem, public_b64

def send_push_to_user(app, user_id, title, body, url='/dashboard'):
    """Envía push a TODAS las suscripciones activas del usuario."""
    if not WEB_PUSH_AVAILABLE: return 0
    keys = get_vapid_keys(app)
    if not keys: return 0
    _, private_pem, _ = keys
    subs = app.supabase.get('web_push_subscriptions', {'user_id': user_id}, select='*')
    sent = 0
    payload = json.dumps({'title': title, 'body': body, 'url': url})
    for s in (subs or []):
        try:
            webpush(
                subscription_info={
                    'endpoint': s['endpoint'],
                    'keys': {'p256dh': s['p256dh'], 'auth': s['auth']},
                },
                data=payload,
                vapid_private_key=private_pem,
                vapid_claims={'sub': 'mailto:noreply@calendarios-map.com'},
            )
            sent += 1
        except WebPushException as e:
            # 410 = suscripción expirada
            if e.response and e.response.status_code in (404, 410):
                app.supabase.delete('web_push_subscriptions', s['id'])
        except Exception:
            pass
    return sent


def get_user_permissions(app, uid):
    """Permisos sueltos concedidos a UNA persona: {'directorio.importar', ...}"""
    val, hit = _user_perms_cache.get(str(uid))
    if hit:
        return val
    try:
        filas = app.supabase.get('user_permissions', {'user_id': uid}, select='permiso') or []
        permisos = {f['permiso'] for f in filas if f.get('permiso')}
    except Exception:
        # La tabla puede no existir aún (migración 023 sin aplicar): sin permisos
        # sueltos el sistema sigue funcionando con los del rol.
        permisos = set()
    _user_perms_cache.set(str(uid), permisos)
    return permisos


def modulos_efectivos(app, uid):
    """Módulos a los que entra esta persona: los de su ROL ACTIVO más los que el
    administrador le haya concedido a ella en particular (ver `grants_efectivos`)."""
    if is_admin():
        return {m[0] for m in ALL_MODULES}
    return grants_efectivos(app, uid)['modules']


def user_can(permiso):
    """¿Puede el usuario actual hacer esto?

    Acepta dos formas:
      * 'directorio'           -> ¿tiene el módulo (por rol o concedido a él)?
      * 'directorio.importar'  -> ¿puede además esa acción concreta?

    Una acción NO sensible sólo exige el módulo. Una SENSIBLE exige además que
    se le haya concedido a esa persona, para poder dar acceso a un módulo sin
    regalar de paso el borrado o la importación masiva."""
    if is_admin():
        return True
    modulo, _, accion = permiso.partition('.')
    if modulo not in modulos_efectivos(current_app, current_user.id):
        return False
    if not accion:
        return True
    if permiso not in ACCIONES_SENSIBLES:
        return True          # acción corriente: basta con tener el módulo
    return permiso in get_user_permissions(current_app, current_user.id)


# ============================================================
#  APPOINTMENT BUILDER
# ============================================================
def _build_appointment(title, cal_id, encargado, tema, client_name, client_email,
                        start_dt, end_dt, tipo, link, lugar, direccion, mapa,
                        ciudad, notificar, notes, user_id):
    return {
        'title': title, 'calendar_id': cal_id, 'encargado': encargado, 'tema': tema,
        'client_name': client_name, 'client_email': client_email,
        'start_time': start_dt.isoformat(), 'end_time': end_dt.isoformat(),
        'status': 'pending', 'notes': notes,
        'invitados': ','.join(notificar) if notificar else '',
        'lugar': lugar, 'direccion': direccion, 'mapa': mapa, 'ciudad': ciudad,
        'meeting_link': link if tipo == 'virtual' else '',
        'created_by': user_id,
    }


# ============================================================
#  RECURRENCE — flexible occurrence generator (materialized)
# ============================================================
# Tope de seguridad: nº máximo de eventos materializados por serie.
# Para recurrencia "indefinida" se materializa hasta este tope.
REC_HARD_CAP = 366


def _add_months(d, months):
    """Suma `months` a la fecha `d`. Devuelve None si el día no existe
    en el mes destino (ej. 31 en un mes de 30 días) — se omite la ocurrencia."""
    m = d.month - 1 + months
    y = d.year + m // 12
    m = m % 12 + 1
    last = _cal.monthrange(y, m)[1]
    if d.day > last:
        return None
    return date(y, m, d.day)


def _generate_recurrence_dates(start_d, freq, interval, weekdays,
                               end_mode, end_date, count, cap=REC_HARD_CAP):
    """Genera la lista de fechas de una serie recurrente.

    freq:     'daily' | 'weekly' | 'monthly' | 'yearly'
    interval: cada N (días/semanas/meses/años), entero >= 1
    weekdays: lista de días [0=Lun..6=Dom] — solo para 'weekly'
    end_mode: 'until' (hasta end_date) | 'count' (N ocurrencias) | 'forever'
    """
    interval = max(1, int(interval or 1))
    count = max(1, int(count or 1))
    out = []

    def _reached_limit():
        if end_mode == 'count' and len(out) >= count:
            return True
        return len(out) >= cap

    if freq == 'daily':
        k = 0
        while len(out) < cap:
            d = start_d + timedelta(days=k * interval)
            if end_mode == 'until' and d > end_date:
                break
            out.append(d)
            if _reached_limit():
                break
            k += 1

    elif freq == 'weekly':
        wds = sorted(set(weekdays)) if weekdays else [start_d.weekday()]
        start_monday = start_d - timedelta(days=start_d.weekday())
        wk = 0
        stop = False
        while len(out) < cap and not stop:
            week_start = start_monday + timedelta(weeks=wk * interval)
            for wd in wds:
                d = week_start + timedelta(days=wd)
                if d < start_d:
                    continue
                if end_mode == 'until' and d > end_date:
                    stop = True
                    break
                out.append(d)
                if _reached_limit():
                    stop = True
                    break
            wk += 1

    elif freq == 'monthly':
        k = 0
        guard = 0
        while len(out) < cap and guard < cap * 3:
            guard += 1
            ref = _add_months(start_d.replace(day=1), k * interval)  # 1° del mes, siempre válido
            if end_mode == 'until' and ref is not None and ref > end_date:
                break
            d = _add_months(start_d, k * interval)
            if d is not None and not (end_mode == 'until' and d > end_date):
                out.append(d)
                if _reached_limit():
                    break
            k += 1

    elif freq == 'yearly':
        k = 0
        guard = 0
        while len(out) < cap and guard < cap * 3:
            guard += 1
            yr = start_d.year + k * interval
            if end_mode == 'until' and date(yr, 1, 1) > end_date:
                break
            try:
                d = date(yr, start_d.month, start_d.day)  # 29-feb se omite en años no bisiestos
            except ValueError:
                d = None
            if d is not None and not (end_mode == 'until' and d > end_date):
                out.append(d)
                if _reached_limit():
                    break
            k += 1

    return out[:cap]


# ============================================================
#  APP FACTORY
# ============================================================
def create_app():
    app = Flask(__name__, template_folder='../templates', static_folder='../static')
    app.config.from_object(Config)
    if not app.config['SECRET_KEY']:
        raise RuntimeError(
            'SECRET_KEY no está configurada. Define la variable de entorno SECRET_KEY '
            '(clave aleatoria y secreta) antes de arrancar la aplicación.'
        )
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

    try:
        app.supabase = SupabaseAPI(app.config['SUPABASE_URL'], app.config['SUPABASE_KEY'])
        print('Supabase OK')
    except Exception as e:
        print(f'Supabase error: {e}'); app.supabase = None

    login_manager.init_app(app)
    login_manager.login_view = 'login'
    limiter.init_app(app)
    app.jinja_env.globals['csrf_token'] = generate_csrf

    @app.context_processor
    def _inject_globals():
        return {'webauthn_available': WEBAUTHN_AVAILABLE,
                'face_login_enabled': True,
                'browser_sync_visible': browser_sync_allowed(),
                # Admin EFECTIVO según el rol activo (no el flag fijo del sistema),
                # para que la interfaz se limite/habilite al alternar de rol.
                'is_effective_admin': is_admin()}

    # ------ PWA: service worker / manifest / offline desde la raíz ------
    @app.route('/sw.js')
    def pwa_service_worker():
        resp = send_from_directory(app.static_folder, 'sw.js',
                                   mimetype='application/javascript')
        resp.headers['Service-Worker-Allowed'] = '/'
        resp.headers['Cache-Control'] = 'no-cache'
        return resp

    @app.route('/manifest.webmanifest')
    def pwa_manifest():
        return send_from_directory(app.static_folder, 'manifest.webmanifest',
                                   mimetype='application/manifest+json')

    @app.route('/offline.html')
    def pwa_offline():
        return send_from_directory(app.static_folder, 'offline.html')

    # ------ Digital Asset Links: vincula la app TWA de Google Play con la web ------
    @app.route('/.well-known/assetlinks.json')
    def well_known_assetlinks():
        return send_from_directory(app.static_folder, 'assetlinks.json',
                                   mimetype='application/json')

    @login_manager.user_loader
    def load_user(uid):
        if app.supabase:
            u = app.supabase.get('users', {'id': uid},
                                 select='id,email,full_name,role,modules,is_active')
            if not u:
                # Si la consulta falla por una columna (PostgREST responde 400 y
                # el cliente devuelve []), esto se ejecuta en CADA petición: sin
                # este respaldo, un nombre de columna equivocado no rompe una
                # pantalla, echa del sistema a todo el mundo a la vez.
                u = app.supabase.get('users', {'id': uid},
                                     select='id,email,full_name,role,modules')
            if u:
                usuario = User(u[0])
                # Cuenta desactivada: se corta AQUÍ, no en el formulario de
                # entrada. Este es el único punto por el que pasan las tres
                # formas de identificarse —clave, huella/Face ID y rostro— y
                # además se ejecuta en cada petición, así que desactivar a
                # alguien lo deja fuera al instante, sin esperar a que su sesión
                # caduque.
                if not usuario.active:
                    return None
                return usuario
        return None

    # ------ Security headers on every response ------
    @app.after_request
    def add_headers(response):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'SAMEORIGIN'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        p = request.path
        if p.startswith('/static/'):
            response.headers['Cache-Control'] = 'public, max-age=86400, immutable'
        elif p in ('/calendar/api/titles', '/calendar/api/encargados',
                   '/calendar/api/temas', '/calendar/api/ciudades', '/calendar/api/clients'):
            response.headers['Cache-Control'] = 'private, max-age=60'
        elif p.startswith('/calendar/api/'):
            response.headers['Cache-Control'] = 'no-store'
        return response

    # ------ Context processor (cached) ------
    @app.context_processor
    def inject_layout_globals():
        connected = False; needs_reauth = False
        try:
            if current_user.is_authenticated and app.supabase:
                cache_key = f'google_status_{current_user.role}'
                val, hit = _google_cache.get(cache_key)
                if hit:
                    connected, needs_reauth = val
                else:
                    tokens = app.supabase.get('google_tokens',
                        {'email': GOOGLE_ACCOUNT_EMAIL}, select='email,token,refresh_token,token_expiry,id')
                    if tokens:
                        if current_user.role == 'admin':
                            connected = get_google_creds(app) is not None
                            needs_reauth = not connected
                        else:
                            connected = True
                    _google_cache.set(cache_key, (connected, needs_reauth))
        except Exception:
            pass
        user_roles_list = []
        active_role_id = None
        active_modules = []
        if current_user.is_authenticated and app.supabase:
            try:
                user_roles_list = get_user_roles(app, current_user.id)
                active_role_id = get_active_role_id(app, current_user.id)
                # Los del rol MÁS los concedidos a esta persona: el menú tiene
                # que enseñar exactamente lo mismo que deja pasar `user_can`, o
                # el módulo existe pero no hay por dónde entrar.
                active_modules = sorted(modulos_efectivos(app, current_user.id))
            except Exception:
                pass
        # Estado de la reconexión automática: permite que el aviso distinga un
        # corte pasajero (se arregla solo) de una revocación real (hace falta
        # que una persona vuelva a autorizar).
        google_health = {'estado': 'ok'}
        if needs_reauth and app.supabase:
            try:
                google_health = _leer_estado_google(app)
            except Exception:
                pass
        # Estado real de las otras dos integraciones. El menú las daba por
        # desconectadas siempre: Microsoft salía «Conectar» aunque estuviera
        # conectado, y el puente con ATLAS no aparecía en ninguna parte pese a
        # existir y estar funcionando. Una integración que miente sobre su
        # estado es peor que no mostrarla.
        #
        # Se comprueba sólo si HAY fila en ms_tokens, no se pide un token
        # válido: get_ms_token puede disparar un refresco contra Microsoft, y
        # eso no puede pasar al pintar cada página. Con caché de 2 minutos,
        # igual que Google.
        ms_connected = False
        try:
            if current_user.is_authenticated and app.supabase:
                val, hit = _ms_cache.get('ms_conectado')
                if hit:
                    ms_connected = val
                else:
                    ms_connected = bool(app.supabase.get('ms_tokens', select='email'))
                    _ms_cache.set('ms_conectado', ms_connected)
        except Exception:
            pass
        try:
            atlas_activo = _atlas.disponible()
        except Exception:
            atlas_activo = False
        return {'google_connected_global': connected, 'google_needs_reauth': needs_reauth,
                'google_health': google_health,
                'ms_connected_global': ms_connected, 'atlas_activo': atlas_activo,
                'user_roles_list': user_roles_list, 'active_role_id': active_role_id,
                'active_modules': active_modules}

    # ============================================================
    #  PUBLIC ROUTES
    # ============================================================
    @app.route('/')
    def home():
        return redirect('/dashboard') if current_user.is_authenticated else render_template('index.html')

    @app.before_request
    def _sesion_deslizante():
        """Marca la sesión como permanente para que Flask le aplique el plazo de
        inactividad de config.py (PERMANENT_SESSION_LIFETIME = 20 min).

        Sin esto la constante se ignora: Flask solo reemite la cookie —y con
        ella renueva la marca de tiempo— cuando la sesión es permanente. Cubre
        de una vez las tres puertas de entrada (contraseña, biometría y rostro),
        en lugar de repetirlo en cada `login_user`.
        """
        if current_user.is_authenticated and not session.permanent:
            session.permanent = True

    @app.before_request
    def _force_password_change():
        """Mientras el usuario arrastre una clave temporal del administrador, no
        puede usar el sistema: solo cambiarla o cerrar sesión."""
        if not current_user.is_authenticated:
            return None
        path = request.path or ''
        if path.startswith(FREE_PATHS_PASSWORD):
            return None
        # El estado se fija al iniciar sesión: así el guard no consulta la base
        # en cada petición.
        if not session.get('must_change_password'):
            return None
        if path.startswith('/api/'):
            return jsonify({'error': 'Debes definir una nueva contraseña',
                            'redirect': '/account/password'}), 403
        flash('Tu contraseña fue restablecida por el administrador. '
              'Define una nueva para continuar.', 'warning')
        return redirect('/account/password')

    @app.route('/login', methods=['GET', 'POST'])
    @limiter.limit('10 per minute')
    def login():
        if current_user.is_authenticated:
            return redirect('/dashboard')
        if request.method == 'POST':
            email = _sanitize(request.form.get('email', ''), 254).lower()
            pw = request.form.get('password', '')
            if not email or not pw:
                flash('Completa todos los campos.', 'danger')
                return render_template('login.html')
            users = app.supabase.get('users', {'email': email})
            if users and check_password_hash(users[0]['password_hash'], pw):
                u = users[0]
                # Una clave temporal caducada no sirve: hay que pedir al
                # administrador que la restablezca de nuevo.
                if _temp_password_expired(u):
                    flash('La contraseña temporal que te entregó el administrador ya caducó. '
                          'Pídele que la restablezca nuevamente.', 'danger')
                    return render_template('login.html')
                # Cuenta desactivada. Se dice con claridad en vez de responder
                # «email o contraseña incorrectos»: la clave es correcta, y
                # mandar a alguien a pelearse con su contraseña cuando el
                # problema es otro sólo termina en una llamada al administrador.
                if not _cuenta_activa(u):
                    flash('Tu cuenta está desactivada. Habla con el administrador '
                          'para que la habilite de nuevo.', 'warning')
                    return render_template('login.html')
                # `remember=True` y `session.permanent` AQUÍ, no en el
                # before_request. El before_request se ejecuta ANTES de la vista,
                # cuando el usuario todavía es anónimo, así que en la respuesta
                # del propio inicio de sesión la cookie salía sin fecha de
                # caducidad: era una cookie de navegador. Si el navegador o la
                # aplicación instalada la descartaba al cerrarse, ese primer
                # inicio de sesión se perdía y había que volver a entrar. Puesto
                # aquí, la cookie ya sale fechada desde la primera respuesta.
                login_user(User(u), remember=True)
                session.permanent = True
                session['must_change_password'] = bool(u.get('must_change_password'))
                roles = get_user_roles(app, u['id'])
                role_ids = {r['id'] for r in roles}
                stored = u.get('active_role_id')
                session['active_role_id'] = stored if stored in role_ids else (roles[0]['id'] if roles else None)
                return redirect(_safe_next_path(request.args.get('next')))
            flash('Email o contraseña incorrectos.', 'danger')
        return render_template('login.html')

    @app.route('/account/active-role', methods=['POST'])
    @login_required
    @csrf_protect
    def set_active_role():
        role_id = request.form.get('role_id')
        valid_ids = {r['id'] for r in get_user_roles(app, current_user.id)}
        if role_id not in valid_ids:
            flash('Rol inválido.', 'danger')
            return redirect(request.referrer or '/dashboard')
        session['active_role_id'] = role_id
        app.supabase.update('users', current_user.id, {'active_role_id': role_id})
        return redirect(request.referrer or '/dashboard')

    # ============================================================
    #  WEBAUTHN — Face ID / huella (passkeys)
    # ============================================================
    def _rp_id():
        return request.host.split(':')[0]

    def _origin():
        return f'{request.scheme}://{request.host}'

    def _wa_guard():
        if not WEBAUTHN_AVAILABLE:
            return jsonify({'error': 'WebAuthn no instalado en el servidor'}), 503
        return None

    @app.route('/webauthn/register/begin', methods=['POST'])
    @login_required
    def webauthn_register_begin():
        guard = _wa_guard()
        if guard:
            return guard
        existing = app.supabase.get('webauthn_credentials',
            {'user_id': str(current_user.id)}, select='credential_id')
        exclude = [PublicKeyCredentialDescriptor(id=base64url_to_bytes(c['credential_id']))
                   for c in existing if c.get('credential_id')]
        opts = generate_registration_options(
            rp_id=_rp_id(),
            rp_name='calendarios-map',
            user_id=str(current_user.id).encode('utf-8'),
            user_name=current_user.email or str(current_user.id),
            user_display_name=current_user.full_name or current_user.email or 'Usuario',
            authenticator_selection=AuthenticatorSelectionCriteria(
                resident_key=ResidentKeyRequirement.PREFERRED,
                user_verification=UserVerificationRequirement.PREFERRED),
            exclude_credentials=exclude,
        )
        session['wa_reg_challenge'] = bytes_to_base64url(opts.challenge)
        return app.response_class(options_to_json(opts), mimetype='application/json')

    @app.route('/webauthn/register/complete', methods=['POST'])
    @login_required
    def webauthn_register_complete():
        guard = _wa_guard()
        if guard:
            return guard
        data = request.get_json(silent=True) or {}
        nombre = (data.pop('nombre', '') or '')[:80]
        challenge = session.pop('wa_reg_challenge', None)
        if not challenge:
            return jsonify({'success': False, 'error': 'Sesion expirada, reintenta'})
        try:
            v = verify_registration_response(
                credential=json.dumps(data),
                expected_challenge=base64url_to_bytes(challenge),
                expected_rp_id=_rp_id(),
                expected_origin=_origin(),
            )
        except Exception as e:
            return jsonify({'success': False, 'error': f'Verificacion fallida: {e}'})
        transports = ','.join((data.get('response', {}) or {}).get('transports', []) or [])
        rec = app.supabase.insert('webauthn_credentials', {
            'user_id': str(current_user.id),
            'credential_id': bytes_to_base64url(v.credential_id),
            'public_key': bytes_to_base64url(v.credential_public_key),
            'sign_count': v.sign_count,
            'transports': transports,
            'nombre': nombre or 'Dispositivo',
        })
        if not rec:
            return jsonify({'success': False,
                            'error': 'No se pudo guardar (¿corriste la migracion 003?)'})
        return jsonify({'success': True})

    @app.route('/webauthn/authenticate/begin', methods=['POST'])
    def webauthn_auth_begin():
        guard = _wa_guard()
        if guard:
            return guard
        opts = generate_authentication_options(
            rp_id=_rp_id(),
            user_verification=UserVerificationRequirement.PREFERRED,
        )
        session['wa_auth_challenge'] = bytes_to_base64url(opts.challenge)
        return app.response_class(options_to_json(opts), mimetype='application/json')

    @app.route('/webauthn/authenticate/complete', methods=['POST'])
    @limiter.limit('10 per minute')
    def webauthn_auth_complete():
        guard = _wa_guard()
        if guard:
            return guard
        data = request.get_json(silent=True) or {}
        challenge = session.pop('wa_auth_challenge', None)
        if not challenge:
            return jsonify({'success': False, 'error': 'Sesion expirada, reintenta'})
        cred_id = data.get('id', '')
        rows = app.supabase.get('webauthn_credentials', {'credential_id': cred_id})
        if not rows:
            return jsonify({'success': False, 'error': 'Dispositivo no reconocido'})
        rec = rows[0]
        try:
            v = verify_authentication_response(
                credential=json.dumps(data),
                expected_challenge=base64url_to_bytes(challenge),
                expected_rp_id=_rp_id(),
                expected_origin=_origin(),
                credential_public_key=base64url_to_bytes(rec['public_key']),
                credential_current_sign_count=rec.get('sign_count', 0) or 0,
                require_user_verification=False,
            )
        except Exception as e:
            return jsonify({'success': False, 'error': f'Autenticacion fallida: {e}'})
        app.supabase.update('webauthn_credentials', rec['id'], {
            'sign_count': v.new_sign_count,
            'last_used_at': datetime.now(timezone.utc).isoformat(),
        })
        users = app.supabase.get('users', {'id': rec['user_id']},
                                 select='id,email,full_name,role,must_change_password,is_active')
        if not users:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'})
        if not _cuenta_activa(users[0]):
            return jsonify({'success': False,
                            'error': 'Tu cuenta está desactivada. Habla con el administrador.'})
        login_user(User(users[0]), remember=True)
        session.permanent = True
        session['must_change_password'] = bool(users[0].get('must_change_password'))
        if session['must_change_password']:
            return jsonify({'success': True, 'redirect': '/account/password'})
        return jsonify({'success': True, 'redirect': '/dashboard'})

    @app.route('/webauthn/credentials', methods=['GET'])
    @login_required
    def webauthn_credentials_list():
        rows = app.supabase.get('webauthn_credentials',
            {'user_id': str(current_user.id)},
            select='id,nombre,created_at,last_used_at')
        return jsonify(rows or [])

    @app.route('/webauthn/credentials/delete/<cred_pk>', methods=['POST'])
    @login_required
    def webauthn_credentials_delete(cred_pk):
        rows = app.supabase.get('webauthn_credentials', {'id': cred_pk}, select='id,user_id')
        if not rows or str(rows[0].get('user_id')) != str(current_user.id):
            return jsonify({'success': False, 'error': 'No autorizado'})
        app.supabase.delete('webauthn_credentials', cred_pk)
        return jsonify({'success': True})

    # ============================================================
    #  FACE LOGIN — reconocimiento facial por cámara (face-api.js)
    #  Descriptor 128-d calculado en el navegador; comparación en
    #  el servidor. Conveniencia: SIN detección de vida (un foto/
    #  pantalla puede engañarlo). La passkey es más segura.
    # ============================================================
    FACE_THRESHOLD = 0.55   # distancia euclidiana máxima para considerar match

    def _face_distance(a, b):
        if len(a) != len(b):
            return 9.9
        return sum((x - y) ** 2 for x, y in zip(a, b)) ** 0.5

    def _valid_descriptor(d):
        return (isinstance(d, list) and len(d) == 128
                and all(isinstance(x, (int, float)) for x in d))

    @app.route('/face/enroll', methods=['POST'])
    @login_required
    def face_enroll():
        data = request.get_json(silent=True) or {}
        desc = data.get('descriptor')
        if not _valid_descriptor(desc):
            return jsonify({'success': False, 'error': 'Descriptor facial invalido'})
        nombre = (data.get('nombre') or 'Rostro')[:80]
        rec = app.supabase.insert('face_descriptors', {
            'user_id': str(current_user.id),
            'descriptor': json.dumps(desc),
            'nombre': nombre,
        })
        if not rec:
            return jsonify({'success': False,
                            'error': 'No se pudo guardar (¿corriste la migracion 004?)'})
        return jsonify({'success': True})

    @app.route('/face/list', methods=['GET'])
    @login_required
    def face_list():
        rows = app.supabase.get('face_descriptors', {'user_id': str(current_user.id)},
                                select='id,nombre,created_at')
        return jsonify(rows or [])

    @app.route('/face/delete/<fid>', methods=['POST'])
    @login_required
    def face_delete(fid):
        rows = app.supabase.get('face_descriptors', {'id': fid}, select='id,user_id')
        if not rows or str(rows[0].get('user_id')) != str(current_user.id):
            return jsonify({'success': False, 'error': 'No autorizado'})
        app.supabase.delete('face_descriptors', fid)
        return jsonify({'success': True})

    @app.route('/face/verify', methods=['POST'])
    @limiter.limit('10 per minute')
    def face_verify():
        data = request.get_json(silent=True) or {}
        email = _sanitize(data.get('email', ''), 254).lower()
        desc = data.get('descriptor')
        if not email or not _valid_descriptor(desc):
            return jsonify({'success': False, 'error': 'Datos invalidos'})
        users = app.supabase.get('users', {'email': email},
                                 select='id,email,full_name,role,must_change_password,is_active')
        # Mensaje genérico para no revelar si el email existe
        if not users:
            return jsonify({'success': False, 'error': 'Rostro no reconocido'})
        if not _cuenta_activa(users[0]):
            return jsonify({'success': False,
                            'error': 'Tu cuenta está desactivada. Habla con el administrador.'})
        uid = str(users[0]['id'])
        stored = app.supabase.get('face_descriptors', {'user_id': uid}, select='descriptor')
        best = 9.9
        for s in (stored or []):
            try:
                v = json.loads(s['descriptor'])
            except Exception:
                continue
            if _valid_descriptor(v):
                best = min(best, _face_distance(desc, v))
        if best <= FACE_THRESHOLD:
            login_user(User(users[0]), remember=True)
            session.permanent = True
            session['must_change_password'] = bool(users[0].get('must_change_password'))
            return jsonify({'success': True,
                            'redirect': '/account/password' if session['must_change_password']
                                        else '/dashboard'})
        return jsonify({'success': False, 'error': 'Rostro no reconocido'})

    @app.route('/register', methods=['GET', 'POST'])
    def register():
        """El registro público está CERRADO: en este sistema entra quien el
        administrador da de alta.

        Antes cualquiera que llegara a esta dirección se creaba una cuenta con
        su propia clave y podía iniciar sesión; lo que esperaba aprobación era
        sólo el acceso a los calendarios, no el ingreso. Para una agenda interna
        con datos de clientes eso es la puerta abierta. El alta vive ahora en
        Administración → Usuarios, junto al resto del gobierno de accesos.

        La ruta se conserva (en vez de borrarla) para que quien tenga el enlace
        guardado reciba una explicación en lugar de un 404 sin sentido."""
        if current_user.is_authenticated:
            return redirect('/dashboard')
        flash('El acceso a este sistema lo concede el administrador. '
              'Solicítale que te dé de alta.', 'info')
        return redirect('/login')

    @app.route('/logout')
    @login_required
    def logout():
        # logout_user() solo quita las claves de flask_login; el resto de la
        # sesión (estado de OAuth, banderas de interfaz, must_change_password)
        # seguía viajando en la cookie. Se vacía entera para no dejar nada del
        # usuario anterior en el dispositivo.
        logout_user()
        session.clear()
        return redirect('/')

    @app.route('/profile', methods=['GET', 'POST'])
    @login_required
    def profile():
        row = (app.supabase.get('users', {'id': current_user.id}) or [{}])[0]
        if request.method == 'POST':
            data = {}
            name = _sanitize(request.form.get('full_name', ''), 100)
            email = _sanitize(request.form.get('email', ''), 254).lower()
            data['phone'] = _sanitize(request.form.get('phone', ''), 30)
            data['position'] = _sanitize(request.form.get('position', ''), 80)
            if name: data['full_name'] = name
            # El email es la credencial de acceso: cambiarlo exige confirmar la
            # clave actual. El cambio de contraseña vive en /account/password,
            # que sí verifica la anterior.
            if email and _validate_email(email) and email != (row.get('email') or '').lower():
                if not check_password_hash(row.get('password_hash') or '',
                                           request.form.get('current_password', '')):
                    flash('Para cambiar tu email debes confirmar tu contraseña actual.', 'danger')
                    return redirect('/profile')
                data['email'] = email
            if data:
                app.supabase.update('users', current_user.id, data)
            flash('Datos actualizados', 'success')
            return redirect('/profile')
        my_cals = get_user_calendars(app, current_user.id)
        return render_template('profile.html', my_calendars=my_cals,
                               all_modules=ALL_MODULES, profile_row=row,
                               min_password=MIN_PASSWORD)

    # ============================================================
    #  MI CLAVE — cambio con verificación de la anterior
    # ============================================================
    @app.route('/account/password', methods=['GET', 'POST'])
    @login_required
    def account_password():
        row = (app.supabase.get('users', {'id': current_user.id}) or [{}])[0]
        forced = bool(row.get('must_change_password'))
        if request.method == 'POST':
            current = request.form.get('current_password', '')
            new = request.form.get('new_password', '')
            confirm = request.form.get('confirm_password', '')
            if not check_password_hash(row.get('password_hash') or '', current):
                flash('La contraseña actual no es correcta.', 'danger')
            elif len(new) < MIN_PASSWORD:
                flash(f'La nueva contraseña debe tener al menos {MIN_PASSWORD} caracteres.', 'danger')
            elif new != confirm:
                flash('La nueva contraseña y su confirmación no coinciden.', 'danger')
            elif check_password_hash(row.get('password_hash') or '', new):
                flash('La nueva contraseña debe ser distinta de la anterior.', 'danger')
            else:
                app.supabase.update('users', current_user.id, {
                    'password_hash': generate_password_hash(new),
                    'must_change_password': False,
                    'temp_password_expires': None,
                    'password_updated_at': datetime.now(timezone.utc).isoformat(),
                })
                session['must_change_password'] = False
                _log_password(app, current_user.id, 'self_change', current_user.id)
                flash('Contraseña actualizada correctamente.', 'success')
                return redirect('/dashboard')
        return render_template('account_password.html', forced=forced,
                               min_password=MIN_PASSWORD)

    # ============================================================
    #  MICROSOFT OAUTH — To-Do
    # ============================================================
    @app.route('/auth/microsoft')
    @login_required
    def auth_microsoft():
        if not is_admin(): return redirect(url_for('planning'))
        cid = app.config.get('MS_CLIENT_ID', '')
        if not cid:
            flash('Configura MS_CLIENT_ID en las variables de entorno.', 'warning')
            return redirect(url_for('planning'))
        redirect_uri = app.config.get('MS_REDIRECT_URI') or request.host_url.rstrip('/') + '/auth/microsoft/callback'
        state = secrets.token_urlsafe(24)
        session['ms_state'] = state
        params = (f'?client_id={cid}'
                  f'&response_type=code'
                  f'&redirect_uri={_url_quote(redirect_uri, safe="")}'
                  f'&scope={_url_quote(MS_SCOPES, safe="")}'
                  f'&response_mode=query'
                  f'&prompt=select_account'
                  f'&state={_url_quote(state, safe="")}')
        return redirect(MS_AUTH_URL + params)

    @app.route('/auth/microsoft/callback')
    @login_required
    def auth_microsoft_callback():
        if not is_admin(): return redirect(url_for('planning'))
        code  = request.args.get('code')
        error = request.args.get('error_description') or request.args.get('error')
        if error:
            flash(f'Microsoft error: {error}', 'danger')
            return redirect(url_for('planning'))
        if not code:
            flash('No se recibió código de autorización.', 'danger')
            return redirect(url_for('planning'))
        expected_state = session.pop('ms_state', None)
        if not expected_state or request.args.get('state') != expected_state:
            flash('Sesión expirada o solicitud inválida. Intenta de nuevo.', 'warning')
            return redirect(url_for('planning'))
        redirect_uri = app.config.get('MS_REDIRECT_URI') or request.host_url.rstrip('/') + '/auth/microsoft/callback'
        try:
            r = req_lib.post(MS_TOKEN_URL, data={
                'client_id':     app.config.get('MS_CLIENT_ID', ''),
                'client_secret': app.config.get('MS_CLIENT_SECRET', ''),
                'grant_type':    'authorization_code',
                'code':          code,
                'redirect_uri':  redirect_uri,
                'scope':         MS_SCOPES,
            }, timeout=(5, 15))
            if r.status_code != 200:
                flash(f'Error al obtener token: {r.text[:200]}', 'danger')
                return redirect(url_for('planning'))
            d = r.json()
            exp = (datetime.now(timezone.utc)
                   + timedelta(seconds=d.get('expires_in', 3600))).isoformat()
            # Get user email from Graph
            me_r = req_lib.get(f'{MS_GRAPH_URL}/me',
                               headers={'Authorization': f'Bearer {d["access_token"]}'},
                               timeout=(5, 10))
            ms_email = me_r.json().get('mail') or me_r.json().get('userPrincipalName', 'microsoft') if me_r.ok else 'microsoft'
            # Upsert token
            existing = app.supabase.get('ms_tokens', {'email': ms_email})
            token_data = {
                'email':         ms_email,
                'access_token':  d['access_token'],
                'refresh_token': d.get('refresh_token', ''),
                'expires_at':    exp,
            }
            if existing:
                app.supabase.update('ms_tokens', existing[0]['id'], token_data)
            else:
                app.supabase.insert('ms_tokens', token_data)
            flash(f'✅ Microsoft To-Do conectado ({ms_email})', 'success')
        except Exception as e:
            flash(f'Error de conexión: {e}', 'danger')
        return redirect(url_for('planning'))

    @app.route('/auth/microsoft/disconnect', methods=['POST'])
    @login_required
    def auth_microsoft_disconnect():
        if not is_admin(): return jsonify({'success': False})
        tokens = app.supabase.get('ms_tokens', select='id')
        for t in (tokens or []):
            app.supabase.delete('ms_tokens', t['id'])
        return jsonify({'success': True})

    # ============================================================
    #  GOOGLE OAUTH
    # ============================================================
    @app.route('/auth/google')
    @login_required
    def google_auth():
        if not is_admin():
            flash('Solo el administrador puede conectar Google Calendar.', 'warning')
            return redirect('/dashboard')
        flow = Flow.from_client_config({'web': {
            'client_id': app.config['GOOGLE_CLIENT_ID'],
            'client_secret': app.config['GOOGLE_CLIENT_SECRET'],
            'auth_uri': 'https://accounts.google.com/o/oauth2/auth',
            'token_uri': 'https://oauth2.googleapis.com/token',
            'redirect_uris': [app.config['GOOGLE_REDIRECT_URI']]}}, scopes=GOOGLE_SCOPES)
        flow.redirect_uri = app.config['GOOGLE_REDIRECT_URI']
        auth_url, state = flow.authorization_url(access_type='offline', prompt='consent')
        session['state'] = state
        return redirect(auth_url)

    @app.route('/auth/google/callback')
    @login_required
    def google_callback():
        state = session.get('state')
        if not state:
            flash('Sesion expirada. Intenta de nuevo.', 'warning')
            return redirect('/dashboard')
        flow = Flow.from_client_config({'web': {
            'client_id': app.config['GOOGLE_CLIENT_ID'],
            'client_secret': app.config['GOOGLE_CLIENT_SECRET'],
            'auth_uri': 'https://accounts.google.com/o/oauth2/auth',
            'token_uri': 'https://oauth2.googleapis.com/token',
            'redirect_uris': [app.config['GOOGLE_REDIRECT_URI']]}},
            scopes=GOOGLE_SCOPES, state=state)
        flow.redirect_uri = app.config['GOOGLE_REDIRECT_URI']
        flow.fetch_token(authorization_response=request.url)
        if save_google_creds(app, flow.credentials):
            flash('Google Calendar conectado correctamente.', 'success')
        else:
            flash('No se pudieron guardar las credenciales de Google. Intenta reconectar.', 'danger')
        return redirect('/dashboard')

    # ============================================================
    #  ADMIN — USERS  (optimized: O(3) queries instead of O(N+3))
    # ============================================================
    @app.route('/admin/users')
    @login_required
    def admin_users():
        if not is_admin():
            return redirect('/dashboard')
        users     = app.supabase.get('users',
                        select='id,email,full_name,role,created_at,phone,position,'
                               'must_change_password,is_active') or []
        all_cals  = _get_calendar_config(app)
        all_perms = app.supabase.get('calendar_permissions',
                        select='id,user_id,calendar_id,status')
        cal_by_id = {c['calendar_id']: c for c in all_cals}
        user_by_id = {u['id']: u for u in users}
        # Solicitudes pendientes (calendar_permissions sigue siendo donde /register las deja)
        pending_perms = defaultdict(list)
        for p in all_perms:
            if p['status'] == 'pending':
                pending_perms[p['user_id']].append(p)
        pending = []; pending_all = []
        for uid, perms_list in pending_perms.items():
            for p in perms_list:
                pending_all.append({'id': p['id'], 'user_id': uid, 'calendar_id': p['calendar_id']})
            u = user_by_id.get(uid, {})
            pending.append({
                'user_id': uid,
                'user_name':  u.get('full_name', ''),
                'user_email': u.get('email', ''),
                'calendars':  [cal_by_id[p['calendar_id']] for p in perms_list
                               if p['calendar_id'] in cal_by_id],
            })
        # Roles: catálogo completo + cuáles tiene cada usuario
        all_roles = app.supabase.get('roles', select='id,name,level') or []
        for r in all_roles:
            r['level'] = r.get('level') or DEFAULT_ROLE_LEVEL
        user_roles_all = app.supabase.get('user_roles', select='user_id,role_id') or []
        roles_by_user = defaultdict(set)
        for ur in user_roles_all:
            roles_by_user[ur['user_id']].add(ur['role_id'])
        for u in users:
            u['role_ids'] = roles_by_user.get(u['id'], set())
            u['activa']   = _cuenta_activa(u)
        # Historial de accesos concedidos y retirados. Se pide ordenado y
        # recortado en el propio servidor: esta tabla sólo crece.
        auditoria = app.supabase.get_q('permission_audit',
                        {'order': 'created_at.desc', 'limit': 60},
                        select='actor_email,target_email,accion,detalle,created_at') or []
        return render_template('admin_users.html', users=users, calendarios=all_cals,
                               pending=pending, pending_all=pending_all,
                               all_roles=all_roles, role_levels=ROLE_LEVELS,
                               all_modules=ALL_MODULES, auditoria=auditoria)

    @app.route('/admin/users/<uid>/permisos', methods=['GET'])
    @login_required
    def admin_user_permisos(uid):
        """Lo que esta persona puede hacer: módulos y, dentro de cada uno, qué
        acciones.

        Devuelve por separado el módulo que le da su ROL y el que le ha dado el
        administrador a ELLA, porque no se gobiernan igual: el del rol se cambia
        en Administración → Roles y afecta a todos los que lo tengan; el suelto
        es de esta persona y sólo suyo."""
        if not is_admin():
            return jsonify({'success': False, 'error': 'Solo admin'}), 403
        filas = app.supabase.get('users', {'id': uid}, select='id,full_name,email,active_role_id')
        if not filas:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        persona = filas[0]
        vacio = {'modules': [], 'calendar_ids': set(), 'project_ids': set(), 'ms_emails': set()}
        rol = role_grants(app, persona['active_role_id']) if persona.get('active_role_id') else vacio
        mio = get_user_grants(app, uid)

        catalogo = [{
            'modulo': mod,
            'etiqueta': etiqueta,
            'por_rol': mod in set(rol['modules']),
            'suelto': mod in mio['modules'],
            'tiene_modulo': mod in set(rol['modules']) or mod in mio['modules'],
            'acciones': [{'clave': f'{mod}.{acc}', 'nombre': nombre, 'sensible': sensible}
                         for acc, nombre, sensible in SUBMODULOS.get(mod, [])],
        } for mod, etiqueta in ALL_MODULES]

        def recursos(items, clave_rol, propios):
            """Cada recurso, diciendo si viene del rol o se le dio a esta persona.
            Se distinguen porque no se gobiernan igual: el del rol se quita en
            Administración → Roles y afecta a todos los que tengan ese rol."""
            return [{**it,
                     'por_rol': it['id'] in rol[clave_rol],
                     'suelto':  it['id'] in propios} for it in items]

        calendarios = [{'id': c['calendar_id'], 'nombre': c.get('name') or c['calendar_id'],
                        'detalle': c.get('email') or '', 'color': c.get('color')}
                       for c in _get_calendar_config(app)]
        proyectos = [{'id': p['id'], 'nombre': p.get('name') or '(sin nombre)', 'detalle': ''}
                     for p in (app.supabase.get('projects', select='id,name') or [])]
        cuentas_ms = [{'id': t['email'], 'nombre': t['email'], 'detalle': 'Microsoft To-Do'}
                      for t in (app.supabase.get('ms_tokens', select='email') or []) if t.get('email')]

        return jsonify({
            'success': True,
            'usuario': {'id': persona['id'], 'nombre': persona.get('full_name'),
                        'email': persona.get('email')},
            'catalogo': catalogo,
            'modulos_sueltos': sorted(mio['modules']),
            'concedidos': sorted(get_user_permissions(app, uid)),
            'recursos': {
                'calendarios': recursos(calendarios, 'calendar_ids', mio['calendar_ids']),
                'proyectos':   recursos(proyectos,   'project_ids',  mio['project_ids']),
                'cuentas_ms':  recursos(cuentas_ms,  'ms_emails',    mio['ms_emails']),
            },
        })

    @app.route('/admin/users/<uid>/permisos', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_user_permisos_guardar(uid):
        """Reemplaza lo concedido a una persona: módulos, recursos y acciones.

        Body: {modulos: [...], calendarios: [...], proyectos: [...],
               cuentas_ms: [...], permisos: [...]}

        Dentro de una categoría se REEMPLAZA: lo que no venga marcado se retira.
        La pantalla envía siempre el estado completo, así que quitar una casilla
        tiene que quitar el acceso — si esto sólo añadiera, un permiso concedido
        por error no habría forma de retirarlo desde aquí.

        Pero una categoría AUSENTE del cuerpo no se toca. La diferencia importa:
        antes, no mencionar `calendarios` equivalía a mandarlos vacíos, y una
        petición que sólo quería cambiar un módulo le retiraba a la persona
        todos sus calendarios sin nombrarlos. Quedó en el historial. Retirar un
        acceso tiene que ser algo que alguien pidió, no lo que ocurre por
        omisión."""
        if not is_admin():
            return jsonify({'success': False, 'error': 'Solo admin'}), 403
        filas = app.supabase.get('users', {'id': uid}, select='id,email')
        if not filas:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        correo = filas[0].get('email')

        cuerpo = request.get_json() or {}
        mio = get_user_grants(app, uid)
        concedido, retirado = [], []

        def sincronizar(tabla, columna, clave_grants, clave_cuerpo, validos, etiqueta):
            """Deja la tabla con exactamente lo pedido. `validos` acota lo que se
            acepta: un identificador inventado desde el navegador no debe poder
            colarse como concesión.

            Si la categoría no viene en el cuerpo se deja intacta; una lista
            vacía sí la vacía. Omitir no es lo mismo que pedir que se quite."""
            actuales = mio[clave_grants]
            if clave_cuerpo not in cuerpo:
                return actuales
            nuevos = {v for v in (cuerpo.get(clave_cuerpo) or []) if v in validos}
            for valor in actuales - nuevos:
                for fila in (app.supabase.get(tabla, {'user_id': uid, columna: valor},
                                              select='id') or []):
                    app.supabase.delete(tabla, fila['id'])
                retirado.append(f'{etiqueta}:{valor}')
            for valor in sorted(nuevos - actuales):
                app.supabase.insert_ignore(tabla, {
                    'user_id': uid, columna: valor, 'granted_by': str(current_user.id)})
                concedido.append(f'{etiqueta}:{valor}')
            return nuevos

        mods_nuevos = sincronizar(
            'user_modules', 'modulo', 'modules', 'modulos',
            {m for m, _ in ALL_MODULES}, 'módulo')
        cals_nuevos = sincronizar(
            'user_calendars', 'calendar_id', 'calendar_ids', 'calendarios',
            {c['calendar_id'] for c in _get_calendar_config(app)}, 'calendario')
        proys_nuevos = sincronizar(
            'user_projects', 'project_id', 'project_ids', 'proyectos',
            {p['id'] for p in (app.supabase.get('projects', select='id') or [])}, 'proyecto')
        ms_nuevos = sincronizar(
            'user_ms_accounts', 'ms_email', 'ms_emails', 'cuentas_ms',
            {t['email'] for t in (app.supabase.get('ms_tokens', select='email') or [])
             if t.get('email')}, 'cuenta MS')

        # ── Acciones dentro de cada módulo ─────────────────────────────────
        actuales = get_user_permissions(app, uid)
        nuevos = actuales
        if 'permisos' in cuerpo:      # ausente = no se tocan (ver la nota de arriba)
            validos = {f'{m}.{a}' for m, lista in SUBMODULOS.items() for a, _, _ in lista}
            nuevos = {p for p in (cuerpo.get('permisos') or []) if p in validos}
            for permiso in actuales - nuevos:
                for fila in (app.supabase.get('user_permissions',
                                              {'user_id': uid, 'permiso': permiso}, select='id') or []):
                    app.supabase.delete('user_permissions', fila['id'])
            por_agregar = [{'user_id': uid, 'permiso': p, 'granted_by': current_user.id}
                           for p in sorted(nuevos - actuales)]
            if por_agregar:
                app.supabase.insert('user_permissions', por_agregar)
            concedido += sorted(nuevos - actuales)
            retirado  += sorted(actuales - nuevos)

        _user_perms_cache.invalidate(str(uid))
        _user_grants_cache.invalidate(str(uid))
        _user_cal_cache.invalidate(str(uid))

        # Sólo se apunta cuando hubo cambio: un historial lleno de «no cambió
        # nada» esconde justo lo que se quiere encontrar.
        if concedido or retirado:
            _auditar_acceso(app, 'permisos', uid, correo, ' · '.join(filter(None, [
                ('concedido: ' + ', '.join(sorted(concedido))) if concedido else '',
                ('retirado: '  + ', '.join(sorted(retirado)))  if retirado  else '',
            ])))

        return jsonify({'success': True, 'concedidos': sorted(nuevos),
                        'modulos_sueltos': sorted(mods_nuevos),
                        'calendarios': sorted(cals_nuevos),
                        'proyectos': sorted(proys_nuevos),
                        'cuentas_ms': sorted(ms_nuevos),
                        'agregados': len(concedido),
                        'quitados': len(retirado)})

    def _crear_usuario(datos):
        """Inserta el usuario tolerando que la migración 027 no esté aplicada.

        `created_by_admin` sólo existe a partir de esa migración. Si todavía no
        se corrió, PostgREST rechaza el INSERT ENTERO y el alta se cae sin que
        se entienda por qué; así que se reintenta sin esa columna. La cuenta
        queda creada igual: lo que se pierde es saber quién la dio de alta, no
        la persona."""
        fila = app.supabase.insert('users', datos)
        if fila:
            return fila
        recorte = {k: v for k, v in datos.items() if k != 'created_by_admin'}
        if recorte == datos:
            return None
        return app.supabase.insert('users', recorte)

    @app.route('/admin/users/create', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_user_create():
        """Alta de una persona. El administrador es el único que abre la puerta.

        No se le pide una contraseña al administrador: se genera una temporal de
        un solo uso, se muestra UNA vez para que se la entregue en mano y la
        persona está obligada a cambiarla al entrar. Así el administrador nunca
        llega a conocer la clave definitiva de nadie."""
        if not is_admin():
            return redirect('/dashboard')
        email  = _sanitize(request.form.get('email', ''), 254).lower()
        nombre = _sanitize(request.form.get('full_name', ''), 100)
        if not email or not nombre:
            flash('Hacen falta el nombre y el correo.', 'danger')
            return redirect('/admin/users')
        if app.supabase.get('users', {'email': email}, select='id'):
            flash(f'Ya existe una cuenta con el correo {email}.', 'warning')
            return redirect('/admin/users')

        temp = _gen_temp_password()
        role_ids = request.form.getlist('roles')
        creado = _crear_usuario({
            'email': email,
            'full_name': nombre,
            'password_hash': generate_password_hash(temp),
            'role': 'staff',
            'position': _sanitize(request.form.get('position', ''), 100) or None,
            'phone':    _sanitize(request.form.get('phone', ''), 40) or None,
            'must_change_password': True,
            'temp_password_expires': (datetime.now(timezone.utc) + timedelta(hours=72)).isoformat(),
            'password_reset_by': str(current_user.id),
            'is_active': True,
            'created_by_admin': str(current_user.id),
        })
        if not creado:
            flash('No se pudo crear la cuenta. Revisa el correo e inténtalo de nuevo.', 'danger')
            return redirect('/admin/users')

        uid = creado[0]['id']
        for rid in role_ids:
            app.supabase.insert_ignore('user_roles', {'user_id': uid, 'role_id': rid})
        if role_ids:
            # Rol activo de partida: sin esto la persona entra sin ningún rol
            # seleccionado y no ve nada, pese a tenerlos asignados.
            app.supabase.update('users', uid, {'active_role_id': role_ids[0]})
        _user_roles_cache.invalidate(uid)
        _log_password(app, uid, 'alta_admin', current_user.id)
        _auditar_acceso(app, 'alta', uid, email,
                        f'alta de {nombre}' + (f' con {len(role_ids)} rol(es)' if role_ids else ' sin roles'))

        flash(f'Cuenta creada para {nombre} ({email}). Contraseña temporal: {temp} — '
              f'entrégasela en persona. Caduca en 72 horas y deberá cambiarla al entrar. '
              f'No se volverá a mostrar.', 'success')
        return redirect('/admin/users')

    @app.route('/admin/users/<uid>/estado', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_user_estado(uid):
        """Activa o desactiva una cuenta sin borrarla.

        Hasta ahora la única forma de cortarle el acceso a alguien era
        ELIMINARLO, y con él se iban sus roles, sus permisos y su rastro. Quien
        se va de la oficina deja de entrar, pero lo que hizo tiene que quedar."""
        if not is_admin():
            return jsonify({'success': False, 'error': 'Solo admin'}), 403
        activar = bool((request.get_json() or {}).get('activar'))
        filas = app.supabase.get('users', {'id': uid}, select='id,email,full_name,role,is_active')
        if not filas:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        persona = filas[0]

        if not activar:
            if str(uid) == str(current_user.id):
                return jsonify({'success': False,
                                'error': 'No puedes desactivar tu propia cuenta.'})
            # Dejar el sistema sin ningún administrador que pueda entrar lo
            # convierte en irrecuperable desde la propia aplicación.
            if persona.get('role') == 'admin':
                otros = [u for u in (app.supabase.get('users', {'role': 'admin'},
                                                      select='id,is_active') or [])
                         if str(u['id']) != str(uid) and _cuenta_activa(u)]
                if not otros:
                    return jsonify({'success': False,
                                    'error': 'Es el único administrador activo: '
                                             'el sistema quedaría sin nadie que pueda entrar.'})

        if not app.supabase.update('users', uid, {'is_active': activar}):
            return jsonify({'success': False, 'error': 'No se pudo cambiar el estado.'})
        _auditar_acceso(app, 'estado', uid, persona.get('email'),
                        'cuenta activada' if activar else 'cuenta desactivada')
        return jsonify({'success': True, 'activa': activar})

    @app.route('/admin/users/<user_id>/reset-password', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_reset_password(user_id):
        """Recuperación de clave olvidada: el administrador genera una clave
        temporal de un solo uso. Se muestra UNA vez (no queda almacenada en
        claro) y caduca a las 72 h. El usuario debe cambiarla al entrar."""
        if not is_admin():
            return redirect('/dashboard')
        rows = app.supabase.get('users', {'id': user_id}, select='id,full_name,email')
        if not rows:
            flash('Usuario no encontrado.', 'danger')
            return redirect('/admin/users')
        temp = _gen_temp_password()
        app.supabase.update('users', user_id, {
            'password_hash': generate_password_hash(temp),
            'must_change_password': True,
            'temp_password_expires': (datetime.now(timezone.utc) + timedelta(hours=72)).isoformat(),
            'password_reset_by': current_user.id,
        })
        _log_password(app, user_id, 'reset_admin', current_user.id)
        _auditar_acceso(app, 'clave', user_id, rows[0].get('email'),
                        'contraseña restablecida por el administrador')
        nombre = rows[0].get('full_name') or rows[0].get('email')
        flash(f'Contraseña temporal de {nombre}: {temp} — entrégasela en persona. '
              f'Caduca en 72 horas y deberá cambiarla al entrar. No se volverá a mostrar.',
              'success')
        return redirect('/admin/users')

    # ============================================================
    #  ADMIN — ROLES (catálogo de roles: módulos + calendarios + proyectos + cuentas MS)
    # ============================================================
    @app.route('/admin/roles')
    @login_required
    def admin_roles():
        if not is_admin():
            return redirect('/dashboard')
        roles = app.supabase.get('roles', select='id,name,description,modules,level,created_at') or []
        all_cals = _get_calendar_config(app)
        all_projects = app.supabase.get('projects', select='id,name') or []
        ms_accounts = [t.get('email','') for t in (app.supabase.get('ms_tokens', select='email') or []) if t.get('email')]
        cal_ids   = app.supabase.get('role_calendars',   select='role_id,calendar_id') or []
        proj_ids  = app.supabase.get('role_projects',    select='role_id,project_id') or []
        ms_ids    = app.supabase.get('role_ms_accounts', select='role_id,ms_email') or []
        task_ids  = app.supabase.get('role_tasks',       select='role_id,task_id') or []
        cals_by_role = defaultdict(set); projs_by_role = defaultdict(set)
        ms_by_role = defaultdict(set);   tasks_by_role = defaultdict(set)
        for r in cal_ids:  cals_by_role[r['role_id']].add(r['calendar_id'])
        for r in proj_ids: projs_by_role[r['role_id']].add(r['project_id'])
        for r in ms_ids:   ms_by_role[r['role_id']].add(r['ms_email'])
        for r in task_ids: tasks_by_role[r['role_id']].add(r['task_id'])
        user_roles_all = app.supabase.get('user_roles', select='user_id,role_id') or []
        users_by_role = defaultdict(int)
        for ur in user_roles_all:
            users_by_role[ur['role_id']] += 1
        for r in roles:
            r['modules_list']  = [m for m in (r.get('modules') or '').split(',') if m]
            r['level']         = r.get('level') or DEFAULT_ROLE_LEVEL
            r['calendar_ids']  = cals_by_role.get(r['id'], set())
            r['project_ids']   = projs_by_role.get(r['id'], set())
            r['ms_emails']     = ms_by_role.get(r['id'], set())
            r['task_ids']      = tasks_by_role.get(r['id'], set())
            r['user_count']    = users_by_role.get(r['id'], 0)
        # Actividades (tareas de proyecto) agrupadas por proyecto, para marcarlas.
        all_tasks = app.supabase.get('tasks', select='id,title,project_id,phase') or []
        proj_name = {p['id']: p['name'] for p in all_projects}
        activities_by_project = defaultdict(list)
        for t in all_tasks:
            if t.get('project_id') and t['project_id'] in proj_name:
                activities_by_project[t['project_id']].append(t)
        projects_activities = [
            {'id': pid, 'name': proj_name[pid],
             'tasks': sorted(activities_by_project[pid], key=lambda x: (x.get('phase') or '', x.get('title') or ''))}
            for pid in proj_name if activities_by_project.get(pid)
        ]
        return render_template('admin_roles.html', roles=roles, calendarios=all_cals,
                               projects=all_projects, ms_accounts=ms_accounts,
                               all_modules=ALL_MODULES, role_levels=ROLE_LEVELS,
                               projects_activities=projects_activities)

    @app.route('/admin/roles/create', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_roles_create():
        if not is_admin(): return jsonify({'success': False})
        name = _sanitize(request.form.get('name', ''), 150)
        if not name:
            flash('El rol necesita un nombre.', 'danger')
            return redirect('/admin/roles')
        level = request.form.get('level', DEFAULT_ROLE_LEVEL)
        if level not in ROLE_LEVEL_IDS:
            level = DEFAULT_ROLE_LEVEL
        data = {
            'name': name,
            'description': _sanitize(request.form.get('description', ''), 500),
            'level': level,
            'modules': ','.join(request.form.getlist('modules')),
            'created_by': str(current_user.id),
        }
        created = app.supabase.insert('roles', data)
        role_id = created[0]['id'] if created else None
        if role_id:
            for cid in request.form.getlist('calendars'):
                app.supabase.insert_ignore('role_calendars', {'role_id': role_id, 'calendar_id': cid})
            for pid in request.form.getlist('projects'):
                app.supabase.insert_ignore('role_projects', {'role_id': role_id, 'project_id': pid})
            for ms in request.form.getlist('ms_accounts'):
                app.supabase.insert_ignore('role_ms_accounts', {'role_id': role_id, 'ms_email': ms})
            _save_role_activities(app, role_id, request.form.getlist('activities'))
        flash('Rol creado', 'success')
        return redirect('/admin/roles')

    @app.route('/admin/roles/update/<rid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_roles_update(rid):
        if not is_admin(): return jsonify({'success': False})
        name = _sanitize(request.form.get('name', ''), 150)
        if not name:
            flash('El rol necesita un nombre.', 'danger')
            return redirect('/admin/roles')
        level = request.form.get('level', DEFAULT_ROLE_LEVEL)
        if level not in ROLE_LEVEL_IDS:
            level = DEFAULT_ROLE_LEVEL
        app.supabase.update('roles', rid, {
            'name': name,
            'description': _sanitize(request.form.get('description', ''), 500),
            'level': level,
            'modules': ','.join(request.form.getlist('modules')),
        })
        for tbl, field, values in (
            ('role_calendars',   'calendar_id', request.form.getlist('calendars')),
            ('role_projects',    'project_id',  request.form.getlist('projects')),
            ('role_ms_accounts', 'ms_email',    request.form.getlist('ms_accounts')),
        ):
            for row in app.supabase.get(tbl, {'role_id': rid}, select='id'):
                app.supabase.delete(tbl, row['id'])
            for v in values:
                app.supabase.insert_ignore(tbl, {'role_id': rid, field: v})
        _save_role_activities(app, rid, request.form.getlist('activities'), replace=True)
        _role_cache.invalidate(rid)
        flash('Rol actualizado', 'success')
        return redirect('/admin/roles')

    @app.route('/admin/roles/delete/<rid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_roles_delete(rid):
        if not is_admin(): return jsonify({'success': False})
        app.supabase.delete('roles', rid)
        _role_cache.invalidate(rid)
        flash('Rol eliminado', 'success')
        return redirect('/admin/roles')

    # ============================================================
    #  ADMIN — DATABASE
    # ============================================================
    ADMIN_DB_TABLES = {'ciudades', 'appointment_titles', 'encargados', 'clients',
                        'appointments', 'calendar_config', 'users'}

    @app.route('/admin/database')
    @login_required
    def admin_database():
        if not is_admin():
            return redirect('/dashboard')
        return render_template('admin_database.html',
            users        = app.supabase.get('users', select='id,email,full_name,role,created_at'),
            ciudades     = app.supabase.get('ciudades'),
            titles       = app.supabase.get('appointment_titles'),
            encargados   = app.supabase.get('encargados'),
            clients      = app.supabase.get('clients'),
            appointments = app.supabase.get('appointments'),
            calendarios  = _get_calendar_config(app))

    @app.route('/admin/database/update', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_db_update():
        if not is_admin(): return jsonify({'success': False})
        table = request.form.get('table'); record_id = request.form.get('id')
        if table not in ADMIN_DB_TABLES:
            flash('Tabla no permitida.', 'danger')
            return redirect('/admin/database')
        data = {k: v for k, v in request.form.items() if k not in ['table', 'id']}
        if data: app.supabase.update(table, record_id, data)
        if table == 'calendar_config':
            _cal_cache.invalidate('all')
            _user_cal_cache.invalidate_prefix('')
        flash('Registro actualizado', 'success')
        return redirect('/admin/database')

    @app.route('/admin/database/delete', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_db_delete():
        if not is_admin(): return jsonify({'success': False})
        table = request.form.get('table')
        if table not in ADMIN_DB_TABLES:
            flash('Tabla no permitida.', 'danger')
            return redirect('/admin/database')
        record_id = request.form.get('id')
        if table == 'users':
            if record_id == str(current_user.id):
                flash('No puedes eliminarte a ti mismo.', 'danger')
                return redirect('/admin/database')
            _delete_user_cascade(app, record_id)
        else:
            app.supabase.delete(table, record_id)
        if table == 'calendar_config':
            _cal_cache.invalidate('all')
            _user_cal_cache.invalidate_prefix('')
        flash('Registro eliminado', 'success')
        return redirect('/admin/database')

    @app.route('/admin/database/insert', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_db_insert():
        if not is_admin(): return jsonify({'success': False})
        table = request.form.get('table')
        if table not in ADMIN_DB_TABLES:
            flash('Tabla no permitida.', 'danger')
            return redirect('/admin/database')
        data = {k: v for k, v in request.form.items() if k not in ['table']}
        if table == 'users' and data.get('password_hash'):
            data['password_hash'] = generate_password_hash(data['password_hash'])
        if data: app.supabase.insert(table, data)
        if table == 'calendar_config':
            _cal_cache.invalidate('all')
            _user_cal_cache.invalidate_prefix('')
        flash('Registro creado', 'success')
        return redirect('/admin/database')

    # ============================================================
    #  ADMIN — SINCRONIZACIÓN DE NAVEGADORES (Avast ⇄ Brave)
    #  Sólo el administrador dueño y sólo en la máquina Windows local.
    # ============================================================
    @app.route('/admin/browser-sync')
    @login_required
    def admin_browser_sync():
        if not browser_sync_allowed():
            return redirect('/dashboard')
        return render_template('admin_browser_sync.html',
                               st=_browser_sync.status())

    @app.route('/admin/browser-sync/run', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_browser_sync_run():
        if not browser_sync_allowed():
            return redirect('/dashboard')
        passphrase = request.form.get('passphrase', '')
        do_bm = request.form.get('do_bookmarks') == 'on'
        do_pw = request.form.get('do_passwords') == 'on'
        if not do_bm and not do_pw:
            flash('Selecciona al menos marcadores o contraseñas.', 'warning')
            return redirect('/admin/browser-sync')
        try:
            report = _browser_sync.run_sync(passphrase, do_bookmarks=do_bm,
                                            do_passwords=do_pw)
        except Exception as e:
            flash(f'Error durante la sincronización: {e}', 'danger')
            return redirect('/admin/browser-sync')
        for m in report.get('messages', []):
            flash(m, 'success' if report.get('ok') else 'warning')
        return render_template('admin_browser_sync.html',
                               st=_browser_sync.status(), report=report)

    @app.route('/admin/browser-sync/passwords/import', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_browser_sync_pw_import():
        if not browser_sync_allowed():
            return redirect('/dashboard')
        passphrase = request.form.get('passphrase', '')
        texts = []
        for f in request.files.getlist('csv_files'):
            if f and f.filename:
                try:
                    texts.append(f.read().decode('utf-8-sig', 'replace'))
                except Exception:
                    flash(f'No se pudo leer {f.filename}.', 'warning')
        if not texts:
            flash('Adjunta al menos un archivo CSV exportado del navegador.', 'warning')
            return redirect('/admin/browser-sync')
        rep = _browser_sync.import_password_csvs(passphrase, texts)
        for m in rep.get('messages', []):
            flash(m, 'success' if rep.get('ok') else 'danger')
        return redirect('/admin/browser-sync')

    @app.route('/admin/browser-sync/passwords/export', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_browser_sync_pw_export():
        if not browser_sync_allowed():
            return redirect('/dashboard')
        passphrase = request.form.get('passphrase', '')
        data, err = _browser_sync.export_password_csv(passphrase)
        if err:
            flash(err, 'danger')
            return redirect('/admin/browser-sync')
        return send_file(io.BytesIO(data), mimetype='text/csv',
                         as_attachment=True,
                         download_name='contrasenas_unificadas.csv')

    @app.route('/admin/user/update/<uid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_update_user(uid):
        if not is_admin(): return redirect('/dashboard')
        data = {}
        if request.form.get('full_name'):
            data['full_name'] = _sanitize(request.form.get('full_name'), 100)
        if request.form.get('email'):
            data['email'] = _sanitize(request.form.get('email'), 254).lower()
        if request.form.get('password'):
            data['password_hash'] = generate_password_hash(request.form.get('password'))
        if request.form.get('role'):
            data['role'] = request.form.get('role')
        if data: app.supabase.update('users', uid, data)
        # Roles asignados
        role_ids = request.form.getlist('roles')
        previos = {p['role_id'] for p in
                   (app.supabase.get('user_roles', {'user_id': uid}, select='id,role_id') or [])}
        for p in app.supabase.get('user_roles', {'user_id': uid}, select='id'):
            app.supabase.delete('user_roles', p['id'])
        for rid in role_ids:
            app.supabase.insert('user_roles', {'user_id': uid, 'role_id': rid})
        _user_roles_cache.invalidate(uid)
        _user_cal_cache.invalidate(uid)
        if previos != set(role_ids):
            nombres = {r['id']: r['name'] for r in
                       (app.supabase.get('roles', select='id,name') or [])}
            _auditar_acceso(app, 'roles', uid, data.get('email'), 'roles: ' + (
                ', '.join(nombres.get(r, r) for r in role_ids) or 'ninguno'))
        flash('Usuario actualizado', 'success')
        return redirect('/admin/users')

    @app.route('/admin/user/delete/<uid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_delete_user(uid):
        if not is_admin(): return jsonify({'success': False, 'error': 'Sin autorización'})
        if uid == str(current_user.id):
            return jsonify({'success': False, 'error': 'No puedes eliminarte a ti mismo'})
        filas = app.supabase.get('users', {'id': uid}, select='email,full_name')
        ok = _delete_user_cascade(app, uid)
        if ok:
            # Se apunta ANTES de que el correo deje de poder consultarse. La
            # tabla de auditoría no tiene clave foránea a users justamente para
            # que la baja sobreviva a la persona.
            _auditar_acceso(app, 'baja', uid, (filas[0].get('email') if filas else None),
                            'cuenta eliminada' + (f" ({filas[0].get('full_name')})" if filas else ''))
        return jsonify({'success': ok, 'error': None if ok else 'No se pudo eliminar el usuario'})

    @app.route('/admin/approve-one/<pid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_approve_one(pid):
        if not is_admin(): return jsonify({'success': False})
        rows = app.supabase.get('calendar_permissions', {'id': pid}, select='id,user_id,calendar_id')
        app.supabase.update('calendar_permissions', pid, {'status': 'approved'})
        if rows:
            _grant_calendar_via_role(app, rows[0]['user_id'], rows[0]['calendar_id'])
        _user_cal_cache.invalidate_prefix('')  # any user might be affected
        return jsonify({'success': True})

    @app.route('/admin/reject-one/<pid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_reject_one(pid):
        if not is_admin(): return jsonify({'success': False})
        app.supabase.update('calendar_permissions', pid, {'status': 'rejected'})
        _user_cal_cache.invalidate_prefix('')
        return jsonify({'success': True})

    @app.route('/admin/approve-all/<uid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_approve_all(uid):
        if not is_admin(): return jsonify({'success': False})
        for p in app.supabase.get('calendar_permissions',
                {'user_id': uid, 'status': 'pending'}, select='id,calendar_id'):
            app.supabase.update('calendar_permissions', p['id'], {'status': 'approved'})
            _grant_calendar_via_role(app, uid, p['calendar_id'])
        _user_cal_cache.invalidate(uid)
        return jsonify({'success': True})

    @app.route('/admin/reject-all/<uid>', methods=['POST'])
    @login_required
    @csrf_protect
    def admin_reject_all(uid):
        if not is_admin(): return jsonify({'success': False})
        for p in app.supabase.get('calendar_permissions',
                {'user_id': uid, 'status': 'pending'}, select='id'):
            app.supabase.update('calendar_permissions', p['id'], {'status': 'rejected'})
        _user_cal_cache.invalidate(uid)
        return jsonify({'success': True})

    # ============================================================
    #  DASHBOARD  (optimized: O(3) queries instead of O(2N+3))
    # ============================================================
    def _dashboard_widgets(app):
        """Calcula las cifras-tarjeta del panel para el usuario logueado."""
        today_iso = date.today().isoformat()
        in_7d_iso = (date.today() + timedelta(days=7)).isoformat()
        # Tareas: aplica los mismos permisos que /planning/api/tasks (rol activo + proyecto)
        all_tasks = app.supabase.get('tasks',
            select='id,status,due_date,priority,source,ms_email,created_by,assigned_to,assigned_email,subtasks,project_id') or []
        all_tasks = _filter_visible_tasks(app, all_tasks, current_user.id)
        pending_all   = [t for t in all_tasks if t.get('status') != 'done']
        overdue       = [t for t in pending_all if t.get('due_date') and t['due_date'] < today_iso]
        today_tasks   = [t for t in pending_all if t.get('due_date') == today_iso]
        week_tasks    = [t for t in pending_all if t.get('due_date') and today_iso <= t['due_date'] <= in_7d_iso]
        manual_pend   = [t for t in pending_all if t.get('source') != 'ms_todo']
        todo_pend     = [t for t in pending_all if t.get('source') == 'ms_todo']
        # Subtareas pendientes (suma global)
        sub_pending = 0
        for t in pending_all:
            for s in (t.get('subtasks') or []):
                if not s.get('done'): sub_pending += 1
        # Citas próximas (siguientes 7 días)
        next_apts = []
        try:
            apt_rows = app.supabase.get('appointments', select='id,title,start_time,status,encargado,calendar_id') or []
            # Filtra por permisos de calendarios del usuario si no es admin
            if not is_admin():
                user_cal_ids = {c.get('calendar_id') for c in get_user_calendars(app, current_user.id)}
                apt_rows = [a for a in apt_rows if a.get('calendar_id') in user_cal_ids]
            for a in apt_rows:
                st = (a.get('start_time') or '')[:10]
                if st and today_iso <= st <= in_7d_iso and a.get('status') != 'cancelled':
                    next_apts.append(a)
            next_apts.sort(key=lambda x: x.get('start_time') or '')
            next_apts = next_apts[:5]
        except Exception:
            next_apts = []
        # Cuentas MS conectadas (solo admin)
        ms_accounts = []
        if is_admin():
            ms_accounts = [t.get('email','') for t in (app.supabase.get('ms_tokens', select='email') or []) if t.get('email')]

        # Directorio y Cronograma. Se consultan sólo si el usuario tiene el
        # módulo: al que no lo tiene no se le hace pagar la consulta, y las
        # tablas pueden no existir todavía si aún no se aplicó la migración 020
        # —de ahí el try, para que el panel no se caiga por eso—.
        contactos_total = sectores_total = 0
        if user_can('directorio'):
            try:
                contactos_total = len(app.supabase.get('contacts', select='id') or [])
                sectores_total = len([s for s in (app.supabase.get('sectors', select='id,active') or [])
                                      if s.get('active') is not False])
            except Exception:
                pass
        planes_total = actividades_vencidas = 0
        if user_can('cronograma'):
            try:
                planes = app.supabase.get('gantt_plans', select='id,created_by,status') or []
                if not is_admin():
                    planes = [p for p in planes if p.get('created_by') == str(current_user.id)]
                planes = [p for p in planes if p.get('status') != 'archived']
                planes_total = len(planes)
                ids_planes = {p['id'] for p in planes}
                for a in (app.supabase.get('gantt_activities', select='plan_id,status,end_date') or []):
                    if (a.get('plan_id') in ids_planes and a.get('status') != 'done'
                            and a.get('end_date') and a['end_date'] < today_iso):
                        actividades_vencidas += 1
            except Exception:
                pass

        return {
            'total_pending':  len(pending_all),
            'overdue':        len(overdue),
            'today_count':    len(today_tasks),
            'week_count':     len(week_tasks),
            'manual_pending': len(manual_pend),
            'todo_pending':   len(todo_pend),
            'sub_pending':    sub_pending,
            'next_apts':      next_apts,
            'ms_accounts':    ms_accounts,
            'contactos':      contactos_total,
            'sectores':       sectores_total,
            'planes':         planes_total,
            'gantt_vencidas': actividades_vencidas,
        }

    @app.route('/dashboard')
    @login_required
    def dashboard():
        if is_admin():
            all_cals   = _get_calendar_config(app)
            all_pending = app.supabase.get('calendar_permissions',
                {'status': 'pending'}, select='id,user_id,calendar_id')
            # Fetch only the users that appear in pending list (single IN query)
            pending_uids = list({p['user_id'] for p in all_pending})
            if pending_uids:
                pend_users = app.supabase.get_in('users', 'id', pending_uids,
                    select='id,full_name,email')
                user_by_id = {u['id']: u for u in pend_users}
            else:
                user_by_id = {}
            cal_by_id   = {c['calendar_id']: c for c in all_cals}
            by_user     = defaultdict(list)
            pending_all = []
            for p in all_pending:
                pending_all.append({'id': p['id'], 'user_id': p['user_id'],
                                    'calendar_id': p['calendar_id']})
                by_user[p['user_id']].append(p)
            pending = []
            for uid, perms_list in by_user.items():
                u = user_by_id.get(uid, {})
                pending.append({
                    'user_id':    uid,
                    'user_name':  u.get('full_name', ''),
                    'user_email': u.get('email', ''),
                    'calendars':  [cal_by_id[p['calendar_id']] for p in perms_list
                                   if p['calendar_id'] in cal_by_id],
                })
            cals = all_cals
        else:
            cals = get_user_calendars(app, current_user.id)
            pending = []; pending_all = []
        google_ok = get_google_creds(app) is not None
        widgets = _dashboard_widgets(app)
        return render_template('dashboard.html', calendarios=cals, pending=pending,
                               pending_all=pending_all, google_connected=google_ok,
                               widgets=widgets,
                               can_planning=user_can('planning'),
                               can_todo=user_can('todo'),
                               can_calendar=user_can('calendar'),
                               can_directorio=user_can('directorio'),
                               can_cronograma=user_can('cronograma'))

    # ============================================================
    #  CALENDAR VIEW
    # ============================================================
    @app.route('/calendar')
    @login_required
    def calendar():
        if not user_can('calendar'):
            flash('No tienes acceso al módulo Calendario.', 'warning')
            return redirect('/dashboard')
        cals = (_get_calendar_config(app) if is_admin()
                else get_user_calendars(app, current_user.id))
        return render_template('calendar.html', calendarios=cals,
                               google_connected=get_google_creds(app) is not None)

    # ============================================================
    #  API — EVENTS  (single query with IN filter for non-admin)
    # ============================================================
    APPT_SELECT_BASE = ('id,title,encargado,start_time,end_time,status,calendar_id,'
                        'tema,client_name,client_email,notes,lugar,direccion,mapa,'
                        'ciudad,meeting_link,google_event_id')
    # Columnas de recurrencia — requieren la migración 002. El SELECT cae al
    # base automáticamente si todavía no existen (ver _events_query abajo).
    APPT_SELECT = APPT_SELECT_BASE + ',is_recurring,parent_event_id'

    @app.route('/calendar/api/events')
    @login_required
    def api_events():
        def _events_query(fetch):
            # fetch(select) -> lista. Intenta con columnas de recurrencia;
            # si vuelve vacío (p.ej. columna inexistente antes de migrar),
            # reintenta con el SELECT base para no ocultar los eventos.
            rows = fetch(APPT_SELECT)
            if not rows:
                rows = fetch(APPT_SELECT_BASE)
            return rows

        if is_admin():
            events = _events_query(
                lambda sel: app.supabase.get('appointments', select=sel))
        else:
            ucal = [c['calendar_id'] for c in get_user_calendars(app, current_user.id)]
            if not ucal:
                return jsonify([])
            # One query with IN — replaces N separate queries
            events = _events_query(
                lambda sel: app.supabase.get_in('appointments', 'calendar_id', ucal, select=sel))
        colors = {'pending': '#f59e0b', 'confirmed': '#10b981', 'cancelled': '#ef4444'}
        result = []
        for e in events:
            is_rec = e.get('is_recurring', False)
            result.append({
                'id': e['id'],
                'title': f"{'R ' if is_rec else ''}{e['title']} — {e.get('encargado', '')}",
                'start': e['start_time'], 'end': e['end_time'],
                'backgroundColor': colors.get(e.get('status'), '#3b82f6'),
                'borderColor':     colors.get(e.get('status'), '#3b82f6'),
                'extendedProps': {
                    'title': e.get('title', ''), 'encargado': e.get('encargado', ''),
                    'tema': e.get('tema', ''), 'client_name': e.get('client_name', ''),
                    'client_email': e.get('client_email', ''),
                    'status': e.get('status', 'pending'),
                    'calendar_id': e.get('calendar_id', ''),
                    'notes': e.get('notes', ''), 'lugar': e.get('lugar', ''),
                    'direccion': e.get('direccion', ''), 'mapa': e.get('mapa', ''),
                    'ciudad': e.get('ciudad', ''), 'meeting_link': e.get('meeting_link', ''),
                    'google_event_id': e.get('google_event_id', ''),
                    'is_recurring': is_rec,
                    'parent_event_id': e.get('parent_event_id', ''), 'id': e['id'],
                },
            })
        return jsonify(result)

    # ============================================================
    #  API — FERIADOS ECUADOR (cálculo automático + traslados de ley)
    # ============================================================
    @app.route('/calendar/api/holidays')
    @login_required
    def api_holidays():
        args = request.args
        start = args.get('start')  # FullCalendar envía start/end del rango visible
        end   = args.get('end')
        try:
            if start and end:
                d0 = date.fromisoformat(start[:10])
                d1 = date.fromisoformat(end[:10])
                data = _feriados_rango(d0, d1)
            else:
                y0 = int(args.get('y_from') or args.get('year') or date.today().year)
                y1 = int(args.get('y_to')   or args.get('year') or y0)
                if y1 < y0:
                    y0, y1 = y1, y0
                y1 = min(y1, y0 + 10)  # tope defensivo
                data = []
                for y in range(y0, y1 + 1):
                    data.extend(_feriados_ec(y))
        except Exception:
            data = _feriados_ec(date.today().year)

        result = []
        for f in data:
            es_local = f.get('ambito') == 'local'
            # Ámbito explícito en la etiqueta: "Nacional" o "Local (Ciudad)"
            if es_local:
                tipo = 'Local' + (' (' + f['ciudad'] + ')' if f.get('ciudad') else '')
            else:
                tipo = 'Nacional'
            etiqueta = f['nombre'] + (' (traslado)' if f['trasladado'] else '') + ' · ' + tipo
            clases = ['feriado-ec']
            clases.append('feriado-local' if es_local else 'feriado-nac')
            if not f['verificado']:
                clases.append('feriado-proj')
            # Nacional=rojo, Local=morado; proyección=ámbar
            if not f['verificado']:
                color = '#f59e0b'
            else:
                color = '#7c3aed' if es_local else '#e11d48'
            result.append({
                'id': 'hol-' + f['fecha'],
                'title': ('🏙️ ' if es_local else '🇪🇨 ') + etiqueta,
                'start': f['fecha'],
                'allDay': True,
                'display': 'block',
                'editable': False,
                'classNames': clases,
                'color': color,
                'extendedProps': {
                    'isHoliday': True,
                    'nombre': f['nombre'],
                    'fecha_real': f['fecha_real'],
                    'trasladado': f['trasladado'],
                    'verificado': f['verificado'],
                    'dia_semana': f['dia_semana'],
                    'ambito': f.get('ambito', 'nacional'),
                    'ciudad': f.get('ciudad'),
                },
            })
        return jsonify(result)

    # ============================================================
    #  API — LOOKUP DATA (projected, browser-cached 60s)
    # ============================================================
    @app.route('/calendar/api/titles')
    @login_required
    def api_titles():
        return jsonify([t['title'] for t in
            app.supabase.get('appointment_titles', select='title')])

    @app.route('/calendar/api/encargados')
    @login_required
    def api_encargados():
        return jsonify([e['name'] for e in
            app.supabase.get('encargados', select='name')])

    @app.route('/calendar/api/temas')
    @login_required
    def api_temas():
        return jsonify([t['description'] for t in
            app.supabase.get('temas', select='description')])

    @app.route('/calendar/api/clients')
    @login_required
    def api_clients():
        rows = app.supabase.get('clients', select='name,email,calendar_id')
        if not is_admin():
            allowed = {c['calendar_id'] for c in get_user_calendars(app, current_user.id)}
            # Los clientes sin calendar_id son heredados de antes de esta restricción:
            # se siguen mostrando a todos para no romper el autocompletado existente.
            rows = [c for c in rows if not c.get('calendar_id') or c['calendar_id'] in allowed]
        return jsonify([{'name': c['name'], 'email': c.get('email', '')} for c in rows])

    @app.route('/calendar/api/ciudades')
    @login_required
    def api_ciudades():
        return jsonify([c['name'] for c in
            app.supabase.get('ciudades', select='name')])

    # ============================================================
    #  API — PENDING
    # ============================================================
    @app.route('/calendar/api/pending')
    @login_required
    def api_pending():
        if is_admin():
            pending = [a for a in
                app.supabase.get('appointments',
                    {'status': 'pending'},
                    select='id,title,encargado,tema,client_name,start_time')
            ]
        else:
            ucal = [c['calendar_id'] for c in get_user_calendars(app, current_user.id)]
            if not ucal:
                return jsonify([])
            all_p = app.supabase.get_in('appointments', 'calendar_id', ucal,
                select='id,title,encargado,tema,client_name,start_time,calendar_id,status')
            pending = [a for a in all_p if a.get('status') == 'pending']
        return jsonify([{
            'id': a['id'], 'title': a['title'], 'encargado': a.get('encargado', ''),
            'tema': a.get('tema', ''), 'client_name': a.get('client_name', ''),
            'date': a['start_time'].split('T')[0],
            'time': a['start_time'].split('T')[1][:5],
            'is_recurring': a.get('is_recurring', False),
        } for a in pending])

    # ============================================================
    #  API — BOOK  (optimized: insert_ignore replaces check+insert)
    # ============================================================
    @app.route('/calendar/api/book', methods=['POST'])
    @login_required
    def api_book():
        try:
            date_str = request.form.get('date', '').strip()
            time_str = request.form.get('time', '').strip()
            dur_sel  = request.form.get('duration', '30')
            dur      = max(15, min(1440, int(request.form.get('custom_duration', dur_sel) or 30)))

            title      = _sanitize(request.form.get('title', ''), 200).upper()
            cal_id     = _sanitize(request.form.get('calendar_id', ''), 100)
            encargado  = _sanitize(request.form.get('encargado', ''), 100).upper()
            tema       = _sanitize(request.form.get('tema', ''), 300)
            client_name  = _sanitize(request.form.get('client_name', ''), 150).upper()
            client_email = _sanitize(request.form.get('client_email', ''), 254).lower()
            notificar    = [e.strip() for e in request.form.getlist('notificar') if e.strip()]
            tipo      = request.form.get('type', 'presencial')
            lugar     = _sanitize(request.form.get('lugar', ''), 150).upper()
            direccion = _sanitize(request.form.get('direccion', ''), 300)
            mapa      = _sanitize(request.form.get('mapa', ''), 500)
            ciudad    = _sanitize(request.form.get('ciudad', 'CUENCA'), 100).upper()
            link      = _sanitize(request.form.get('meeting_link', ''), 500)
            notes     = _sanitize(request.form.get('notes', ''), 1000)

            sessions_present = bool(request.form.get('sessions', '').strip())
            if not all([title, cal_id, encargado, tema]) or \
               (not sessions_present and not all([date_str, time_str])):
                return jsonify({'success': False, 'error': 'Faltan campos obligatorios'})
            if not is_admin() and not user_has_calendar_access(app, current_user.id, cal_id):
                return jsonify({'success': False, 'error': 'Sin autorizacion para este calendario'})

            # Upsert lookup tables — insert_ignore skips if already exists
            if ciudad:     app.supabase.insert_ignore('ciudades', {'name': ciudad})
            if title:      app.supabase.insert_ignore('appointment_titles', {'title': title})
            if encargado:  app.supabase.insert_ignore('encargados', {'name': encargado})
            if tema:       app.supabase.insert_ignore('temas', {'description': tema})
            if client_name:
                app.supabase.insert_ignore('clients',
                    {'name': client_name, 'email': client_email, 'created_by': current_user.id,
                     'calendar_id': cal_id})

            # ---- Recurring (flexible: daily/weekly/monthly/yearly) ----
            is_recurring = request.form.get('is_recurring') == 'true'
            if is_recurring:
                freq     = request.form.get('rec_freq', 'weekly')
                end_mode = request.form.get('rec_end_mode', 'until')
                try:
                    interval  = max(1, min(366, int(request.form.get('rec_interval', '1') or 1)))
                    start_d   = datetime.strptime(date_str, '%Y-%m-%d').date()
                    weekdays  = json.loads(request.form.get('rec_weekdays', '[]') or '[]')
                    rec_count = max(1, min(REC_HARD_CAP, int(request.form.get('rec_count', '1') or 1)))
                    rec_end   = None
                    if end_mode == 'until':
                        rec_end = datetime.strptime(
                            request.form.get('rec_end_date', ''), '%Y-%m-%d').date()
                except Exception as ex:
                    return jsonify({'success': False, 'error': f'Datos de recurrencia invalidos: {ex}'})

                if freq not in ('daily', 'weekly', 'monthly', 'yearly'):
                    return jsonify({'success': False, 'error': 'Frecuencia invalida'})
                if end_mode == 'until' and (rec_end is None or rec_end < start_d):
                    return jsonify({'success': False, 'error': 'Fecha fin debe ser posterior a inicio'})
                if freq == 'weekly' and weekdays and any(w < 0 or w > 6 for w in weekdays):
                    return jsonify({'success': False, 'error': 'Dias de semana invalidos'})

                dates_to_create = _generate_recurrence_dates(
                    start_d, freq, interval, weekdays, end_mode, rec_end, rec_count)
                if not dates_to_create:
                    return jsonify({'success': False, 'error': 'La recurrencia no genera ninguna fecha'})

                rule_json = json.dumps({
                    'freq': freq, 'interval': interval, 'weekdays': weekdays,
                    'end_mode': end_mode,
                    'end_date': rec_end.isoformat() if rec_end else None,
                    'count': rec_count,
                }, ensure_ascii=False)
                # Aviso si se topó el límite de materialización (recurrencia muy larga/indefinida)
                capped = len(dates_to_create) >= REC_HARD_CAP
                rec_notes = f'[SERIE {len(dates_to_create)} eventos] {notes}'.strip()

                created_ids = []; parent_id = None
                for d in dates_to_create:
                    local_dt = TIMEZONE.localize(
                        datetime.strptime(f'{d.isoformat()} {time_str}:00', '%Y-%m-%d %H:%M:%S'))
                    s_dt = local_dt.astimezone(pytz.UTC)
                    e_dt = s_dt + timedelta(minutes=dur)
                    record = _build_appointment(title, cal_id, encargado, tema,
                        client_name, client_email, s_dt, e_dt, tipo, link,
                        lugar, direccion, mapa, ciudad, notificar, rec_notes, current_user.id)
                    record['is_recurring'] = True
                    if parent_id:
                        record['parent_event_id'] = parent_id
                    else:
                        record['recurrence_rule'] = rule_json
                        if rec_end:
                            record['recurrence_end_date'] = rec_end.isoformat()
                    r = app.supabase.insert('appointments', record)
                    if not r:
                        # Fallback si aún no se corrió la migración de columnas de recurrencia
                        for col in ('is_recurring', 'parent_event_id',
                                    'recurrence_rule', 'recurrence_end_date'):
                            record.pop(col, None)
                        r = app.supabase.insert('appointments', record)
                    if r:
                        aid = r[0]['id']
                        created_ids.append(aid)
                        if parent_id is None:
                            parent_id = aid
                            try:
                                app.supabase.update('appointments', aid, {'parent_event_id': aid})
                            except Exception:
                                pass
                if created_ids:
                    return jsonify({'success': True, 'count': len(created_ids),
                                    'recurring': True, 'capped': capped})
                return jsonify({'success': False, 'error': 'No se pudieron crear los eventos'})

            # ---- Varias fechas / sesiones manuales (distinta hora y duración) ----
            if sessions_present:
                try:
                    sessions = json.loads(request.form.get('sessions', '[]'))
                except Exception:
                    return jsonify({'success': False, 'error': 'Sesiones invalidas'})
                if not isinstance(sessions, list) or not sessions:
                    return jsonify({'success': False, 'error': 'Agrega al menos una fecha'})
                if len(sessions) > 60:
                    return jsonify({'success': False, 'error': 'Maximo 60 sesiones por serie'})

                parsed = []
                for s in sessions:
                    try:
                        sd = datetime.strptime((s.get('date') or ''), '%Y-%m-%d').date()
                        st = (s.get('time') or '')
                        sdur = max(15, min(1440, int(s.get('duration', 60) or 60)))
                        local_dt = TIMEZONE.localize(datetime.strptime(
                            f'{sd.isoformat()} {st}:00', '%Y-%m-%d %H:%M:%S'))
                    except Exception:
                        return jsonify({'success': False,
                                        'error': 'Fecha u hora invalida en una sesion'})
                    parsed.append((local_dt, sdur))
                parsed.sort(key=lambda x: x[0])

                ses_notes = f'[SERIE {len(parsed)} sesiones] {notes}'.strip()
                created_ids = []; parent_id = None
                for local_dt, sdur in parsed:
                    s_dt = local_dt.astimezone(pytz.UTC)
                    e_dt = s_dt + timedelta(minutes=sdur)
                    record = _build_appointment(title, cal_id, encargado, tema,
                        client_name, client_email, s_dt, e_dt, tipo, link,
                        lugar, direccion, mapa, ciudad, notificar, ses_notes, current_user.id)
                    record['is_recurring'] = True
                    if parent_id:
                        record['parent_event_id'] = parent_id
                    r = app.supabase.insert('appointments', record)
                    if not r:
                        for col in ('is_recurring', 'parent_event_id'):
                            record.pop(col, None)
                        r = app.supabase.insert('appointments', record)
                    if r:
                        aid = r[0]['id']; created_ids.append(aid)
                        if parent_id is None:
                            parent_id = aid
                            try:
                                app.supabase.update('appointments', aid, {'parent_event_id': aid})
                            except Exception:
                                pass
                if created_ids:
                    return jsonify({'success': True, 'count': len(created_ids), 'recurring': True})
                return jsonify({'success': False, 'error': 'No se pudieron crear las sesiones'})

            # ---- Single event ----
            local_dt = TIMEZONE.localize(
                datetime.strptime(f'{date_str} {time_str}:00', '%Y-%m-%d %H:%M:%S'))
            s_dt = local_dt.astimezone(pytz.UTC)
            e_dt = s_dt + timedelta(minutes=dur)
            record = _build_appointment(title, cal_id, encargado, tema,
                client_name, client_email, s_dt, e_dt, tipo, link,
                lugar, direccion, mapa, ciudad, notificar, notes, current_user.id)
            result = app.supabase.insert('appointments', record)
            if result:
                return jsonify({'success': True, 'id': result[0]['id']})
            return jsonify({'success': False, 'error': 'Error en base de datos'})

        except Exception as e:
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)})

    # ============================================================
    #  API — APPROVE / REJECT / DELETE / DELETE-SERIES / SYNC
    # ============================================================
    @app.route('/calendar/api/approve/<aid>', methods=['POST'])
    @login_required
    def api_approve(aid):
        if not user_can('calendar.aprobar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para aprobar citas.'}), 403
        apts = app.supabase.get('appointments', {'id': aid},
            select='id,title,encargado,tema,client_name,client_email,start_time,end_time,'
                   'status,calendar_id,invitados,lugar,direccion,ciudad,mapa,notes,'
                   'meeting_link,google_event_id,google_cal_id')
        if not apts: return jsonify({'success': False})
        apt = apts[0]
        if not is_admin() and not user_has_calendar_access(app, current_user.id, apt.get('calendar_id')):
            return jsonify({'success': False, 'error': 'Sin autorizacion'})

        # Idempotente: si ya tiene evento en Google, solo confirmar — no duplicar
        if apt.get('google_event_id'):
            app.supabase.update('appointments', aid, {'status': 'confirmed'})
            return jsonify({'success': True, 'message': 'Confirmada (ya sincronizada con Google)'})

        # Reclamo atómico: solo una petición concurrente puede pasar de 'pending' a
        # 'confirmed' (usamos el mismo valor de status final, sin inventar uno nuevo,
        # para no depender de qué valores acepte un posible CHECK constraint en la BD).
        # Evita que dos aprobaciones simultáneas creen el evento de Google dos veces.
        claimed = app.supabase.update_where('appointments',
            {'id': aid, 'status': 'pending'}, {'status': 'confirmed'})
        if not claimed:
            return jsonify({'success': False, 'error': 'Esta cita ya fue procesada por otra solicitud'})

        creds = get_google_creds(app)
        if not creds:
            app.supabase.update('appointments', aid, {'status': 'confirmed'})
            return jsonify({'success': True, 'message': 'Aprobada (sin sincronizacion Google)'})
        try:
            service  = build('calendar', 'v3', credentials=creds)
            all_cals = _get_calendar_config(app)
            email_map, gcal_id_map = _make_cal_maps(all_cals)
            cal_id  = apt.get('calendar_id')
            gcal_id = gcal_id_map.get(cal_id, 'primary')
            attendees = _build_attendees(apt, email_map)
            event = _build_google_event(apt, attendees)
            # Buscar si ya existe en Google Calendar para evitar duplicado
            existing = service.events().list(
                calendarId=gcal_id, timeMin=apt['start_time'],
                timeMax=apt['end_time'], q=apt['title'], maxResults=1).execute()
            if existing.get('items'):
                # Ya existe: vincular sin reenviar notificaciones
                gev_id = existing['items'][0]['id']
                app.supabase.update('appointments', aid,
                    {'status': 'confirmed', 'google_event_id': gev_id, 'google_cal_id': gcal_id})
                return jsonify({'success': True, 'message': 'Confirmada (evento ya existía en Google)'})
            # Nuevo evento — notificar a todos los asistentes una sola vez
            created = service.events().insert(
                calendarId=gcal_id, body=event, sendUpdates='all').execute()
            app.supabase.update('appointments', aid,
                {'status': 'confirmed', 'google_event_id': created.get('id'),
                 'google_cal_id': gcal_id})
            return jsonify({'success': True,
                'message': f'Aprobada — {len(attendees)} invitado(s) notificado(s)'})
        except google.auth.exceptions.RefreshError:
            app.supabase.update('appointments', aid, {'status': 'confirmed'})
            return jsonify({'success': True,
                'message': 'Aprobada. Reconecta Google en /auth/google para sincronizar.'})
        except Exception as e:
            if _is_invalid_grant(e):
                app.supabase.update('appointments', aid, {'status': 'confirmed'})
                return jsonify({'success': True, 'message': 'Aprobada. Reconecta Google.'})
            # No dejar la cita atascada en 'approving': volver a 'pending' para poder reintentar.
            app.supabase.update('appointments', aid, {'status': 'pending'})
            return jsonify({'success': False, 'error': str(e)})

    @app.route('/calendar/api/reject/<aid>', methods=['POST'])
    @login_required
    def api_reject(aid):
        # Rechazar es la otra mitad de aprobar: el mismo permiso decide las dos.
        if not user_can('calendar.aprobar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para rechazar citas.'}), 403
        apts = app.supabase.get('appointments', {'id': aid}, select='id,calendar_id')
        if not apts: return jsonify({'success': False})
        if not is_admin() and not user_has_calendar_access(app, current_user.id, apts[0].get('calendar_id')):
            return jsonify({'success': False, 'error': 'Sin autorizacion'})
        app.supabase.update('appointments', aid, {'status': 'cancelled'})
        return jsonify({'success': True})

    @app.route('/calendar/api/delete/<aid>', methods=['POST'])
    @login_required
    def api_delete(aid):
        if not user_can('calendar.eliminar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para eliminar citas.'}), 403
        apts = app.supabase.get('appointments', {'id': aid},
            select='id,calendar_id,google_event_id,google_cal_id')
        if not apts: return jsonify({'success': False})
        apt = apts[0]
        if not is_admin() and not user_has_calendar_access(app, current_user.id, apt.get('calendar_id')):
            return jsonify({'success': False, 'error': 'Sin autorizacion'})
        if apt.get('google_event_id'):
            creds = get_google_creds(app)
            if creds:
                gcal_id = apt.get('google_cal_id') or 'primary'
                try:
                    build('calendar', 'v3', credentials=creds).events().delete(
                        calendarId=gcal_id, eventId=apt['google_event_id']).execute()
                except Exception:
                    pass
        app.supabase.delete('appointments', aid)
        return jsonify({'success': True})

    @app.route('/calendar/api/delete-series/<parent_id>', methods=['POST'])
    @login_required
    def api_delete_series(parent_id):
        # Borrar una serie entera es borrar, sólo que muchas veces de golpe.
        if not user_can('calendar.eliminar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para eliminar citas.'}), 403
        all_apts = app.supabase.get('appointments',
            select='id,calendar_id,google_event_id,google_cal_id,parent_event_id')
        series = [a for a in all_apts
                  if a.get('parent_event_id') == parent_id or a.get('id') == parent_id]
        if not series: return jsonify({'success': False, 'error': 'Serie no encontrada'})
        if not is_admin() and not user_has_calendar_access(
                app, current_user.id, series[0].get('calendar_id')):
            return jsonify({'success': False, 'error': 'Sin autorizacion'})
        creds = get_google_creds(app); deleted = 0
        for apt in series:
            if apt.get('google_event_id') and creds:
                gcal_id = apt.get('google_cal_id') or 'primary'
                try:
                    build('calendar', 'v3', credentials=creds).events().delete(
                        calendarId=gcal_id, eventId=apt['google_event_id']).execute()
                except Exception:
                    pass
            app.supabase.delete('appointments', apt['id']); deleted += 1
        return jsonify({'success': True, 'deleted': deleted})

    @app.route('/calendar/api/sync', methods=['POST'])
    @login_required
    def api_sync():
        if not is_admin(): return jsonify({'success': False, 'error': 'Solo admin'})
        creds = get_google_creds(app)
        if not creds: return jsonify({'success': False, 'error': 'Google no conectado'})
        synced = 0; errors = 0; skipped = 0
        all_cals = _get_calendar_config(app)
        email_map, gcal_id_map = _make_cal_maps(all_cals)
        service = build('calendar', 'v3', credentials=creds)
        for apt in app.supabase.get('appointments',
                select='id,title,encargado,tema,client_name,start_time,end_time,calendar_id,'
                       'invitados,direccion,ciudad,lugar,mapa,notes,meeting_link,status,google_event_id'):
            if apt.get('status') != 'confirmed' or apt.get('google_event_id'):
                continue
            try:
                cal_id  = apt.get('calendar_id')
                gcal_id = gcal_id_map.get(cal_id, 'primary')
                existing = service.events().list(
                    calendarId=gcal_id, timeMin=apt['start_time'],
                    timeMax=apt['end_time'], q=apt['title'], maxResults=1).execute()
                if existing.get('items'):
                    app.supabase.update('appointments', apt['id'],
                        {'google_event_id': existing['items'][0]['id'],
                         'google_cal_id': gcal_id})
                    skipped += 1; continue
                attendees = _build_attendees(apt, email_map)
                event = _build_google_event(apt, attendees)
                created = service.events().insert(calendarId=gcal_id,
                    body=event, sendUpdates='all').execute()
                app.supabase.update('appointments', apt['id'],
                    {'google_event_id': created.get('id'), 'google_cal_id': gcal_id})
                synced += 1
            except google.auth.exceptions.RefreshError:
                return jsonify({'success': False, 'synced': synced, 'skipped': skipped,
                    'errors': errors, 'error': 'Google desconectado. Reconecta en /auth/google.'})
            except Exception as e:
                if _is_invalid_grant(e):
                    return jsonify({'success': False, 'synced': synced, 'skipped': skipped,
                        'errors': errors, 'error': 'Google desconectado. Reconecta en /auth/google.'})
                errors += 1
        return jsonify({'success': True, 'synced': synced, 'skipped': skipped, 'errors': errors})

    # ============================================================
    #  RETROACTIVE FIX — patch all existing Google events with
    #  correct location + link (presencial / virtual)
    # ============================================================
    @app.route('/calendar/api/fix-events', methods=['POST'])
    @login_required
    def api_fix_events():
        if not is_admin(): return jsonify({'success': False, 'error': 'Solo admin'})
        creds = get_google_creds(app)
        if not creds: return jsonify({'success': False, 'error': 'Google no conectado'})
        updated = 0; errors = 0; skipped = 0
        try:
            service  = build('calendar', 'v3', credentials=creds)
            all_cals = _get_calendar_config(app)
            email_map, gcal_id_map = _make_cal_maps(all_cals)
            apts = app.supabase.get('appointments',
                select='id,title,encargado,tema,client_name,start_time,end_time,calendar_id,'
                       'invitados,direccion,ciudad,lugar,mapa,notes,meeting_link,status,'
                       'google_event_id,google_cal_id')
            for apt in apts:
                if apt.get('status') != 'confirmed': continue
                gid = apt.get('google_event_id')
                if not gid: skipped += 1; continue
                try:
                    cal_id  = apt.get('calendar_id')
                    gcal_id = apt.get('google_cal_id') or gcal_id_map.get(cal_id, 'primary')
                    attendees = _build_attendees(apt, email_map)
                    ev = _build_google_event(apt, attendees)
                    patch = {'description': ev['description']}
                    if ev.get('location'): patch['location'] = ev['location']
                    service.events().patch(
                        calendarId=gcal_id, eventId=gid, body=patch).execute()
                    updated += 1
                except Exception:
                    errors += 1
            return jsonify({'success': True, 'updated': updated,
                            'skipped': skipped, 'errors': errors})
        except google.auth.exceptions.RefreshError:
            return jsonify({'success': False,
                'error': 'Google desconectado. Reconecta en /auth/google.'})
        except Exception as e:
            if _is_invalid_grant(e):
                return jsonify({'success': False,
                    'error': 'Google desconectado. Reconecta en /auth/google.'})
            return jsonify({'success': False, 'error': str(e)})

    # ============================================================
    #  APPOINTMENT UPDATE — re-sync Google Calendar on calendar change
    # ============================================================
    @app.route('/calendar/api/appointment/<aid>', methods=['PATCH'])
    @login_required
    def api_update_appointment(aid):
        """Update appointment fields.
        If calendar_id changes and the appointment is confirmed, deletes the
        old Google Calendar event and creates a new one in the correct calendar.
        """
        if not is_admin(): return jsonify({'success': False, 'error': 'Solo admin'})
        d = request.get_json() or {}
        if not d: return jsonify({'success': False, 'error': 'Sin datos'})

        apts = app.supabase.get('appointments', {'id': aid},
            select='id,calendar_id,google_event_id,google_cal_id,status,'
                   'title,encargado,tema,client_name,client_email,'
                   'start_time,end_time,invitados,lugar,direccion,ciudad,mapa,notes,meeting_link')
        if not apts: return jsonify({'success': False, 'error': 'No encontrado'})
        apt = apts[0]

        new_cal_id = d.get('calendar_id')
        old_cal_id = apt.get('calendar_id')
        cal_changed = new_cal_id and new_cal_id != old_cal_id
        is_confirmed = apt.get('status') == 'confirmed'

        sync_warning = None
        if cal_changed and apt.get('google_event_id') and is_confirmed:
            creds = get_google_creds(app)
            if creds:
                all_cals = _get_calendar_config(app)
                email_map, gcal_id_map = _make_cal_maps(all_cals)
                try:
                    service = build('calendar', 'v3', credentials=creds)
                    # Borrar del calendario anterior
                    old_gcal = apt.get('google_cal_id') or gcal_id_map.get(old_cal_id, 'primary')
                    try:
                        service.events().delete(
                            calendarId=old_gcal, eventId=apt['google_event_id']).execute()
                    except Exception:
                        pass
                    # Crear en el nuevo calendario
                    merged = {**apt, **d}
                    new_gcal = gcal_id_map.get(new_cal_id, 'primary')
                    attendees = _build_attendees(merged, email_map)
                    event = _build_google_event(merged, attendees)
                    created = service.events().insert(
                        calendarId=new_gcal, body=event, sendUpdates='all').execute()
                    d['google_event_id'] = created.get('id')
                    d['google_cal_id']   = new_gcal
                except Exception as e:
                    print(f'[api_update_appointment] Google error: {e}')
                    # El evento anterior ya pudo haberse borrado en Google: no dejar el
                    # google_event_id viejo apuntando a un evento que ya no existe.
                    d['google_event_id'] = None
                    d['google_cal_id']   = None
                    sync_warning = ('Se guardaron los cambios, pero falló la sincronización '
                                     'con Google Calendar. Usa "Reparar eventos" o vuelve a '
                                     'intentar el cambio de calendario.')
            else:
                sync_warning = ('Se guardó el cambio de calendario, pero Google no está '
                                 'conectado: el evento no se movió en Google Calendar.')

        ok = app.supabase.update('appointments', aid, d)
        return jsonify({'success': ok, 'warning': sync_warning})

    # ============================================================
    #  PLANNING MODULE
    # ============================================================
    @app.route('/planning')
    @login_required
    def planning():
        if not user_can('planning'):
            flash('No tienes acceso al módulo Planificación.', 'warning')
            return redirect('/dashboard')
        ms_connected = bool(get_ms_token(app))
        return render_template('planning.html', ms_connected=ms_connected,
                               is_admin_user=is_admin(), scope='planning',
                               page_title='Proyectos', page_sub='Tareas internas del equipo')

    @app.route('/todo')
    @login_required
    def todo():
        if not user_can('todo'):
            flash('No tienes acceso al módulo To-Do externo.', 'warning')
            return redirect('/dashboard')
        ms_connected = bool(get_ms_token(app))
        return render_template('planning.html', ms_connected=ms_connected,
                               is_admin_user=is_admin(), scope='todo',
                               page_title='Actividades',
                               page_sub='Pendientes sincronizados en ambos sentidos con Microsoft To-Do')

    @app.route('/planning/api/projects', methods=['GET'])
    @login_required
    def planning_projects():
        return jsonify(get_user_projects(app, current_user.id))

    @app.route('/planning/api/projects', methods=['POST'])
    @login_required
    def planning_create_project():
        if not user_can('planning'):
            return jsonify({'success': False, 'error': 'Sin acceso al módulo Planificación'})
        if not user_can('planning.proyectos'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para crear proyectos.'}), 403
        d = request.get_json() or {}
        d['created_by'] = current_user.id
        d['name'] = _sanitize(d.get('name', ''), 200)
        if not d['name']: return jsonify({'success': False, 'error': 'Nombre requerido'})
        if 'color' in d: d['color'] = _sanitize_hex_color(d.get('color'))
        r = app.supabase.insert('projects', d)
        if r and not is_admin():
            # Otorgar el proyecto recién creado al rol activo del creador: si no,
            # el control de acceso por proyecto le ocultaría su propio proyecto.
            rid = get_active_role_id(app, current_user.id)
            if rid:
                app.supabase.insert_ignore('role_projects', {'role_id': rid, 'project_id': r[0]['id']})
                _role_cache.invalidate(rid)
        return jsonify({'success': bool(r), 'project': r[0] if r else None})

    PROJECT_EDITABLE_FIELDS = {'name', 'description', 'color', 'status', 'priority',
                               'start_date', 'due_date', 'owner'}

    @app.route('/planning/api/projects/<pid>', methods=['PATCH'])
    @login_required
    def planning_update_project(pid):
        if not is_admin() and not (user_can('planning') and user_has_project_access(app, current_user.id, pid)):
            return jsonify({'success': False, 'error': 'Sin permisos sobre este proyecto'}), 403
        if not user_can('planning.proyectos'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para editar proyectos.'}), 403
        body = request.get_json() or {}
        d = {k: v for k, v in body.items() if k in PROJECT_EDITABLE_FIELDS}
        if 'name' in d:
            d['name'] = _sanitize(d.get('name', ''), 200)
            if not d['name']:
                return jsonify({'success': False, 'error': 'Nombre requerido'})
        if 'color' in d:
            d['color'] = _sanitize_hex_color(d.get('color'))
        if not d:
            return jsonify({'success': False, 'error': 'Nada que actualizar'})
        ok = app.supabase.update('projects', pid, d)
        return jsonify({'success': ok})

    @app.route('/planning/api/projects/<pid>', methods=['DELETE'])
    @login_required
    def planning_delete_project(pid):
        if not is_admin(): return jsonify({'success': False, 'error': 'Solo admin'})
        ok = app.supabase.delete('projects', pid)
        return jsonify({'success': ok})

    @app.route('/planning/api/tasks', methods=['GET'])
    @login_required
    def planning_tasks():
        pid    = request.args.get('project_id')
        scope  = request.args.get('scope', 'all')   # all | planning | todo
        if pid:
            rows = app.supabase.get('tasks', {'project_id': pid}, select='*')
        else:
            rows = app.supabase.get('tasks', select='*')
        rows = rows or []
        # Filtrar por scope (planning = manual; todo = MS)
        if scope == 'planning':
            rows = [t for t in rows if t.get('source') != 'ms_todo']
        elif scope == 'todo':
            rows = [t for t in rows if t.get('source') == 'ms_todo']
        # Permisos por usuario
        rows = _filter_visible_tasks(app, rows, current_user.id)
        return jsonify(rows)

    # ============================================================
    #  WEB PUSH — endpoints
    # ============================================================
    @app.route('/api/push/vapid-public', methods=['GET'])
    @login_required
    def push_vapid_public():
        if not WEB_PUSH_AVAILABLE:
            return jsonify({'available': False, 'error': 'pywebpush no instalado en el servidor'})
        keys = get_vapid_keys(app)
        if not keys:
            return jsonify({'available': False})
        return jsonify({'available': True, 'public_key': keys[2]})

    @app.route('/api/push/subscribe', methods=['POST'])
    @login_required
    def push_subscribe():
        body = request.get_json() or {}
        sub = body.get('subscription') or {}
        endpoint = sub.get('endpoint')
        keys = sub.get('keys') or {}
        if not endpoint or not keys.get('p256dh') or not keys.get('auth'):
            return jsonify({'success': False, 'error': 'subscription incompleta'})
        # Upsert por (user_id, endpoint)
        existing = app.supabase.get('web_push_subscriptions',
            {'user_id': current_user.id, 'endpoint': endpoint}, select='id')
        data = {
            'user_id':    current_user.id,
            'endpoint':   endpoint,
            'p256dh':     keys['p256dh'],
            'auth':       keys['auth'],
            'user_agent': (request.headers.get('User-Agent') or '')[:255],
        }
        if existing:
            app.supabase.update('web_push_subscriptions', existing[0]['id'], data)
        else:
            app.supabase.insert('web_push_subscriptions', data)
        return jsonify({'success': True})

    @app.route('/api/push/unsubscribe', methods=['POST'])
    @login_required
    def push_unsubscribe():
        body = request.get_json() or {}
        endpoint = body.get('endpoint', '')
        if endpoint:
            rows = app.supabase.get('web_push_subscriptions',
                {'user_id': current_user.id, 'endpoint': endpoint}, select='id')
            for r in (rows or []):
                app.supabase.delete('web_push_subscriptions', r['id'])
        return jsonify({'success': True})

    @app.route('/api/push/test', methods=['POST'])
    @login_required
    def push_test():
        n = send_push_to_user(app, current_user.id,
                              '🔔 Prueba de notificación',
                              'Si ves este mensaje, las notificaciones están funcionando.',
                              '/dashboard')
        return jsonify({'success': True, 'sent': n})

    @app.route('/api/push/notify-overdue', methods=['POST'])
    @login_required
    def push_notify_overdue():
        """Disparable por el cron externo: envía notificación a cada usuario con tareas vencidas hoy."""
        if not is_admin(): return jsonify({'success': False, 'error': 'Solo admin'})
        today_iso = date.today().isoformat()
        all_users = app.supabase.get('users', select='id,full_name,email,active_role_id') or []
        # Independiente del usuario: se consulta una sola vez fuera del bucle.
        rows = app.supabase.get('tasks',
            select='id,due_date,status,created_by,assigned_to,assigned_email,ms_email,source,project_id') or []
        sent_total = 0
        empty_grants = {'modules': [], 'calendar_ids': set(), 'project_ids': set(), 'ms_emails': set()}
        for u in all_users:
            uid = u['id']
            # Job en segundo plano: se usa el rol activo PERSISTIDO de cada usuario
            # (users.active_role_id), no session['active_role_id'] (eso es del admin
            # que disparó el cron, no del usuario que se está evaluando).
            grants = role_grants(app, u['active_role_id']) if u.get('active_role_id') else empty_grants
            mods = grants['modules']
            if 'planning' not in mods and 'todo' not in mods: continue
            has_todo = 'todo' in mods
            overdue = []
            for t in rows:
                if t.get('status') == 'done': continue
                d = t.get('due_date')
                if not d or d >= today_iso: continue
                pid = t.get('project_id')
                if pid and pid not in grants['project_ids']: continue
                if t.get('source') == 'ms_todo':
                    if has_todo and (t.get('ms_email') or '') in grants['ms_emails']: overdue.append(t)
                else:
                    if (t.get('created_by') == uid or t.get('assigned_to') == uid or
                        (t.get('assigned_email') or '').lower() == (u.get('email') or '').lower()):
                        overdue.append(t)
            if overdue:
                sent_total += send_push_to_user(app, uid,
                    f'⛔ Tienes {len(overdue)} tareas vencidas',
                    'Abre el sistema para revisar y reorganizar tus pendientes.',
                    '/todo?tf=overdue')
        return jsonify({'success': True, 'sent': sent_total})

    @app.route('/planning/api/ms-accounts', methods=['GET'])
    @login_required
    def planning_ms_accounts():
        if not is_admin(): return jsonify([])
        rows = app.supabase.get('ms_tokens', select='email') or []
        return jsonify([r.get('email') for r in rows if r.get('email')])

    @app.route('/planning/api/ms-lists', methods=['GET'])
    @login_required
    def planning_ms_lists():
        if not is_admin(): return jsonify([])
        email = request.args.get('email', '')
        if not email: return jsonify([])
        token = get_ms_token_for(app, email)
        if not token: return jsonify([])
        try:
            r = req_lib.get(f'{MS_GRAPH_URL}/me/todo/lists',
                            headers={'Authorization': f'Bearer {token}'}, timeout=(5,15))
            if r.status_code != 200: return jsonify([])
            lists = r.json().get('value', [])
            return jsonify([{'id': l['id'], 'name': l.get('displayName','To-Do')} for l in lists])
        except Exception:
            return jsonify([])

    @app.route('/planning/api/todo-target', methods=['GET', 'POST'])
    @login_required
    def planning_todo_target():
        """Lista de Microsoft To-Do a la que se envían las tareas creadas en el
        sistema. Es lo que hace que la sincronización sea de ida y vuelta sin que
        nadie tenga que elegir cuenta y lista en cada tarea."""
        if request.method == 'GET':
            return jsonify({'target': get_todo_default_target(app),
                            'puede_configurar': is_admin()})
        if not is_admin():
            return jsonify({'success': False, 'error': 'Sólo un administrador puede fijar la lista por defecto'}), 403
        cuerpo = request.get_json() or {}
        email   = _sanitize(cuerpo.get('email'), 200)
        list_id = _sanitize(cuerpo.get('list_id'), 300)
        if email and not app.supabase.get('ms_tokens', {'email': email}, select='id'):
            return jsonify({'success': False, 'error': 'Esa cuenta de Microsoft no está conectada'})
        ok = set_todo_default_target(app, email, list_id, _sanitize(cuerpo.get('list_name'), 200))
        return jsonify({'success': ok, 'target': get_todo_default_target(app)})

    @app.route('/planning/api/push-pending-to-todo', methods=['POST'])
    @login_required
    def planning_push_pending():
        """Sube a Microsoft To-Do las tareas del sistema que todavía no están allá.

        Fijar la lista por defecto sólo afecta a las tareas NUEVAS. Sin esto, todo
        lo creado en el sistema antes de configurarla se quedaba abajo para
        siempre, y la promesa de que ambas caras muestran lo mismo era falsa para
        el histórico. Esta ruta hace ese arrastre una vez."""
        if not user_can('todo.sincronizar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para sincronizar con Microsoft.'}), 403
        destino = get_todo_default_target(app)
        if not (destino and destino.get('email') and destino.get('list_id')):
            return jsonify({'success': False,
                            'error': 'Primero elige la lista de To-Do por defecto.'})

        cuerpo = request.get_json(silent=True) or {}
        incluir_completadas = bool(cuerpo.get('incluir_completadas'))
        pendientes = [t for t in (app.supabase.get('tasks', select='*') or [])
                      if not t.get('ms_email') and not t.get('source_id')
                      and (incluir_completadas or t.get('status') != 'done')]

        import time as _time
        limite = _time.monotonic() + 90       # mismo tope que el resto de lotes
        subidas, errores, parcial = 0, 0, False
        for tarea in pendientes:
            if _time.monotonic() > limite:
                parcial = True
                break
            tarea['ms_email']   = destino['email']
            tarea['ms_list_id'] = destino['list_id']
            ok, nuevo_id = push_task_to_ms(app, tarea)
            if ok and nuevo_id:
                app.supabase.update('tasks', tarea['id'], {
                    'ms_email':   destino['email'],
                    'ms_list_id': destino['list_id'],
                    'source_id':  nuevo_id,
                    'source':     'ms_todo',
                    'last_synced_at': datetime.now(timezone.utc).isoformat()})
                subidas += 1
            else:
                errores += 1
        return jsonify({'success': True, 'subidas': subidas, 'errores': errores,
                        'pendientes': len(pendientes), 'parcial': parcial})

    @app.route('/planning/api/deps/<dep_id>', methods=['DELETE'])
    @login_required
    def planning_delete_dep(dep_id):
        if not is_admin():
            rows = app.supabase.get('task_deps', {'id': dep_id}, select='task_id')
            if not rows:
                return jsonify({'success': False, 'error': 'No encontrada'}), 404
            task_rows = app.supabase.get('tasks', {'id': rows[0]['task_id']},
                select='id,created_by,assigned_to,assigned_email,ms_email,source,project_id')
            if not task_rows or not _user_owns_task(app, task_rows[0], current_user.id):
                return jsonify({'success': False, 'error': 'Sin permisos'}), 403
        ok = app.supabase.delete('task_deps', dep_id)
        return jsonify({'success': ok})

    @app.route('/planning/api/tasks/<tid>/refresh-subtasks', methods=['POST'])
    @login_required
    def planning_refresh_subtasks(tid):
        """Trae las subtareas más recientes desde Microsoft To-Do para una tarea concreta."""
        if not is_admin(): return jsonify({'success': False, 'error': 'Solo admin'})
        rows = app.supabase.get('tasks', {'id': tid}, select='*')
        if not rows: return jsonify({'success': False, 'error': 'Tarea no encontrada'})
        task = rows[0]
        ms_email = task.get('ms_email'); list_id = task.get('ms_list_id')
        src_id   = task.get('source_id')
        if not (ms_email and src_id):
            return jsonify({'success': False, 'error': 'Esta tarea no es de Microsoft To-Do'})
        token = get_ms_token_for(app, ms_email)
        if not token: return jsonify({'success': False, 'error': 'Token MS no disponible'})
        headers = {'Authorization': f'Bearer {token}'}
        # Si no sabemos en qué lista vive la tarea, la buscamos
        if not list_id:
            try:
                lr = req_lib.get(f'{MS_GRAPH_URL}/me/todo/lists', headers=headers, timeout=(5,15))
                if lr.status_code != 200:
                    return jsonify({'success': False, 'error': f'MS lists error {lr.status_code}'})
                for lst in lr.json().get('value', []):
                    cand = lst['id']
                    chk = req_lib.get(
                        f'{MS_GRAPH_URL}/me/todo/lists/{cand}/tasks/{src_id}',
                        headers=headers, timeout=(5,10))
                    if chk.status_code == 200:
                        list_id = cand
                        # Guardar para no buscar de nuevo
                        app.supabase.update('tasks', tid, {'ms_list_id': list_id})
                        break
                if not list_id:
                    return jsonify({'success': False, 'error': 'Tarea no encontrada en MS (¿borrada?)'})
            except Exception as e:
                return jsonify({'success': False, 'error': f'Buscando lista: {str(e)[:200]}'})
        items = []
        url = f'{MS_GRAPH_URL}/me/todo/lists/{list_id}/tasks/{src_id}/checklistItems?$top=200'
        try:
            while url:
                r = req_lib.get(url, headers=headers, timeout=(5,15))
                if r.status_code != 200:
                    return jsonify({'success': False, 'error': f'MS error {r.status_code}'})
                d = r.json()
                items.extend(d.get('value', []))
                url = d.get('@odata.nextLink')
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)[:200]})
        subs = [{
            'id':   ci.get('id',''),
            'name': (ci.get('displayName') or '').strip(),
            'done': bool(ci.get('isChecked')),
            'checked_at': ci.get('checkedDateTime'),
        } for ci in items]
        prog = int(sum(1 for s in subs if s.get('done')) * 100 / len(subs)) if subs else (task.get('progress_pct') or 0)
        app.supabase.update('tasks', tid, {
            'subtasks':       subs,
            'progress_pct':   prog,
            'last_synced_at': datetime.now(timezone.utc).isoformat()})
        return jsonify({'success': True, 'count': len(subs), 'progress': prog, 'subtasks': subs})

    @app.route('/planning/api/tasks/<tid>/subtask/<sid>', methods=['PATCH'])
    @login_required
    def planning_toggle_subtask(tid, sid):
        """Marca/desmarca una subtarea. body: {done: true|false}"""
        body = request.get_json() or {}
        done = bool(body.get('done'))
        rows = app.supabase.get('tasks', {'id': tid}, select='*')
        if not rows: return jsonify({'success': False, 'error': 'Tarea no encontrada'})
        task = rows[0]
        if not is_admin() and not _user_owns_task(app, task, current_user.id):
            return jsonify({'success': False, 'error': 'Sin permisos'}), 403
        subs = task.get('subtasks') or []
        changed = False
        for s in subs:
            if s.get('id') == sid:
                s['done'] = done
                s['checked_at'] = datetime.now(timezone.utc).isoformat() if done else None
                changed = True
                break
        if not changed: return jsonify({'success': False, 'error': 'Subtarea no encontrada'})
        # Recalcular progreso
        prog = int(sum(1 for s in subs if s.get('done')) * 100 / len(subs)) if subs else 0
        upd = {'subtasks': subs, 'progress_pct': prog,
               'updated_at': datetime.now(timezone.utc).isoformat()}
        ok = app.supabase.update('tasks', tid, upd)
        # Push a Microsoft si corresponde
        pushed = False
        if ok and task.get('source') == 'ms_todo' and task.get('ms_email') and task.get('ms_list_id') and task.get('source_id'):
            token = get_ms_token_for(app, task['ms_email'])
            if token:
                try:
                    r = req_lib.patch(
                        f"{MS_GRAPH_URL}/me/todo/lists/{task['ms_list_id']}/tasks/{task['source_id']}/checklistItems/{sid}",
                        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
                        json={'isChecked': done}, timeout=(5,15))
                    pushed = r.status_code in (200, 204)
                except Exception:
                    pass
        return jsonify({'success': ok, 'pushed_to_ms': pushed, 'progress': prog})

    @app.route('/planning/api/tasks', methods=['POST'])
    @login_required
    def planning_create_task():
        d = request.get_json() or {}
        wants_ms = bool(d.get('ms_email') and d.get('ms_list_id'))
        if not user_can('todo' if wants_ms else 'planning'):
            return jsonify({'success': False, 'error': 'Sin acceso a este módulo'})
        if wants_ms and not is_admin() and d['ms_email'] not in get_user_ms_emails(app, current_user.id):
            return jsonify({'success': False, 'error': 'No tienes autorización sobre esa cuenta de Microsoft'})
        if d.get('project_id') and not is_admin() and not user_has_project_access(app, current_user.id, d['project_id']):
            return jsonify({'success': False, 'error': 'No tienes acceso a ese proyecto'})
        d['created_by'] = current_user.id
        d['title'] = _sanitize(d.get('title', ''), 300)
        if not d['title']: return jsonify({'success': False, 'error': 'Título requerido'})
        d.setdefault('status', 'pending')
        d.setdefault('priority', 'medium')
        d.setdefault('phase', 'General')
        d.setdefault('progress_pct', 0)
        d.setdefault('alert_days', 3)
        # Sin destino explícito se usa la lista por defecto, de modo que toda
        # tarea creada aquí aparezca también en Microsoft To-Do. Es lo que cierra
        # el círculo de la sincronización: antes sólo bajaba, ahora también sube.
        # Sólo para quien tiene el módulo To-Do: si no lo tuviera, la tarea
        # pasaría a source='ms_todo' y dejaría de verla el propio autor.
        if not d.get('ms_email') and d.get('sync_todo', True) and user_can('todo'):
            destino = get_todo_default_target(app)
            if destino and destino.get('email') and destino.get('list_id'):
                permitidas = get_user_ms_emails(app, current_user.id)
                if is_admin() or destino['email'] in permitidas:
                    d['ms_email']   = destino['email']
                    d['ms_list_id'] = destino['list_id']
        d.pop('sync_todo', None)
        # Si se solicita sincronizar con MS, marcar como ms_todo
        if d.get('ms_email') and d.get('ms_list_id'):
            d.setdefault('source', 'ms_todo')
        r = app.supabase.insert('tasks', d)
        task = r[0] if r else None
        # Push a Microsoft si corresponde
        if task and task.get('ms_email') and task.get('ms_list_id'):
            ok, new_src = push_task_to_ms(app, task)
            if ok and new_src:
                app.supabase.update('tasks', task['id'],
                    {'source_id': new_src,
                     'last_synced_at': datetime.now(timezone.utc).isoformat()})
                task['source_id'] = new_src
        return jsonify({'success': bool(r), 'task': task})

    @app.route('/planning/api/tasks/<tid>', methods=['PATCH'])
    @login_required
    def planning_update_task(tid):
        if not is_admin():
            rows = app.supabase.get('tasks', {'id': tid},
                select='id,created_by,assigned_to,assigned_email,ms_email,source,project_id')
            task = rows[0] if rows else None
            if not task or not _user_owns_task(app, task, current_user.id):
                return jsonify({'success': False, 'error': 'Sin permisos'}), 403
        d = request.get_json() or {}
        # No basta con validar el estado PREVIO de la tarea: si el body intenta
        # reasignarla a un proyecto o cuenta MS fuera del rol activo, se rechaza.
        if not is_admin():
            if d.get('project_id') and not user_has_project_access(app, current_user.id, d['project_id']):
                return jsonify({'success': False, 'error': 'No tienes acceso a ese proyecto'}), 403
            if d.get('ms_email') and d['ms_email'] not in get_user_ms_emails(app, current_user.id):
                return jsonify({'success': False, 'error': 'No tienes autorización sobre esa cuenta de Microsoft'}), 403
        d['updated_at'] = datetime.now(timezone.utc).isoformat()
        if d.get('status') == 'done' and not d.get('completed_date'):
            d['completed_date'] = date.today().isoformat()
        ok = app.supabase.update('tasks', tid, d)
        pushed = False
        if ok:
            current = app.supabase.get('tasks', {'id': tid}, select='*')
            if current:
                pushed, new_src = push_task_to_ms(app, current[0])
                if pushed:
                    upd = {'last_synced_at': datetime.now(timezone.utc).isoformat()}
                    if new_src and not current[0].get('source_id'):
                        upd['source_id'] = new_src
                        upd['source']    = 'ms_todo'
                    app.supabase.update('tasks', tid, upd)
        return jsonify({'success': ok, 'pushed_to_ms': pushed})

    @app.route('/planning/api/tasks/bulk', methods=['POST'])
    @login_required
    def planning_bulk_update():
        """Body: {ids: [...], patch: {...}}. Aplica patch a varias tareas y empuja a MS si toca."""
        body = request.get_json() or {}
        ids = body.get('ids') or []
        patch = body.get('patch') or {}
        if not ids or not patch:
            return jsonify({'success': False, 'error': 'ids y patch son obligatorios'})
        patch['updated_at'] = datetime.now(timezone.utc).isoformat()
        if patch.get('status') == 'done' and not patch.get('completed_date'):
            patch['completed_date'] = date.today().isoformat()
            patch['progress_pct']   = 100
        # Permisos: si no es admin, valida que todas le pertenezcan (dueño/asignado/cuenta MS + proyecto del rol activo)
        if not is_admin():
            if patch.get('project_id') and not user_has_project_access(app, current_user.id, patch['project_id']):
                return jsonify({'success': False, 'error': 'No tienes acceso a ese proyecto'}), 403
            if patch.get('ms_email') and patch['ms_email'] not in get_user_ms_emails(app, current_user.id):
                return jsonify({'success': False, 'error': 'No tienes autorización sobre esa cuenta de Microsoft'}), 403
            rows = app.supabase.get_in('tasks', 'id', ids,
                select='id,created_by,assigned_to,assigned_email,ms_email,source,project_id')
            ids = [t['id'] for t in rows if _user_owns_task(app, t, current_user.id)]
            if not ids:
                return jsonify({'success': False, 'error': 'Sin permisos'})
        import time as _time
        DEADLINE = _time.monotonic() + 90
        updated = 0; pushed = 0; errors = 0; partial = False
        for tid in ids:
            if _time.monotonic() > DEADLINE: partial = True; break
            ok = app.supabase.update('tasks', tid, patch)
            if ok:
                updated += 1
                # Push a MS si la tarea es de To-Do
                try:
                    rows = app.supabase.get('tasks', {'id': tid}, select='*')
                    if rows:
                        pushed_ok, _ = push_task_to_ms(app, rows[0])
                        if pushed_ok: pushed += 1
                except Exception:
                    pass
            else:
                errors += 1
        return jsonify({'success': True, 'updated': updated, 'pushed_to_ms': pushed,
                        'errors': errors, 'partial': partial})

    @app.route('/planning/api/tasks/bulk-delete', methods=['POST'])
    @login_required
    def planning_bulk_delete():
        body = request.get_json() or {}
        ids = body.get('ids') or []
        if not ids: return jsonify({'success': False, 'error': 'ids vacío'})
        # Borrar de golpe es borrar, así que pide el mismo permiso que borrar de
        # una en una. Antes era «solo admin», y eso dejaba el interruptor
        # «Eliminar tareas» sin efecto por este camino.
        if not (is_admin() or user_can('todo.eliminar') or user_can('planning.eliminar')):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para eliminar tareas.'}), 403
        import time as _time
        DEADLINE = _time.monotonic() + 90
        deleted = 0; deleted_ms = 0; partial = False; omitidas = 0
        for tid in ids:
            if _time.monotonic() > DEADLINE: partial = True; break
            rows = app.supabase.get('tasks', {'id': tid}, select='*')
            task = rows[0] if rows else None
            # A diferencia del borrado de una sola tarea, aquí no había ninguna
            # comprobación por tarea: bastaba mandar los ids. Como ahora entra
            # gente que no es administrador, cada una se verifica —que sea suya
            # y que tenga el permiso del módulo al que pertenece— en vez de
            # confiar en la lista que llega del navegador.
            if task and not is_admin():
                ambito = 'todo' if task.get('ms_email') else 'planning'
                if not (_user_owns_task(app, task, current_user.id)
                        and user_can(f'{ambito}.eliminar')):
                    omitidas += 1
                    continue
            if app.supabase.delete('tasks', tid):
                deleted += 1
                if task and delete_task_in_ms(app, task):
                    deleted_ms += 1
        return jsonify({'success': True, 'deleted': deleted, 'deleted_in_ms': deleted_ms,
                        'omitidas': omitidas, 'partial': partial})

    @app.route('/planning/api/tasks/<tid>', methods=['DELETE'])
    @login_required
    def planning_delete_task(tid):
        # Obtener tarea antes de borrar para poder eliminarla en MS
        rows = app.supabase.get('tasks', {'id': tid}, select='*')
        task = rows[0] if rows else None
        if not task:
            return jsonify({'success': False, 'error': 'No encontrada'}), 404
        if not is_admin() and not _user_owns_task(app, task, current_user.id):
            return jsonify({'success': False, 'error': 'Sin permisos'}), 403
        # El permiso de borrado depende de dónde vive la tarea: si está atada a
        # una cuenta de Microsoft es del módulo Actividades, si no es de
        # Proyectos. Es la misma regla que usa el alta (planning_create_task).
        ambito = 'todo' if task.get('ms_email') else 'planning'
        if not user_can(f'{ambito}.eliminar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para eliminar tareas.'}), 403
        ok = app.supabase.delete('tasks', tid)
        deleted_ms = False
        if ok and task:
            deleted_ms = delete_task_in_ms(app, task)
        return jsonify({'success': ok, 'deleted_in_ms': deleted_ms})

    @app.route('/planning/api/deps/<tid>', methods=['GET'])
    @login_required
    def planning_task_deps(tid):
        if not is_admin():
            task_rows = app.supabase.get('tasks', {'id': tid},
                select='id,created_by,assigned_to,assigned_email,ms_email,source,project_id')
            if not task_rows or not _user_owns_task(app, task_rows[0], current_user.id):
                return jsonify({'success': False, 'error': 'Sin permisos'}), 403
        rows = app.supabase.get('task_deps', {'task_id': tid}, select='id,depends_on')
        return jsonify(rows or [])

    @app.route('/planning/api/deps', methods=['POST'])
    @login_required
    def planning_add_dep():
        d = request.get_json() or {}
        if not is_admin():
            for field in ('task_id', 'depends_on'):
                tid = d.get(field)
                task_rows = app.supabase.get('tasks', {'id': tid},
                    select='id,created_by,assigned_to,assigned_email,ms_email,source,project_id') if tid else []
                if not task_rows or not _user_owns_task(app, task_rows[0], current_user.id):
                    return jsonify({'success': False, 'error': 'Sin permisos sobre una de las tareas'}), 403
        r = app.supabase.insert_ignore('task_deps', d)
        return jsonify({'success': bool(r)})

    @app.route('/planning/api/import-todo', methods=['POST'])
    @login_required
    def planning_import_todo():
        # Antes era «sólo admin» y por eso el interruptor «Sincronizar con
        # Microsoft» de la pantalla de permisos no servía de nada: se podía
        # conceder y seguía sin dejar pulsar el botón. Ahora manda el permiso,
        # que para un administrador siempre es cierto.
        if not user_can('todo.sincronizar'):
            return jsonify({'success': False,
                            'error': 'No tienes permiso para sincronizar con Microsoft.'}), 403
        accounts = get_all_ms_tokens(app)
        # Filtro opcional: ?email=jomap@... para sincronizar una sola cuenta
        only_email = (request.args.get('email') or '').strip().lower()
        if only_email:
            accounts = [(e, t) for (e, t) in accounts if e.lower() == only_email]
        if not accounts:
            return jsonify({'success': False, 'needs_auth': True,
                'error': 'Microsoft To-Do no está conectado. Conecta primero desde Planificación.'})
        try:
            return jsonify(sync_ms_todo(app, accounts, current_user.id))
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)[:300]})

    @app.route('/planning/api/sync-todo-cron', methods=['POST', 'GET'])
    def planning_sync_todo_cron():
        """Sincronización automática To-Do ⇄ Sistema disparada por un cron externo.

        No requiere sesión: se autentica con un secreto (header X-Cron-Secret o
        ?secret=...). Trae a Microsoft los cambios hechos en el sistema ya se empujan
        en cada edición; aquí jalamos los cambios hechos directamente en To-Do, aunque
        nadie tenga la página abierta.
        """
        secret = app.config.get('CRON_SECRET') or ''
        given  = request.headers.get('X-Cron-Secret') or request.args.get('secret') or ''
        if not secret or given != secret:
            return jsonify({'success': False, 'error': 'No autorizado'}), 401
        accounts = get_all_ms_tokens(app)
        if not accounts:
            return jsonify({'success': False, 'error': 'Sin cuentas Microsoft conectadas'})
        # created_by para las tareas nuevas: el primer admin del sistema
        admins = app.supabase.get('users', {'role': 'admin'}, select='id') or []
        created_by = admins[0]['id'] if admins else None
        try:
            return jsonify(sync_ms_todo(app, accounts, created_by))
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)[:300]})

    @app.route('/planning/api/export-excel')
    @login_required
    def planning_export_excel():
        if not OPENPYXL_AVAILABLE:
            return jsonify({'error': 'openpyxl no instalado'}), 500
        pid = request.args.get('project_id')
        tasks = (app.supabase.get('tasks', {'project_id': pid}, select='*')
                 if pid else app.supabase.get('tasks', select='*'))
        tasks = _filter_visible_tasks(app, tasks or [], current_user.id)
        projects_map = {p['id']: p['name']
                        for p in (app.supabase.get('projects', select='id,name') or [])}
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = 'Tareas'
        hdr_fill = PatternFill('solid', fgColor='4F46E5')
        hdr_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='CCCCCC')
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ['Proyecto','Fase','Título','Descripción','Estado','Prioridad',
                   'Asignado a','Email','F. Inicio','F. Vencimiento',
                   'Días restantes','Progreso %','Alerta (días)','Etiquetas','Notas','Fuente']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = brd
        ws.row_dimensions[1].height = 20
        today_d = date.today()
        for ri, t in enumerate(tasks or [], 2):
            due_str = (t.get('due_date') or '')[:10]
            days_left = None
            if due_str:
                try:
                    days_left = (datetime.strptime(due_str,'%Y-%m-%d').date() - today_d).days
                except Exception: pass
            row = [
                _xlsx_safe(projects_map.get(t.get('project_id'), '')),
                _xlsx_safe(t.get('phase','')), _xlsx_safe(t.get('title','')), _xlsx_safe(t.get('description','')),
                t.get('status',''), t.get('priority',''),
                _xlsx_safe(t.get('assigned_to','')), _xlsx_safe(t.get('assigned_email','')),
                t.get('start_date','')[:10] if t.get('start_date') else '',
                due_str, days_left,
                t.get('progress_pct',0), t.get('alert_days',3),
                _xlsx_safe(t.get('tags','')), _xlsx_safe(t.get('notes','')), t.get('source',''),
            ]
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = brd
                if ci == 11 and val is not None:
                    c.font = Font(color='EF4444' if val < 0 else ('F59E0B' if val <= 3 else '000000'))
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 50)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f'tareas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/planning/api/import-excel', methods=['POST'])
    @login_required
    def planning_import_excel():
        if not user_can('planning'):
            return jsonify({'success': False, 'error': 'Sin acceso al módulo Planificación'})
        if not OPENPYXL_AVAILABLE:
            return jsonify({'success': False, 'error': 'openpyxl no instalado'}), 500
        f = request.files.get('file')
        if not f: return jsonify({'success': False, 'error': 'No se recibió archivo'})
        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            raw_headers = [str(c.value or '').strip().lower() for c in ws[1]]
            # Normalize headers — support Spanish and English
            alias = {
                'título': 'title', 'titulo': 'title',
                'descripción': 'description', 'descripcion': 'description',
                'fase': 'phase',
                'estado': 'status',
                'prioridad': 'priority',
                'asignado a': 'assigned_to',
                'f. inicio': 'start_date', 'fecha inicio': 'start_date',
                'f. vencimiento': 'due_date', 'fecha vencimiento': 'due_date',
                'progreso %': 'progress_pct', 'progreso': 'progress_pct',
                'alerta (días)': 'alert_days', 'alerta días': 'alert_days',
                'etiquetas': 'tags',
                'notas': 'notes',
                'email': 'assigned_email',
                'proyecto': '_project_name',
            }
            headers = [alias.get(h, h) for h in raw_headers]
            # Build project name map -- solo proyectos que el usuario puede ver (rol activo),
            # igual que planning_create_task/planning_update_task; si no, el import podría
            # asignar tareas a un proyecto fuera de su rol con solo adivinar el nombre.
            proj_by_name = {p['name'].lower(): p['id']
                            for p in get_user_projects(app, current_user.id)}
            imported = 0; errors = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None or str(v).strip() == '' for v in row): continue
                rd = dict(zip(headers, row))
                title = str(rd.get('title', '')).strip()
                if not title: continue
                try:
                    td = {
                        'title': title,
                        'description': str(rd.get('description', '') or ''),
                        'phase': str(rd.get('phase', 'General') or 'General').strip(),
                        'status': str(rd.get('status', 'pending') or 'pending').lower().replace(' ','_'),
                        'priority': str(rd.get('priority', 'medium') or 'medium').lower(),
                        'assigned_to': str(rd.get('assigned_to', '') or ''),
                        'assigned_email': str(rd.get('assigned_email', '') or ''),
                        'tags': str(rd.get('tags', '') or ''),
                        'notes': str(rd.get('notes', '') or ''),
                        'source': 'excel',
                        'created_by': current_user.id,
                    }
                    # Map project name → id
                    pname = str(rd.get('_project_name', '') or '').strip().lower()
                    if pname and pname in proj_by_name:
                        td['project_id'] = proj_by_name[pname]
                    # Parse dates
                    for fld in ('start_date', 'due_date'):
                        val = rd.get(fld)
                        if val:
                            if isinstance(val, datetime): td[fld] = val.strftime('%Y-%m-%d')
                            elif isinstance(val, date):   td[fld] = val.isoformat()
                            elif str(val).strip():
                                try: td[fld] = datetime.strptime(str(val).strip()[:10],'%Y-%m-%d').strftime('%Y-%m-%d')
                                except Exception: pass
                    # Progress
                    try: td['progress_pct'] = max(0, min(100, int(float(str(rd.get('progress_pct') or 0)))))
                    except Exception: td['progress_pct'] = 0
                    try: td['alert_days'] = max(0, int(float(str(rd.get('alert_days') or 3))))
                    except Exception: td['alert_days'] = 3
                    app.supabase.insert('tasks', td)
                    imported += 1
                except Exception: errors += 1
            return jsonify({'success': True, 'imported': imported, 'errors': errors})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    # ============================================================
    #  MÓDULOS EN ARCHIVO APARTE
    #  Directorio y Cronograma viven en app/directorio.py y app/cronograma.py.
    #  Reciben los ayudantes de este módulo por parámetro (no por import) para no
    #  crear un ciclo de importación entre los archivos.
    # ============================================================
    _ctx_modulos = {
        'is_admin':              is_admin,
        'user_can':              user_can,
        '_sanitize':             _sanitize,
        '_sanitize_hex_color':   _sanitize_hex_color,
        '_xlsx_safe':            _xlsx_safe,
        '_filter_visible_tasks': _filter_visible_tasks,
    }
    registrar_directorio(app, _ctx_modulos)
    registrar_cronograma(app, _ctx_modulos)

    # Contadores del menú. Se cachean por usuario: el navegador los pide en cada
    # pantalla y sin esto serían cuatro consultas por página vista.
    _badges_cache = TTLCache(ttl=45, maxsize=128)

    @app.route('/api/nav-badges')
    @login_required
    def nav_badges():
        """Cuánto trabajo hay detrás de cada módulo, para el menú lateral.

        Va por JavaScript y no en el context processor a propósito: así el
        pintado de la página no espera a estas consultas. Si algo falla se
        devuelve cero y el menú simplemente no muestra distintivo — un contador
        no puede tumbar una pantalla."""
        clave = f'{current_user.id}:{get_active_role_id(app, current_user.id)}'
        val, hit = _badges_cache.get(clave)
        if hit:
            return jsonify(val)

        hoy = date.today().isoformat()
        datos = {'actividades': 0, 'proyectos': 0, 'calendario': 0,
                 'cronograma': 0, 'directorio': 0, 'usuarios': 0}
        try:
            if user_can('todo') or user_can('planning'):
                tareas = _filter_visible_tasks(app, app.supabase.get(
                    'tasks', select='id,status,due_date,source,ms_email,created_by,'
                                    'assigned_to,assigned_email,project_id') or [],
                    current_user.id)
                vencidas = [t for t in tareas
                            if t.get('status') != 'done' and t.get('due_date')
                            and t['due_date'] < hoy]
                datos['actividades'] = sum(1 for t in vencidas if t.get('source') == 'ms_todo')
                datos['proyectos']   = sum(1 for t in vencidas if t.get('source') != 'ms_todo')
        except Exception:
            pass
        try:
            if is_admin():
                # Citas por aprobar y solicitudes de registro: son cosas que
                # esperan una decisión de un administrador.
                datos['calendario'] = len(app.supabase.get(
                    'appointments', {'status': 'pending'}, select='id') or [])
                datos['usuarios'] = len(app.supabase.get(
                    'calendar_permissions', {'status': 'pending'}, select='id') or [])
        except Exception:
            pass
        try:
            if user_can('cronograma'):
                planes = app.supabase.get('gantt_plans', select='id,created_by,status') or []
                if not is_admin():
                    planes = [p for p in planes if p.get('created_by') == str(current_user.id)]
                ids = {p['id'] for p in planes if p.get('status') != 'archived'}
                datos['cronograma'] = sum(
                    1 for a in (app.supabase.get('gantt_activities',
                                                 select='plan_id,status,end_date') or [])
                    if a.get('plan_id') in ids and a.get('status') != 'done'
                    and a.get('end_date') and a['end_date'] < hoy)
        except Exception:
            pass
        try:
            if user_can('directorio'):
                datos['directorio'] = len(app.supabase.get('contacts', select='id') or [])
        except Exception:
            pass

        _badges_cache.set(clave, datos)
        return jsonify(datos)

    # ============================================================
    #  UTILITY
    # ============================================================
    @app.route('/health')
    def health():
        return jsonify({'status': 'ok'})

    # ============================================================
    #  PUENTE CON ATLAS
    # ============================================================
    @app.route('/calendar/api/atlas-sync', methods=['POST'])
    @login_required
    def api_atlas_sync():
        """Sincroniza en los dos sentidos la agenda de ATLAS con la de aquí."""
        if not user_can('calendar'):
            return jsonify({'success': False, 'error': 'Sin acceso al Calendario'}), 403
        return jsonify(_atlas.sincronizar(app, TIMEZONE))

    @app.route('/calendar/api/atlas-estado', methods=['GET'])
    @login_required
    def api_atlas_estado():
        estado = _atlas.estado()
        if estado['disponible']:
            try:
                enlazadas = app.supabase.get(
                    'appointments', {'calendar_id': _atlas.CALENDARIO_ATLAS}, select='atlas_reunion_id') or []
                estado['enlazadas'] = sum(1 for c in enlazadas if c.get('atlas_reunion_id'))
                estado['total_calendario'] = len(enlazadas)
            except Exception:
                pass
        return jsonify(estado)

    @app.route('/calendar/api/atlas-sync-cron', methods=['POST', 'GET'])
    def api_atlas_sync_cron():
        """Misma pasada, disparada por un cron externo (sin sesión)."""
        secreto = app.config.get('CRON_SECRET') or ''
        recibido = request.headers.get('X-Cron-Secret') or request.args.get('secret') or ''
        if not secreto or recibido != secreto:
            return jsonify({'success': False, 'error': 'No autorizado'}), 401
        return jsonify(_atlas.sincronizar(app, TIMEZONE))

    @app.route('/calendar/api/google-reconnect', methods=['POST'])
    @login_required
    def api_google_reconnect():
        """Fuerza un ciclo de reconexión sin esperar al hilo de fondo.

        Si el permiso sigue vivo (el caso habitual tras un corte de red) esto
        arregla la conexión y sube las citas atrasadas sin que nadie tenga que
        volver a pasar por la pantalla de Google."""
        if not is_admin():
            return jsonify({'success': False, 'error': 'Solo admin'}), 403
        return jsonify(google_autoreconectar(app))

    @app.route('/calendar/api/google-keepalive', methods=['POST', 'GET'])
    def api_google_keepalive():
        """Mismo ciclo, disparado por un cron externo (sin sesión).

        Existe porque en despliegues donde el hilo de fondo no arranca —Windows,
        o un plan que apaga el proceso cuando no hay tráfico— el token se moriría
        igualmente por inactividad."""
        secreto = app.config.get('CRON_SECRET') or ''
        recibido = request.headers.get('X-Cron-Secret') or request.args.get('secret') or ''
        if not secreto or recibido != secreto:
            return jsonify({'success': False, 'error': 'No autorizado'}), 401
        return jsonify(google_autoreconectar(app))

    @app.route('/api/google-status')
    @login_required
    def google_status():
        if not is_admin():
            return jsonify({'connected': False, 'error': 'No autorizado'})
        tokens = app.supabase.get('google_tokens', {'email': GOOGLE_ACCOUNT_EMAIL})
        if not tokens:
            return jsonify({'connected': False, 'message': 'No hay token. Ve a /auth/google.'})
        t = tokens[0]
        creds = get_google_creds(app)
        if creds:
            return jsonify({'connected': True, 'email': t['email'],
                'expiry': t.get('token_expiry'), 'has_refresh_token': bool(t.get('refresh_token'))})
        return jsonify({'connected': False, 'email': t['email'],
            'message': 'Token invalido. Reconecta en /auth/google.'})

    # Trabajos de fondo (un solo worker se los queda mediante flock).
    if app.supabase:
        start_todo_autosync(app, interval_min=5)
        # Mantiene vivo el permiso de Google y sube las citas que quedaron
        # atrasadas en cuanto la conexión vuelve.
        start_google_autoheal(app, interval_min=60)
        # Puente con la agenda de ATLAS, en los dos sentidos.
        _atlas.arrancar_autosync(app, TIMEZONE, interval_min=10)

    return app
