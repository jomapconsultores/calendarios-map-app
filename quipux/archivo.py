# ------------------------------------------------------------
# Desarrollado por Marco Antonio Posligua San Martín
# ------------------------------------------------------------
"""Dejar los documentos en el disco con nombre propio y sitio propio.

Un archivo llamado `descarga(3).pdf` en la carpeta de descargas está perdido:
existe, ocupa espacio y no se puede encontrar. El trabajo de este módulo es que
eso no pase — que dentro de seis meses, buscando «matriz de requerimientos», el
Explorador de Windows lo encuentre sin tener que abrir nada.

Cómo queda:

    Quipux/
      GESTION DE PLANIFICACIÓN Y PROYECTOS/
        2026/
          Recibidos/
            2026-09-03_Oficio_DGPG-2050-2026_ACTUALIZACION-MATRIZ-DE-REQUERIMIENTOS/
              _FICHA.txt                     ← de quién, plazo, y el enlace
              01_matriz-requerimientos.xlsx
              02_oficio-firmado.pdf
      OBSERVATORIO DE SEGURIDAD CIUDADANA/
        ...
      INDICE.xlsx                            ← todo, con enlaces y plazos
      INDICE.html                            ← lo mismo, para abrir de un clic

La carpeta empieza por la FECHA para que el orden alfabético sea el
cronológico, sigue por el tipo y el número —que es como se cita un documento en
un oficio— y termina con el asunto recortado, que es lo que uno recuerda. Los
adjuntos van numerados en el orden en que aparecen, porque en un trámite el
orden importa: el primero suele ser el principal.

`_FICHA.txt` lleva el enlace a CuencaDOC. Es la pieza que pidió expresamente y
la más fácil de olvidar: sin ella, tener el archivo no sirve para contestarlo,
porque contestar exige volver al sistema y encontrar el documento otra vez.
"""
import json
import os
import re
import unicodedata
from datetime import datetime

# Windows no admite estos caracteres en un nombre de archivo, y los asuntos de
# los oficios vienen llenos de ellos: comillas, dos puntos, barras de fechas.
PROHIBIDOS = r'<>:"/\|?*'

# Nombres reservados de Windows: una carpeta llamada CON o PRN no se puede
# crear ni borrar por medios normales.
RESERVADOS = {'CON', 'PRN', 'AUX', 'NUL', 'COM1', 'COM2', 'COM3', 'COM4',
              'LPT1', 'LPT2', 'LPT3'}

# Windows se planta en 260 caracteres de ruta completa. Con la carpeta de
# Documentos, el área y el año ya gastados, al nombre le queda poco: se recorta
# aquí a conciencia en vez de que falle la escritura a mitad de la descarga.
LARGO_ASUNTO = 60
LARGO_NOMBRE = 70


def limpiar(texto, largo=LARGO_NOMBRE):
    """Un trozo de texto convertido en algo que Windows acepta como nombre."""
    t = unicodedata.normalize('NFKD', str(texto or ''))
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = ''.join('-' if c in PROHIBIDOS else c for c in t)
    t = re.sub(r'[\x00-\x1f]', '', t)
    t = re.sub(r'\s+', ' ', t).strip(' .-')
    t = t.replace(' ', '-')
    t = re.sub(r'-{2,}', '-', t)
    if t.upper().split('.')[0] in RESERVADOS:
        t = '_' + t
    return t[:largo].strip(' .-') or 'sin-nombre'


def nombre_de_carpeta(registro):
    """2026-09-03_Oficio_DGPG-2050-2026_ACTUALIZACION-MATRIZ..."""
    fecha = (registro.get('fecha_doc') or '')[:10] or 'sin-fecha'
    tipo = limpiar(registro.get('tipo') or 'Documento', 20)
    numero = limpiar(registro.get('numero') or registro.get('id') or 's-n', 40)
    asunto = limpiar(registro.get('asunto') or '', LARGO_ASUNTO)
    return f'{fecha}_{tipo}_{numero}_{asunto}'.strip('_')


def carpeta_del_documento(raiz, area, registro, bandeja):
    """Área / año / bandeja / documento."""
    anio = (registro.get('fecha_doc') or '')[:4]
    if not re.fullmatch(r'\d{4}', anio or ''):
        anio = 'sin-fecha'
    return os.path.join(raiz, limpiar(area, 60), anio,
                        limpiar(bandeja, 30), nombre_de_carpeta(registro))


def nombre_de_anexo(indice, nombre_original, url):
    """01_matriz-requerimientos.xlsx — numerado y con su extensión intacta."""
    base = nombre_original or url.rsplit('/', 1)[-1] or 'anexo'
    base = base.split('?')[0]
    raiz, ext = os.path.splitext(base)
    if not ext or len(ext) > 6:
        _, ext = os.path.splitext(url.split('?')[0])
    ext = (ext or '.bin').lower()
    return f'{indice:02d}_{limpiar(raiz, 50)}{ext}'


def enlace_al_documento(base, id_doc, numero):
    """El enlace que devuelve a la ficha en CuencaDOC.

    Quipux abre los documentos desde dentro de su marco, así que no hay una URL
    limpia que se pueda pegar en el navegador y funcione siempre. Se guarda la
    que más se acerca, y además el identificador y el número: con el buscador de
    la propia bandeja, el número encuentra el documento en dos segundos aunque
    la dirección haya dejado de valer."""
    return f'{base.rstrip("/")}/index_frames.php?doc={id_doc}&nro={numero}'


def escribir_ficha(carpeta, registro, plazo, base, texto=''):
    """El _FICHA.txt: para qué es este documento y dónde está el original."""
    fecha, origen_plazo, seguro = plazo
    lineas = [
        registro.get('asunto') or '(sin asunto)',
        '=' * min(len(registro.get('asunto') or '(sin asunto)'), 78),
        '',
        f"Número      : {registro.get('numero', '')}",
        f"Tipo        : {registro.get('tipo', '')}",
        f"Fecha       : {registro.get('fecha_doc', '')}",
        f"De          : {registro.get('de', '')}",
        f"Trámite     : {registro.get('tramite', '')}",
        f"Referencia  : {registro.get('referencia', '')}",
        f"Categoría   : {registro.get('categoria', '')}",
        '',
    ]
    if fecha:
        sello = 'según el sistema' if seguro else 'DEDUCIDO del texto — confírmalo'
        lineas += [f'PLAZO       : {fecha.isoformat()}  ({origen_plazo}; {sello})', '']
    else:
        lineas += ['PLAZO       : no consta', '']
    lineas += [
        f"Enlace      : {enlace_al_documento(base, registro.get('id',''), registro.get('numero',''))}",
        f"Buscar por  : {registro.get('numero', '')}   (pégalo en «Texto a Buscar» de la bandeja)",
        '',
        f"Descargado  : {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    if texto:
        lineas += ['', '-' * 60, 'TEXTO DEL DOCUMENTO', '-' * 60, '', texto[:8000]]
    ruta = os.path.join(carpeta, '_FICHA.txt')
    with open(ruta, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lineas))
    return ruta


# ============================================================
#  EL ÍNDICE
# ============================================================
CABECERAS = ['Área', 'Bandeja', 'Fecha', 'Tipo', 'Número', 'Asunto', 'De',
             'Trámite', 'Plazo', 'Origen del plazo', 'Confirmado', 'Estado',
             'Adjuntos', 'Carpeta', 'Enlace']


def _fila(d):
    plazo = d.get('plazo') or {}
    return [
        d.get('area', ''), d.get('bandeja', ''), d.get('fecha_doc', ''),
        d.get('tipo', ''), d.get('numero', ''), d.get('asunto', ''),
        d.get('de', ''), d.get('tramite', ''),
        plazo.get('fecha', ''), plazo.get('origen', ''),
        'sí' if plazo.get('seguro') else ('no' if plazo.get('fecha') else ''),
        d.get('estado', ''), d.get('n_adjuntos', 0),
        d.get('carpeta', ''), d.get('enlace', ''),
    ]


def escribir_indice_excel(ruta, documentos):
    """Un Excel que se puede ordenar y filtrar. La columna del plazo va primero
    en el orden de lectura de quien abre esto un lunes: qué se debe y cuándo."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Quipux'
    hoja.append(CABECERAS)
    cabecera = Font(bold=True, color='FFFFFF')
    fondo = PatternFill('solid', fgColor='1F4E79')
    for celda in hoja[1]:
        celda.font = cabecera
        celda.fill = fondo
        celda.alignment = Alignment(vertical='center', wrap_text=True)
    hoy = datetime.now().date().isoformat()
    vencido = PatternFill('solid', fgColor='FFC7CE')
    proximo = PatternFill('solid', fgColor='FFEB9C')
    for d in documentos:
        hoja.append(_fila(d))
        plazo = (d.get('plazo') or {}).get('fecha') or ''
        if plazo and d.get('estado') != 'cerrado':
            fila = hoja[hoja.max_row]
            if plazo < hoy:
                for c in fila: c.fill = vencido
            elif plazo <= (datetime.now().date().replace(
                    day=min(28, datetime.now().day)) ).isoformat():
                pass
        # El enlace, pinchable.
        celda = hoja.cell(row=hoja.max_row, column=len(CABECERAS))
        if celda.value:
            celda.hyperlink = celda.value
            celda.style = 'Hyperlink'
    anchos = [28, 16, 12, 14, 24, 60, 24, 16, 12, 26, 11, 10, 9, 50, 40]
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.freeze_panes = 'A2'
    hoja.auto_filter.ref = f'A1:{get_column_letter(len(CABECERAS))}{hoja.max_row}'
    libro.save(ruta)
    return ruta


def escribir_indice_html(ruta, documentos, base):
    """La misma información, pero de un clic: enlaces que abren la carpeta y el
    documento. El Excel es para trabajar; esto es para mirar."""
    import html as _h
    hoy = datetime.now().date().isoformat()

    def clase(d):
        p = (d.get('plazo') or {}).get('fecha') or ''
        if not p or d.get('estado') == 'cerrado':
            return ''
        return 'vencido' if p < hoy else 'pendiente'

    filas = []
    for d in sorted(documentos,
                    key=lambda x: ((x.get('plazo') or {}).get('fecha') or '9999',
                                   x.get('fecha_doc') or '')):
        plazo = d.get('plazo') or {}
        marca = '' if plazo.get('seguro') else ' <span class="ded">deducido</span>'
        carpeta = d.get('carpeta', '')
        filas.append(
            f'<tr class="{clase(d)}">'
            f'<td>{_h.escape(plazo.get("fecha") or "—")}{marca if plazo.get("fecha") else ""}</td>'
            f'<td>{_h.escape(d.get("fecha_doc", ""))}</td>'
            f'<td>{_h.escape(d.get("tipo", ""))}</td>'
            f'<td class="num">{_h.escape(d.get("numero", ""))}</td>'
            f'<td>{_h.escape(d.get("asunto", ""))}</td>'
            f'<td>{_h.escape(d.get("de", ""))}</td>'
            f'<td>{_h.escape(d.get("area", ""))}<br><small>{_h.escape(d.get("bandeja", ""))}</small></td>'
            f'<td>{d.get("n_adjuntos", 0)}</td>'
            f'<td><a href="file:///{_h.escape(carpeta.replace(chr(92), "/"))}">carpeta</a> · '
            f'<a href="{_h.escape(d.get("enlace", ""))}" target="_blank">Quipux</a></td>'
            f'</tr>')

    pendientes = [d for d in documentos if (d.get('plazo') or {}).get('fecha')]
    vencidos = [d for d in pendientes if (d['plazo']['fecha'] < hoy
                                          and d.get('estado') != 'cerrado')]
    doc = f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<title>Quipux — lo que hay que hacer</title>
<style>
 body{{font:14px system-ui,Segoe UI,sans-serif;margin:24px;color:#0f172a;background:#f8fafc}}
 h1{{font-size:20px;margin:0 0 4px}} .sub{{color:#64748b;margin:0 0 18px}}
 .tarjetas{{display:flex;gap:12px;margin-bottom:18px;flex-wrap:wrap}}
 .t{{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px 18px;min-width:120px}}
 .t b{{display:block;font-size:24px}} .t.rojo b{{color:#b91c1c}}
 table{{border-collapse:collapse;width:100%;background:#fff;font-size:13px}}
 th,td{{border-bottom:1px solid #e2e8f0;padding:7px 9px;text-align:left;vertical-align:top}}
 th{{background:#1f4e79;color:#fff;position:sticky;top:0}}
 tr.vencido td:first-child{{color:#b91c1c;font-weight:700}}
 tr.pendiente td:first-child{{color:#a16207;font-weight:600}}
 .num{{font-family:ui-monospace,Consolas,monospace;white-space:nowrap}}
 .ded{{background:#fef3c7;color:#92400e;font-size:10px;padding:1px 5px;border-radius:6px}}
 small{{color:#64748b}} a{{color:#1d4ed8}}
</style></head><body>
<h1>Quipux — lo que hay que hacer</h1>
<p class="sub">Marco Antonio Posligua San Martín · actualizado el {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
<div class="tarjetas">
  <div class="t"><b>{len(documentos)}</b>documentos</div>
  <div class="t"><b>{len(pendientes)}</b>con plazo</div>
  <div class="t rojo"><b>{len(vencidos)}</b>vencidos</div>
</div>
<table><thead><tr><th>Plazo</th><th>Fecha doc.</th><th>Tipo</th><th>Número</th>
<th>Asunto</th><th>De</th><th>Área / bandeja</th><th>Adj.</th><th>Abrir</th></tr></thead>
<tbody>{''.join(filas)}</tbody></table>
<p class="sub">Un plazo marcado como <span class="ded">deducido</span> se sacó leyendo el
texto del documento, no de un campo del sistema: confírmalo antes de fiarte.</p>
</body></html>"""
    with open(ruta, 'w', encoding='utf-8') as f:
        f.write(doc)
    return ruta


CAMPOS_INDICE = ('id', 'numero', 'asunto', 'de', 'tipo', 'fecha_doc', 'tramite',
                 'referencia', 'categoria', 'ultima', 'area', 'bandeja', 'estado',
                 'carpeta', 'enlace', 'n_adjuntos', 'plazo', 'nuevo')


def escribir_indice_json(ruta, documentos):
    """El mismo índice, en el formato que lee la plataforma.

    El Excel es para trabajar y el HTML para mirar; esto es para que la pantalla
    de calendarios·map pueda enseñar lo mismo sin tener que abrir un archivo de
    Office ni depender de que la base de datos esté en pie. Es un archivo suelto
    a propósito: lo que se recogió sigue estando aunque el servidor no conteste."""
    datos = {
        'actualizado': datetime.now().isoformat(timespec='seconds'),
        'total': len(documentos),
        'documentos': [{k: d.get(k) for k in CAMPOS_INDICE} for d in documentos],
    }
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(datos, f, ensure_ascii=False, indent=1)
    return ruta


def leer_indice_json(ruta):
    try:
        with open(ruta, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'actualizado': None, 'total': 0, 'documentos': []}


def escribir_estado(ruta, documentos):
    """Lo ya descargado, para que la próxima pasada no repita el trabajo ni
    vuelva a bajar cien adjuntos que ya están en el disco."""
    datos = {d['id']: {'numero': d.get('numero'), 'carpeta': d.get('carpeta'),
                       'ultima': d.get('ultima'), 'adjuntos': d.get('n_adjuntos', 0)}
             for d in documentos if d.get('id')}
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(datos, f, ensure_ascii=False, indent=1)
    return ruta


def leer_estado(ruta):
    try:
        with open(ruta, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}
